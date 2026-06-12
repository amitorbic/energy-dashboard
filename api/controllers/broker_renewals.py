import datetime
from io import BytesIO

import openpyxl
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

try:
    import phpserialize
    _HAS_PHP = True
except ImportError:
    _HAS_PHP = False

try:
    from dateutil.parser import parse as _dparse
    _HAS_DATEUTIL = True
except ImportError:
    _HAS_DATEUTIL = False


# ── Brokers list ───────────────────────────────────────────────────────────

async def get_brokers(db: AsyncSession) -> list:
    """
    Mirrors renewals.php SQL:
      SELECT broker_id, name, role FROM contract_user ORDER BY name ASC
    Excludes role==1 (admins) — mirrors PHP `if($row['role']=='1') continue`.
    """
    res = await db.execute(
        text("SELECT broker_id, name, role FROM contract_user ORDER BY name ASC")
    )
    return [
        {"broker_id": str(r.broker_id), "name": str(r.name)}
        for r in res.fetchall()
        if str(r.role) != "1"
    ]


# ── Change Company Name ────────────────────────────────────────────────────

async def change_company_name(db: AsyncSession, esiid: str, cname: str) -> dict:
    """
    Mirrors change_com_name.php Cname_update class:
      SELECT company_name FROM contract_renewal WHERE premise_id = :esiid
      UPDATE contract_renewal SET company_name = :cname WHERE premise_id = :esiid
    Returns PHP message "Updated Succesfully" (original typo preserved).
    """
    await db.execute(
        text("SELECT company_name FROM contract_renewal WHERE premise_id = :esiid"),
        {"esiid": esiid},
    )
    await db.execute(
        text("UPDATE contract_renewal SET company_name = :cname WHERE premise_id = :esiid"),
        {"cname": cname, "esiid": esiid},
    )
    await db.commit()
    return {"success": True, "message": "Updated Succesfully"}  # PHP typo preserved


# ── Renewal Offer Upload ───────────────────────────────────────────────────

def _parse_date(cell_val) -> datetime.date | None:
    """
    Mirrors renewal_offer_upload.php date handling:
      - datetime object from openpyxl → date
      - string with "/" → dateutil parse
      - Excel serial number → openpyxl datetime conversion
    """
    if isinstance(cell_val, datetime.datetime):
        return cell_val.date()
    if isinstance(cell_val, datetime.date):
        return cell_val
    if cell_val is None:
        return None
    s = str(cell_val).strip()
    if "/" in s and _HAS_DATEUTIL:
        try:
            return _dparse(s).date()
        except Exception:
            return None
    # Fall back: try Excel serial
    try:
        import openpyxl.utils.datetime as xl_dt
        return xl_dt.from_excel(float(s))
    except Exception:
        return None


async def upload_renewal_offer(db: AsyncSession, file_bytes: bytes) -> dict:
    """
    Mirrors renewal_offer_upload.php Excel processing.

    Expected column headers (row 1):
      A=Cust ID, B=Premise ID, C=Company Name, D=Load Profile,
      E=Service Address, I=Agent Code, J=Agent Name, L=Cust Email,
      M=Contract End Date

    Logic:
      1. Validate column headers (PHP lines 26-37).
      2. Build company → broker_code map (PHP $final_array).
      3. Truncate renewal_offer table.
      4. For each company, collect ESIDs and determine expiry status.
         "has expired" if date < today OR concatenated text contains "On Default".
         "will expire soon" otherwise.
      5. Insert using phpserialize.dumps() for esid_list
         (matches PHP serialize($esid_array)).
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    def cell(row: int, col: int) -> str:
        return str(ws.cell(row=row, column=col).value or "").strip()

    # ── Column header validation ─────────────────────────────────────────────
    errors = []
    if cell(1, 1) != "Cust ID":
        errors.append("Cust ID column is not mateched")       # PHP typo preserved
    if cell(1, 2) not in ("Premise ID", "premise_id"):
        errors.append("Premise ID column is not mateched")
    if cell(1, 3) != "Company Name":
        errors.append("Company Name column is not mateched")
    if cell(1, 4) not in ("Load Profile", "Load_Profile"):
        errors.append("Load Profile column is not mateched")
    if cell(1, 5) != "Service Address":
        errors.append("Service Address column is not mateched")
    if cell(1, 9) != "Agent Code":
        errors.append("Agent Code column is not mateched")
    if cell(1, 10) != "Agent Name":
        errors.append("Agent Name column is not mateched")
    if cell(1, 12) != "Cust Email":
        errors.append("Cust Email column is not mateched")
    if errors:
        return {"success": False, "message": ", ".join(errors), "records": []}

    max_row = ws.max_row

    # ── Build company → broker_code map (PHP $final_array) ───────────────────
    final_array: dict[str, str] = {}
    for i in range(2, max_row + 1):
        broker_name_val = cell(i, 10)
        if not broker_name_val:
            return {
                "success": False,
                "message": f"Broker Code Not Exist in Column J{i}",
                "records": [],
            }
        com = cell(i, 3)
        if com:
            final_array[com] = cell(i, 9)   # broker_code from col I

    # ── Truncate renewal_offer (PHP line 5: mysql_query("truncate table renewal_offer")) ─
    await db.execute(text("TRUNCATE TABLE renewal_offer"))

    today = datetime.date.today()
    records = []

    for com_name, broker_code in final_array.items():
        esid_array: list[str] = []
        date_expired = False
        text_parts: list[str] = []
        startdate_str = ""
        email = ""
        broker_name_col = ""

        for i in range(2, max_row + 1):
            if cell(i, 3) != com_name:
                continue

            esid_array.append(cell(i, 2))
            broker_name_col = cell(i, 10)
            email = cell(i, 12)

            raw_m = ws.cell(row=i, column=13).value
            d = _parse_date(raw_m)
            if d:
                startdate_str = d.strftime("%m/%d/%Y")
                if d < today:
                    date_expired = True
                else:
                    text_parts.append(startdate_str)
            else:
                text_parts.append(str(raw_m or ""))

        # PHP: $pos = strstr($string, "On Default"); if($pos) { $text1 = "has expired"; }
        combined = "".join(text_parts)
        if "On Default" in combined:
            date_expired = True

        status_text = "has expired" if date_expired else "will expire soon"

        # PHP: serialize($esid_array) — replicate with phpserialize
        if _HAS_PHP:
            esid_serial = phpserialize.dumps(
                {i: v.encode() for i, v in enumerate(esid_array)}
            ).decode("utf-8", errors="replace")
        else:
            # Fallback: simple comma-join if phpserialize not available
            esid_serial = ",".join(esid_array)

        await db.execute(
            text(
                "INSERT INTO renewal_offer "
                "(broker_code, broker_name, com_name, email, esid_list, start_date, text) "
                "VALUES (:bc, :bn, :cn, :em, :el, :sd, :tx)"
            ),
            {
                "bc":  broker_code,
                "bn":  broker_name_col,
                "cn":  com_name,
                "em":  email,
                "el":  esid_serial,
                "sd":  startdate_str,
                "tx":  status_text,
            },
        )
        records.append({
            "broker_code":  broker_code,
            "broker_name":  broker_name_col,
            "com_name":     com_name,
            "email":        email,
            "esid_list":    esid_array,
            "start_date":   startdate_str,
            "text":         status_text,
        })

    await db.commit()
    return {
        "success": True,
        "message": "Uploaded",
        "records": records,
    }


# ── Active Renewals ────────────────────────────────────────────────────────

async def get_active_renewals(
    db: AsyncSession,
    broker_id: str,
    search: str = "",
) -> dict:
    """
    Mirrors active_renewals.php / renewal_custom.php.

    PHP SQL (browse-all path):
      SELECT DISTINCT company_name, contract_end_date
      FROM contract_renewal
      WHERE broker_code = :broker_id
      [AND company_name LIKE '%:search%']

    Companies sorted ascending by earliest contract_end_date (PHP ksort on strtotime).
    For each company:
      SELECT * FROM contract_renewal
      WHERE company_name = :comp AND broker_code = :broker_id

    Returns:
      has_data: False if no companies found (mirrors PHP redirect to error=1)
      companies: list of {company_name, end_date, esiids: [...]}
    """
    if search:
        q = text(
            "SELECT DISTINCT company_name, contract_end_date "
            "FROM contract_renewal "
            "WHERE broker_code = :bid AND company_name LIKE :search"
        )
        res = await db.execute(q, {"bid": broker_id, "search": f"%{search}%"})
    else:
        q = text(
            "SELECT DISTINCT company_name, contract_end_date "
            "FROM contract_renewal "
            "WHERE broker_code = :bid"
        )
        res = await db.execute(q, {"bid": broker_id})

    rows = res.fetchall()
    if not rows:
        return {"has_data": False, "companies": []}

    # Build company → earliest end_date map, then ksort (ascending by date)
    import time as _time
    comp_date: dict[str, int] = {}
    for row in rows:
        cname = row[0] or ""
        end_str = row[1] or ""
        try:
            ts = int(_time.mktime(_time.strptime(end_str, "%m/%d/%Y")))
        except Exception:
            ts = 0
        # Keep earliest date per company
        if cname not in comp_date or ts < comp_date[cname]:
            comp_date[cname] = ts

    # Sort by earliest end date ascending (PHP ksort)
    sorted_companies = sorted(comp_date.items(), key=lambda x: x[1])

    companies = []
    for cname, end_ts in sorted_companies:
        end_display = (
            _time.strftime("%m/%d/%Y", _time.localtime(end_ts)) if end_ts else ""
        )

        esiid_res = await db.execute(
            text(
                "SELECT premise_id, premise_address2, premise_city, "
                "premise_state, premise_zip, contract_end_date "
                "FROM contract_renewal "
                "WHERE company_name = :comp AND broker_code = :bid"
            ),
            {"comp": cname, "bid": broker_id},
        )
        esiids = []
        for r in esiid_res.fetchall():
            pid = r[0] or ""
            if not pid or pid == "Array":
                continue
            city = r[2] or ""
            state = r[3] or ""
            zipcode = r[4] or ""
            csz = f"{city}/{state}/{zipcode}" if state else city
            end_date = r[5] or ""
            # PHP: if contract_end_date == "12/31/1969" → "Default"
            if end_date == "12/31/1969":
                end_date = "Default"
            esiids.append({
                "premise_id":    pid,
                "address":       r[1] or "",
                "city_state_zip": csz,
                "end_date":      end_date,
            })

        companies.append({
            "company_name": cname,
            "end_date":     end_display,
            "esiids":       esiids,
        })

    return {"has_data": True, "companies": companies}
