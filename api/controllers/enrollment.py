import os
import re
import time
import tempfile
from datetime import date, datetime, timedelta
from typing import List, Optional

import pandas as pd
import phpserialize
from fastapi import UploadFile
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession


# ── helpers ────────────────────────────────────────────────────────────────────


def _get_difference_months(start_date: str, end_date: str) -> int:
    """
    Exact replication of PHP _get_difference_months()
    from view_enrollments.php lines 249-262.
    Both dates are MM/DD/YYYY strings.
    """
    year  = int(end_date[6:10])  - int(start_date[6:10])
    month = int(end_date[0:2])   - int(start_date[0:2])
    days  = int(end_date[3:5])   - int(start_date[3:5])
    if days > 15:
        month += 1
    return (year * 12) + month


def _decode_volumes(volumes_str: str) -> List[str]:
    """
    Decode PHP-serialized volumes field from confirmation_log.
    PHP: $loade_profile = unserialize($row['volumes'])
         foreach ($loade_profile as $key => $value) { echo $key; }
    Returns list of load profile key strings.
    """
    if not volumes_str:
        return []
    try:
        decoded = phpserialize.loads(
            volumes_str.encode('latin-1'), decode_strings=True
        )
        if isinstance(decoded, dict):
            return list(decoded.keys())
        return []
    except Exception:
        return []


def _fmt_ts(ts_str: str) -> str:
    """
    Format Unix timestamp varchar to MM/DD/YYYY hh:mm:ss AM/PM.
    Mirrors PHP: date("m/d/Y h:i:s A", $row['date_modified'])
    """
    try:
        return datetime.fromtimestamp(int(ts_str)).strftime('%m/%d/%Y %I:%M:%S %p')
    except Exception:
        return ts_str


def _derive_zone(esid: str) -> str:
    """
    Exact replication of zone derivation in add_enrollment.php lines 236-247.
    """
    if esid[:10] == '1044372000':
        return 'north'
    if esid[:5] == '10089':
        return 'coast'
    if esid[:7] == '1003278':
        return 'south'
    return 'N/A'


def _derive_meter_fees(plan_type: str) -> str:
    """
    Exact replication of meter_fees switch in add_enrollment.php lines 103-200.
    """
    _map = {
        'PR1503060001': '0',
        'PR1503090001': '2.99',
        'PR1503060003': '5.00',
        'PR1503090002': '4.99',
        'PR1503050002': '5.99',
        'PR1503060002': '7.99',
        'PR1503050003': '10.00',
        'PR1503080002': '10.00',
        'PR1503080003': '5.99',
        'PR1503050005': '10.00',
        'PR1503050001': '5.99',
        'PR1503080001': '9.99',
    }
    return _map.get(plan_type, '')


def _c(row: "pd.Series", n: int) -> str:
    """
    Get cell at 1-indexed column n from a pandas row.
    Mirrors PHP: $data->sheets[0]['cells'][$i][N]
    """
    try:
        v = str(row.iloc[n - 1])
        return '' if v in ('nan', 'None', 'NaN') else v.strip()
    except IndexError:
        return ''


# ── upload ────────────────────────────────────────────────────────────────────


async def upload_enrollment(
    file: UploadFile, db: AsyncSession, username: str
) -> dict:
    """
    Exact replication of add_enrollment.php lines 22-356.

    PHP reads rows starting at $i=3 (rows 1-2 are title/header).
    pandas skiprows=2 aligns with PHP row 3 as first data row.
    All column numbers are 1-indexed, matching PHP.

    Column map:
      col 1      batch_no
      col 17     broker_code
      col 18     esid  (trim)
      col 23     company_name primary  (strip #')
      col 24+25  company_name fallback (first + last joined if col 23 blank)
      col 36     email_id
      col 109    contract_term
      col 127    contract_rate  (x100 on store)
      col 120    commission
      col 52-59  tax_exempt1-8
      col 34     zip
      col 108    contract_end_date
      col 125    plan_type -> meter_fees via switch
      col 20     cell3  (start date logic)
      col 21     cell1  (M/S flag)
      col 22     cell2  (fallback start date)

    NOTE: archive and comment_active columns are referenced in PHP line 255
    but are absent from the verified DB schema — excluded from all SQL.
    """
    suffix   = os.path.splitext(file.filename)[1].lower()
    tmp_path = os.path.join(tempfile.gettempdir(), file.filename)
    with open(tmp_path, 'wb') as f:
        f.write(await file.read())

    engine = 'xlrd' if suffix == '.xls' else 'openpyxl'
    df = pd.read_excel(tmp_path, header=None, dtype=str, engine=engine, skiprows=2)
    df = df.fillna('')

    success_esids: List[str] = []
    updated_esids: List[str] = []
    fail_esids:    List[str] = []
    total_esids = success = fail = 0

    for _, row in df.iterrows():
        esid = _c(row, 18).strip()
        if not esid:
            continue

        batch_no    = _c(row, 1)
        broker_code = _c(row, 17)

        raw_company = _c(row, 23)
        if raw_company:
            company_name = re.sub(r"[#']", '', raw_company)
        else:
            company_name = (_c(row, 24) + ' ' + _c(row, 25)).strip()

        email_id          = _c(row, 36)
        contract_term     = _c(row, 109)
        raw_rate          = _c(row, 127)
        contract_rate     = str(float(raw_rate) * 100) if raw_rate else '0'
        commission        = _c(row, 120)
        zip_code          = _c(row, 34)
        contract_end_date = _c(row, 108)
        plan_type         = _c(row, 125)
        meter_fees        = _derive_meter_fees(plan_type)
        zone              = _derive_zone(esid)

        tax_exempt1 = _c(row, 52)
        tax_exempt2 = _c(row, 53)
        tax_exempt3 = _c(row, 54)
        tax_exempt4 = _c(row, 55)
        tax_exempt5 = _c(row, 56)
        tax_exempt6 = _c(row, 57)
        tax_exempt7 = _c(row, 58)
        tax_exempt8 = _c(row, 59)

        # contract_start_date derivation — add_enrollment.php lines 204-221
        cell1 = _c(row, 21)
        cell2 = _c(row, 22)
        cell3 = _c(row, 20)
        contract_start_date = ''
        if cell1 == 'M':
            contract_start_date = cell3
        elif cell1 == 'S':
            if cell2 == '':
                contract_start_date = cell3
            elif cell2 != '' and cell3 != '':
                contract_start_date = cell2

        date_added = int(time.time())

        existing = (await db.execute(
            text("SELECT esid FROM enrollment "
                 "WHERE esid LIKE :esid AND type LIKE 'enrollment'"),
            {'esid': esid}
        )).fetchone()

        if existing:
            # UPDATE — reset all check flags to '' matching PHP line 255.
            # archive and comment_active omitted — absent from verified DB schema.
            await db.execute(text("""
                UPDATE enrollment SET
                    company_name = :company_name,
                    email_id = :email_id,
                    broker_code = :broker_code,
                    contract_rate = :contract_rate,
                    contract_term = :contract_term,
                    commission = :commission,
                    contract_start_date = :contract_start_date,
                    contract_end_date = :contract_end_date,
                    zone = :zone,
                    meter_fees = :meter_fees,
                    tax_exempt1 = :tax_exempt1,
                    tax_exempt2 = :tax_exempt2,
                    tax_exempt3 = :tax_exempt3,
                    tax_exempt4 = :tax_exempt4,
                    tax_exempt5 = :tax_exempt5,
                    tax_exempt6 = :tax_exempt6,
                    tax_exempt7 = :tax_exempt7,
                    tax_exempt8 = :tax_exempt8,
                    zip = :zip,
                    date_added = :date_added,
                    batch_no = :batch_no,
                    enrollment_status = '',
                    meter_fee_check = '',
                    enrolled_status = '',
                    billed_30 = '',
                    enrollment_cleared = '',
                    additional_esid_check = '',
                    comparison_comment = '',
                    cleared_comment = '',
                    contract_rate_check = '',
                    commission_check = '',
                    contract_term_check = '',
                    broker_code_check = '',
                    districts_cnp = '',
                    cnp_check = '',
                    compare_check = '',
                    load_profile_isigma = '',
                    enroll_check = '',
                    tax_exempt_check = '',
                    tax_exempt_comment = '',
                    certificate_check = ''
                WHERE esid LIKE :esid
            """), {
                'company_name': company_name,  'email_id': email_id,
                'broker_code': broker_code,    'contract_rate': contract_rate,
                'contract_term': contract_term,'commission': commission,
                'contract_start_date': contract_start_date,
                'contract_end_date': contract_end_date,
                'zone': zone,                  'meter_fees': meter_fees,
                'tax_exempt1': tax_exempt1,    'tax_exempt2': tax_exempt2,
                'tax_exempt3': tax_exempt3,    'tax_exempt4': tax_exempt4,
                'tax_exempt5': tax_exempt5,    'tax_exempt6': tax_exempt6,
                'tax_exempt7': tax_exempt7,    'tax_exempt8': tax_exempt8,
                'zip': zip_code,               'date_added': date_added,
                'batch_no': batch_no,          'esid': esid,
            })
            updated_esids.append(esid)
            total_esids += 1
        else:
            try:
                await db.execute(text("""
                    INSERT INTO enrollment
                        (esid, batch_no, company_name, broker_code, email_id,
                         zone, contract_rate, contract_term, commission,
                         contract_start_date, contract_end_date, meter_fees,
                         tax_exempt1, tax_exempt2, tax_exempt3, tax_exempt4,
                         tax_exempt5, tax_exempt6, tax_exempt7, tax_exempt8,
                         zip, date_added, type, status)
                    VALUES
                        (:esid, :batch_no, :company_name, :broker_code, :email_id,
                         :zone, :contract_rate, :contract_term, :commission,
                         :contract_start_date, :contract_end_date, :meter_fees,
                         :tax_exempt1, :tax_exempt2, :tax_exempt3, :tax_exempt4,
                         :tax_exempt5, :tax_exempt6, :tax_exempt7, :tax_exempt8,
                         :zip, :date_added, 'enrollment', 1)
                """), {
                    'esid': esid.strip(),          'batch_no': batch_no,
                    'company_name': company_name,  'broker_code': broker_code.strip(),
                    'email_id': email_id,          'zone': zone,
                    'contract_rate': contract_rate,'contract_term': contract_term,
                    'commission': commission,
                    'contract_start_date': contract_start_date,
                    'contract_end_date': contract_end_date,
                    'meter_fees': meter_fees,
                    'tax_exempt1': tax_exempt1,    'tax_exempt2': tax_exempt2,
                    'tax_exempt3': tax_exempt3,    'tax_exempt4': tax_exempt4,
                    'tax_exempt5': tax_exempt5,    'tax_exempt6': tax_exempt6,
                    'tax_exempt7': tax_exempt7,    'tax_exempt8': tax_exempt8,
                    'zip': zip_code,               'date_added': date_added,
                })
                success_esids.append(esid)
                success = 1
                total_esids += 1
            except Exception:
                fail_esids.append(esid)
                fail = 1

    # enrollment_log INSERT — add_enrollment.php line 347
    # esid column omitted from INSERT: upload-level log has no per-ESID reference
    await db.execute(text("""
        INSERT INTO enrollment_log (user, date_modified, num_esid, comments)
        VALUES (:user, :date_modified, :num_esid, :comments)
    """), {
        'user':          username,
        'date_modified': str(int(time.time())),
        'num_esid':      total_esids,
        'comments':      f'Uploaded enrollment sheet with {total_esids} esids',
    })

    await db.commit()
    os.remove(tmp_path)

    return {
        'success':       success,
        'fail':          fail,
        'total_esids':   total_esids,
        'success_esids': success_esids,
        'updated_esids': updated_esids,
        'fail_esids':    fail_esids,
    }


# ── view enrollments ──────────────────────────────────────────────────────────


async def get_enrollments_view(sort: Optional[str], db: AsyncSession) -> List[dict]:
    """
    Replication of view_enrollments.php lines 262-380.

    sort=None (default) : NOT completed/cleared/deleted, ORDER date_added DESC
    sort='date'         : same without completed filter, ORDER date_added DESC
    sort='comment'      : all non-cleared, term-mismatch rows first (Python sort)
    sort='status'       : with-status rows first, empty-status rows second
    """
    if sort == 'date':
        rows = (await db.execute(text("""
            SELECT * FROM enrollment
            WHERE type = 'enrollment'
              AND enroll_check <> 1
              AND enrollment_cleared <> 1
              AND enrollment_status NOT LIKE 'Delete'
            ORDER BY date_added DESC
        """))).fetchall()
        return [dict(r._mapping) for r in rows]

    if sort == 'status':
        with_status = (await db.execute(text("""
            SELECT * FROM enrollment
            WHERE type = 'enrollment'
              AND enroll_check <> 1
              AND enrollment_status <> ''
              AND enrollment_cleared <> 1
              AND enrollment_status NOT LIKE 'Delete'
            ORDER BY enrollment_status, enrollment_status_comment, date_added DESC
        """))).fetchall()
        without_status = (await db.execute(text("""
            SELECT * FROM enrollment
            WHERE type = 'enrollment'
              AND enroll_check <> 1
              AND enrollment_status = ''
              AND enrollment_cleared <> 1
              AND enrollment_status NOT LIKE 'Delete'
            ORDER BY enrollment_status_comment DESC, date_added DESC
        """))).fetchall()
        return [dict(r._mapping) for r in with_status] + \
               [dict(r._mapping) for r in without_status]

    if sort == 'comment':
        rows = (await db.execute(text("""
            SELECT * FROM enrollment
            WHERE type = 'enrollment'
              AND enroll_check <> 1
              AND enrollment_cleared <> 1
              AND enrollment_status NOT LIKE 'Delete'
            ORDER BY date_added DESC
        """))).fetchall()
        records = [dict(r._mapping) for r in rows]
        mismatch, match = [], []
        for r in records:
            try:
                diff = _get_difference_months(
                    r['contract_start_date'], r['contract_end_date']
                )
                (mismatch if diff != int(r['contract_term'] or 0) else match).append(r)
            except Exception:
                match.append(r)
        return mismatch + match

    # Default
    rows = (await db.execute(text("""
        SELECT * FROM enrollment
        WHERE type = 'enrollment'
          AND enroll_check <> 1
          AND enrollment_cleared <> 1
          AND enrollment_status NOT LIKE '%Completed%'
          AND enrollment_status NOT LIKE 'Delete'
        ORDER BY date_added DESC
    """))).fetchall()
    return [dict(r._mapping) for r in rows]


# ── completed enrollments ─────────────────────────────────────────────────────


async def get_completed_enrollments(sort: Optional[str], db: AsyncSession) -> List[dict]:
    """
    Replication of completed_enrollments.php lines 244-381.

    Default : enrollment_status LIKE '%Completed%', last 8 months,
              green (Completed) rows first then grey.
    sort='date'    : all non-cleared, green first
    sort='comment' : all non-cleared, mismatch first (Python sort)
    sort='status'  : with-status first then empty-status
    """
    eight_months_ago = int((datetime.now() - timedelta(days=240)).timestamp())

    if sort == 'date':
        rows = (await db.execute(text("""
            SELECT * FROM enrollment
            WHERE type = 'enrollment'
              AND enroll_check <> 1
              AND enrollment_cleared <> 1
            ORDER BY date_added DESC
        """))).fetchall()
        records = [dict(r._mapping) for r in rows]
        green = [r for r in records if r['enrollment_status'][:9] == 'Completed']
        grey  = [r for r in records if r['enrollment_status'][:9] != 'Completed']
        return green + grey

    if sort == 'status':
        with_s = (await db.execute(text("""
            SELECT * FROM enrollment
            WHERE type = 'enrollment'
              AND enroll_check <> 1
              AND enrollment_status <> ''
              AND enrollment_cleared <> 1
            ORDER BY enrollment_status, enrollment_status_comment
        """))).fetchall()
        without_s = (await db.execute(text("""
            SELECT * FROM enrollment
            WHERE type = 'enrollment'
              AND enroll_check <> 1
              AND enrollment_status = ''
              AND enrollment_cleared <> 1
            ORDER BY enrollment_status_comment DESC
        """))).fetchall()
        return [dict(r._mapping) for r in with_s] + \
               [dict(r._mapping) for r in without_s]

    if sort == 'comment':
        rows = (await db.execute(text("""
            SELECT * FROM enrollment
            WHERE type = 'enrollment'
              AND enroll_check <> 1
              AND enrollment_cleared <> 1
            ORDER BY date_added DESC
        """))).fetchall()
        records = [dict(r._mapping) for r in rows]
        mismatch, match = [], []
        for r in records:
            try:
                diff = _get_difference_months(
                    r['contract_start_date'], r['contract_end_date']
                )
                (mismatch if diff != int(r['contract_term'] or 0) else match).append(r)
            except Exception:
                match.append(r)
        return mismatch + match

    # Default
    rows = (await db.execute(text("""
        SELECT * FROM enrollment
        WHERE date_added >= :eight_months_ago
          AND type = 'enrollment'
          AND enroll_check <> 1
          AND enrollment_cleared <> 1
          AND enrollment_status LIKE '%Completed%'
    """), {'eight_months_ago': eight_months_ago})).fetchall()
    records = [dict(r._mapping) for r in rows]
    green = [r for r in records if r['enrollment_status'][:9] == 'Completed']
    grey  = [r for r in records if r['enrollment_status'][:9] != 'Completed']
    return green + grey


# ── canceled enrollments ──────────────────────────────────────────────────────


async def get_canceled_enrollments(db: AsyncSession) -> List[dict]:
    """Replication of canceled_enrollment.php line 46."""
    rows = (await db.execute(text(
        "SELECT * FROM enrollment WHERE enrollment_status = 'Cancelled'"
    ))).fetchall()
    return [dict(r._mapping) for r in rows]


# ── user log ──────────────────────────────────────────────────────────────────


async def get_user_log(db: AsyncSession) -> List[dict]:
    """
    Replication of enrollment_user_log.php lines 21-23.

    PHP: date_modified > date('Y-m-d', strtotime("-1 month"))
    MySQL evaluates the unquoted date arithmetic (e.g. 2025-5-10 = 2010) as
    a numeric threshold. All Unix timestamps (~1.7B) exceed this value so
    all records are returned. Replicated exactly via CAST comparison.
    """
    d         = date.today() - timedelta(days=30)
    threshold = d.year - d.month - d.day   # e.g. 2025 - 5 - 10 = 2010

    rows = (await db.execute(text("""
        SELECT * FROM enrollment_log
        WHERE CAST(date_modified AS SIGNED) > :threshold
        ORDER BY CAST(date_modified AS SIGNED) DESC, user DESC
    """), {'threshold': threshold})).fetchall()
    return [dict(r._mapping) for r in rows]


# ── dashboard stats ───────────────────────────────────────────────────────────


async def get_stats(db: AsyncSession) -> dict:
    """Replication of report_home.php lines 19-28 — five COUNT queries."""
    start = '1341100800'   # PHP hardcoded: July 1, 2012 (Unix ts stored as varchar)

    total_confirmations = (await db.execute(text("""
        SELECT COUNT(*) FROM confirmation_log
        WHERE date_modified >= :start
          AND (type_of_contract LIKE 'New' OR type_of_contract LIKE 'Addition')
    """), {'start': start})).scalar()

    total_enrollments = (await db.execute(text(
        "SELECT COUNT(*) FROM enrollment WHERE type LIKE 'enrollment'"
    ))).scalar()

    enrollments_checked = (await db.execute(text(
        "SELECT COUNT(*) FROM enrollment "
        "WHERE enroll_check = 1 AND type LIKE 'enrollment'"
    ))).scalar()

    enrollments_unchecked = (await db.execute(text(
        "SELECT COUNT(*) FROM enrollment "
        "WHERE enroll_check = 0 AND type LIKE 'enrollment'"
    ))).scalar()

    confirmations_unchecked = (await db.execute(text("""
        SELECT COUNT(*) FROM confirmation_log
        WHERE enroll_check = 0
          AND (type_of_contract LIKE 'New' OR type_of_contract LIKE 'Addition')
          AND date_modified >= :start
    """), {'start': start})).scalar()

    return {
        'total_confirmations':     int(total_confirmations    or 0),
        'total_enrollments':       int(total_enrollments      or 0),
        'enrollments_checked':     int(enrollments_checked    or 0),
        'enrollments_unchecked':   int(enrollments_unchecked  or 0),
        'confirmations_unchecked': int(confirmations_unchecked or 0),
    }


# ── enrollment list ───────────────────────────────────────────────────────────


async def get_enrollment_list(db: AsyncSession) -> List[dict]:
    """Replication of view_enrollment_list.php line 66."""
    rows = (await db.execute(text(
        "SELECT * FROM enrollment WHERE enroll_check = 0 AND enrollment_cleared = 0"
    ))).fetchall()
    return [dict(r._mapping) for r in rows]


# ── template CRUD ─────────────────────────────────────────────────────────────


async def get_templates(search: Optional[str], db: AsyncSession) -> List[dict]:
    """Replication of template_list.php lines 22-67. Joins broker_new for company_name."""
    if search:
        rows = (await db.execute(text("""
            SELECT * FROM additional_esid_template
            WHERE customer_name LIKE :search ORDER BY sid DESC
        """), {'search': f'%{search}%'})).fetchall()
    else:
        rows = (await db.execute(text(
            "SELECT * FROM additional_esid_template ORDER BY sid DESC"
        ))).fetchall()

    results = []
    for r in rows:
        row = dict(r._mapping)
        bn = (await db.execute(
            text("SELECT company_name FROM broker_new WHERE broker_code LIKE :bc"),
            {'bc': row['broker_code']}
        )).fetchone()
        row['broker_name'] = bn[0] if bn else ''
        results.append(row)
    return results


async def get_brokers_for_dropdown(db: AsyncSession) -> List[dict]:
    """Replication of add_template.php lines 257-270."""
    rows = (await db.execute(text("""
        SELECT broker_code, company_name FROM broker_new
        WHERE confirmation_email <> '' AND confirmation_flag = 1
        ORDER BY company_name ASC
    """))).fetchall()
    return [dict(r._mapping) for r in rows]


async def create_template(payload, db: AsyncSession) -> dict:
    """Replication of add_template.php lines 128-133 (INSERT branch)."""
    await db.execute(text("""
        INSERT INTO additional_esid_template
            (customer_name, contract_rate, commission, broker_code,
             meter_fee, tax_exempt, contract_end_date)
        VALUES (:customer_name, :contract_rate, :commission, :broker_code,
                :meter_fee, :tax_exempt, :contract_end_date)
    """), {
        'customer_name': payload.customer_name, 'contract_rate': payload.contract_rate,
        'commission': payload.commission,        'broker_code': payload.broker_code,
        'meter_fee': payload.meter_fee,          'tax_exempt': payload.tax_exempt,
        'contract_end_date': payload.contract_end_date,
    })
    await db.commit()
    return {'message': 'Template added successfully'}


async def update_template(sid: int, payload, db: AsyncSession) -> dict:
    """Replication of add_template.php lines 122-126 (UPDATE branch)."""
    await db.execute(text("""
        UPDATE additional_esid_template SET
            customer_name = :customer_name, contract_rate = :contract_rate,
            commission = :commission,        broker_code = :broker_code,
            meter_fee = :meter_fee,          tax_exempt = :tax_exempt,
            contract_end_date = :contract_end_date
        WHERE sid = :sid
    """), {
        'customer_name': payload.customer_name, 'contract_rate': payload.contract_rate,
        'commission': payload.commission,        'broker_code': payload.broker_code,
        'meter_fee': payload.meter_fee,          'tax_exempt': payload.tax_exempt,
        'contract_end_date': payload.contract_end_date, 'sid': sid,
    })
    await db.commit()
    return {'message': 'Template updated successfully'}


async def delete_template(sid: int, db: AsyncSession) -> dict:
    """Replication of delete_esid_template.php lines 8-10."""
    existing = (await db.execute(
        text("SELECT sid FROM additional_esid_template WHERE sid = :sid"), {'sid': sid}
    )).fetchone()
    if not existing:
        return {'result': 0}
    await db.execute(
        text("DELETE FROM additional_esid_template WHERE sid = :sid"), {'sid': sid}
    )
    await db.commit()
    return {'result': 1}


# ── download completed (download_compleat_enrrol.php) ────────────────────────


async def download_completed(start: str, end: str, db: AsyncSession) -> bytes:
    """
    Replication of download_compleat_enrrol.php lines 243-301.
    JOIN enrollment_log + enrollment, date range filter, 14-column XLSX.
    comment split by '<br />' joined with commas (line 289).
    """
    import openpyxl
    from io import BytesIO
    from openpyxl.styles import PatternFill, Font

    start_ts = int(datetime.strptime(start, '%Y-%m-%d').timestamp())
    end_ts   = int(datetime.strptime(end,   '%Y-%m-%d').timestamp())

    rows = (await db.execute(text("""
        SELECT e.esid, e.company_name, e.broker_code, e.contract_rate,
               e.contract_term, e.commission, e.contract_start_date,
               e.contract_end_date, e.meter_fees, e.enrolled_status,
               e.enrollment_status, e.enrollment_status_comment, el.date_modified
        FROM enrollment_log AS el
        LEFT JOIN enrollment AS e ON el.esid = e.esid
        WHERE el.date_modified >= :s AND el.date_modified <= :e
          AND e.enrollment_status LIKE '%Completed%'
          AND el.comments LIKE '%Changed enrollment status from - to Completed-%'
          AND e.enroll_check <> 1 AND e.enrollment_cleared <> 1
        ORDER BY el.date_modified DESC
    """), {'s': str(start_ts), 'e': str(end_ts)})).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Completed Enrollments'
    hfill = PatternFill('solid', fgColor='FFFF00')
    hfont = Font(color='FF0000')
    for col, h in enumerate([
        'S No.', 'Date added', 'Company name', 'Broker code', 'Premise id',
        'Contract rate', 'Comm', 'Term', 'Contract start date',
        'Contract end date', 'Meter fees', 'Enrollment Status',
        'Broker Name', 'Enrollment Comment',
    ], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont

    for sno, r in enumerate(rows, 1):
        row = dict(r._mapping)
        bn = (await db.execute(
            text("SELECT company_name FROM broker_new WHERE broker_code = :bc"),
            {'bc': row.get('broker_code') or ''}
        )).fetchone()
        ws.append([
            sno,
            _fmt_ts(str(row.get('date_modified') or '')),
            row.get('company_name') or '',
            row.get('broker_code') or '',
            row.get('esid') or '',
            str(float(row.get('contract_rate') or 0) / 100),
            row.get('commission') or '',
            row.get('contract_term') or '',
            row.get('contract_start_date') or '',
            row.get('contract_end_date') or '',
            row.get('meter_fees') or '',
            row.get('enrollment_status') or '',
            bn[0] if bn else '',
            ','.join((row.get('enrollment_status_comment') or '').split('<br />')),
        ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── download pending (download_pending.php) ───────────────────────────────────


async def download_pending(db: AsyncSession) -> bytes:
    """
    Replication of download_pending.php lines 39-92.
    Same exclusion logic as get_pending_confirmations.
    12-column XLSX. Decodes volumes; applies character replacements (lines 74-76).
    """
    import openpyxl
    from io import BytesIO
    from openpyxl.styles import PatternFill, Font

    start = '1341100800'
    all_confs = (await db.execute(text("""
        SELECT * FROM confirmation_log
        WHERE date_modified >= :start
          AND (type_of_contract LIKE 'New' OR type_of_contract LIKE 'Addition')
          AND enroll_check <> 1
        ORDER BY SUBSTRING(start_date,7,4), SUBSTRING(start_date,1,2),
                 SUBSTRING(start_date,4,2)
    """), {'start': start})).fetchall()

    matched = (await db.execute(text(f"""
        SELECT c.sid FROM enrollment e
        LEFT JOIN confirmation_log c ON {_CONF_FUZZY_ON}
        WHERE e.type LIKE 'enrollment' AND e.enroll_check <> 1
          AND c.enroll_check <> 1 AND e.enrollment_cleared = 1 AND e.status = 1
          AND (c.type_of_contract LIKE 'New' OR c.type_of_contract LIKE 'Addition')
    """))).fetchall()
    matched_sids = {r[0] for r in matched}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pending Enrollments'
    hfill = PatternFill('solid', fgColor='FFFF00')
    hfont = Font(color='FF0000')
    for col, h in enumerate([
        'S No.', 'Date modified', 'Customer Name', 'Broker name',
        'Contract rate', 'Commission', 'Term', 'Meter Fee',
        'Tax exempt', 'Start date', 'Number of ESIIDS', 'Comments',
    ], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont

    sno = 1
    for r in all_confs:
        row = dict(r._mapping)
        if row['sid'] in matched_sids:
            continue
        name = row.get('customer_name') or ''
        name = name.replace('â', "'")
        name = name.replace('â ', '-')
        name = name.replace('â', 'e')
        ws.append([
            sno,
            _fmt_ts(str(row.get('date_modified') or '')),
            name,
            row.get('broker_name') or '',
            row.get('contract_rate') or '',
            row.get('commission') or '',
            row.get('term') or '',
            row.get('meter_fees') or '',
            row.get('tax_exempt') or '',
            row.get('start_date') or '',
            row.get('esid_count') or '',
            row.get('comment_enrollment') or '',
        ])
        sno += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── get single enrollment (pre-fill for edit modal) ───────────────────────────


async def get_enrollment_by_esid(esid: str, db: AsyncSession) -> Optional[dict]:
    """SELECT * FROM enrollment WHERE esid = ? — used by edit modal prefill."""
    row = (await db.execute(
        text("SELECT * FROM enrollment WHERE esid = :esid"), {'esid': esid}
    )).fetchone()
    return dict(row._mapping) if row else None


async def get_enrollment_log(esid: str, db: AsyncSession) -> List[dict]:
    """Replication of log_check.php line 19."""
    rows = (await db.execute(
        text("SELECT * FROM enrollment_log WHERE esid LIKE :esid ORDER BY date_modified"),
        {'esid': esid}
    )).fetchall()
    return [dict(r._mapping) for r in rows]


# ── fuzzy confirmation JOIN condition ─────────────────────────────────────────
# Exact replication of enrollment_report.php lines 286-293.
# Used in: get_comparison, get_pending_confirmations,
#          get_no_confirmations, download_pending.

_CONF_FUZZY_ON = """
    (c.broker_code LIKE e.broker_code AND c.customer_name LIKE e.company_name)
    OR TRIM(c.customer_name) LIKE TRIM(e.company_name)
    OR REPLACE(c.customer_name, ',', '') LIKE REPLACE(e.company_name, ',', '')
    OR REPLACE(c.customer_name, ' ', '') LIKE REPLACE(e.company_name, ' ', '')
    OR REPLACE(c.customer_name, '’', '') LIKE e.company_name
    OR c.customer_name LIKE CONCAT('%', e.company_name, '%')
    OR TRIM(c.customer_name) LIKE CONCAT(TRIM(e.company_name), '(Revision)')
    OR c.customer_name LIKE e.company_name
    OR TRIM(c.customer_name) LIKE CONCAT(TRIM(e.company_name), '(Revision)(Revision)')
    OR UPPER(c.customer_name) LIKE UPPER(e.company_name)
"""


# ── enrollment / confirmation comparison ──────────────────────────────────────


async def get_comparison(
    start: Optional[str], end: Optional[str], db: AsyncSession
) -> List[dict]:
    """
    Replication of enrollment_report.php lines 285-559.

    Default (no date params):
      WHERE c.date_modified in [1341100800, now]
            AND e.date_added = MAX(date_added)   <- latest batch only
            AND e.type='enrollment', enroll_check<>1, status=1
            AND c.type_of_contract New/Addition
    With date params:
      WHERE c.date_modified in [strtotime(start), strtotime(end)]
            AND e.date_added in same range

    Excludes company names present in additional_esid_template (lines 278-283).
    Auto-updates meter_fee_check when enrollment_meter == confirmation_meter (line 482).
    Contract expiry check (line 507) omitted: PHP concatenates start_date+term+"month"
    producing an unparseable strtotime string, making it dead code.
    """
    template_rows = (await db.execute(
        text("SELECT customer_name FROM additional_esid_template")
    )).fetchall()
    template_names = {r[0] for r in template_rows}

    if start and end:
        start_ts = int(datetime.strptime(start, '%Y-%m-%d').timestamp())
        end_ts   = int(datetime.strptime(end,   '%Y-%m-%d').timestamp())
        sql = text(f"""
            SELECT e.company_name, e.esid, e.broker_code,
                   e.commission AS enrol_comm, e.contract_rate, e.contract_term,
                   e.zone, e.email_id, e.enrollment_status, e.enrolled_status,
                   e.billed_30, e.load_profile_isigma,
                   e.meter_fees AS enrollment_meter, e.meter_fee_check,
                   e.tax_exempt_check, e.tax_exempt1, e.tax_exempt2, e.tax_exempt3,
                   e.tax_exempt4, e.tax_exempt5, e.tax_exempt6, e.tax_exempt7,
                   e.tax_exempt8, e.comparison_comment, e.certificate_check,
                   e.contract_term_check, e.contract_rate_check, e.commission_check,
                   e.contract_start_date, e.date_added,
                   c.sid, c.customer_name, c.commission AS conf_comm,
                   c.contract_rate AS contract_rate_comm, c.term, c.volumes,
                   c.date_modified, c.start_date,
                   c.meter_fees AS confirmation_meter, c.tax_exempt, c.customer_email
            FROM enrollment e
            LEFT JOIN confirmation_log c ON {_CONF_FUZZY_ON}
            WHERE c.date_modified >= :start_ts
              AND c.date_modified <= :end_ts
              AND e.date_added   >= :start_ts
              AND e.date_added   <= :end_ts
              AND e.type LIKE 'enrollment'
              AND e.enroll_check <> 1
              AND e.status = 1
              AND (c.type_of_contract LIKE 'New' OR c.type_of_contract LIKE 'Addition')
            ORDER BY c.sid DESC
        """)
        params: dict = {'start_ts': str(start_ts), 'end_ts': str(end_ts)}
    else:
        max_row = (await db.execute(
            text("SELECT MAX(date_added) AS d FROM enrollment")
        )).fetchone()
        max_date = max_row.d if max_row else 0
        start_ts = 1341100800
        end_ts   = int(time.time())
        sql = text(f"""
            SELECT e.company_name, e.esid, e.broker_code,
                   e.commission AS enrol_comm, e.contract_rate, e.contract_term,
                   e.zone, e.email_id, e.enrollment_status, e.enrolled_status,
                   e.billed_30, e.load_profile_isigma,
                   e.meter_fees AS enrollment_meter, e.meter_fee_check,
                   e.tax_exempt_check, e.tax_exempt1, e.tax_exempt2, e.tax_exempt3,
                   e.tax_exempt4, e.tax_exempt5, e.tax_exempt6, e.tax_exempt7,
                   e.tax_exempt8, e.comparison_comment, e.certificate_check,
                   e.contract_term_check, e.contract_rate_check, e.commission_check,
                   e.contract_start_date, e.date_added,
                   c.sid, c.customer_name, c.commission AS conf_comm,
                   c.contract_rate AS contract_rate_comm, c.term, c.volumes,
                   c.date_modified, c.start_date,
                   c.meter_fees AS confirmation_meter, c.tax_exempt, c.customer_email
            FROM enrollment e
            LEFT JOIN confirmation_log c ON {_CONF_FUZZY_ON}
            WHERE c.date_modified >= :start_ts
              AND c.date_modified <= :end_ts
              AND e.date_added = :max_date
              AND e.type LIKE 'enrollment'
              AND e.enroll_check <> 1
              AND e.status = 1
              AND (c.type_of_contract LIKE 'New' OR c.type_of_contract LIKE 'Addition')
            ORDER BY c.sid DESC
        """)
        params = {
            'start_ts': str(start_ts), 'end_ts': str(end_ts), 'max_date': max_date
        }

    raw_rows = (await db.execute(sql, params)).fetchall()
    results  = []

    for r in raw_rows:
        row = dict(r._mapping)
        if row.get('company_name') in template_names:
            continue

        remarks        = ''
        flag_remarks   = 0
        tax_error      = 0
        tax_error_cert = 0
        tax_cert       = 0
        clean_flag     = 0
        billed_flag    = 0

        if row.get('contract_rate_comm') != row.get('contract_rate'):
            if str(row.get('contract_rate_check', 0)) != '1':
                remarks = 'Contract rate'
                flag_remarks = 1

        if row.get('contract_term') != row.get('term'):
            if str(row.get('contract_term_check', 0)) != '1':
                remarks = (remarks + ', Contract term') if remarks else 'Contract term'
                flag_remarks = 1

        if (row.get('enrol_comm') != row.get('conf_comm') and
                str(row.get('commission_check', 0)) == '0'):
            remarks = (remarks + ', Commission') if remarks else 'Commission'
            flag_remarks = 1

        # PHP: chop($enrollment_meter, '$') strips trailing '$' chars
        enrollment_meter   = str(row.get('enrollment_meter')   or '').rstrip('$')
        confirmation_meter = str(row.get('confirmation_meter') or '')
        if enrollment_meter == confirmation_meter:
            await db.execute(
                text("UPDATE enrollment SET meter_fee_check = 1 WHERE esid LIKE :esid"),
                {'esid': row['esid']}
            )
            row['meter_fee_check'] = 1
        elif str(row.get('meter_fee_check', 0)) == '0':
            flag_remarks = 1
            remarks = (remarks + ', Meter fees') if remarks else 'Meter fees'

        if flag_remarks == 1:
            remarks += ' not matching'

        tax_exempt = row.get('tax_exempt') or ''
        t = {i: str(row.get(f'tax_exempt{i}') or '') for i in (1, 2, 3, 4, 5, 8)}
        if tax_exempt == 'Certificate':
            tax_cert = 1
            if all(t[i] == '100' for i in (1, 2, 3, 4, 5, 8)):
                if str(row.get('certificate_check', 0)) not in ('1', '2'):
                    tax_error_cert = 1
            else:
                if str(row.get('tax_exempt_check', 0)) == '0':
                    tax_error = 1

        if (str(row.get('load_profile_isigma', 0)) == '1' and
                str(row.get('billed_30', 0)) == '1' and
                str(row.get('meter_fee_check', 0)) == '1' and
                tax_error_cert != 1 and tax_error != 1):
            clean_flag = 1
        if str(row.get('billed_30', 0)) == '1':
            billed_flag = 1

        results.append({
            **row,
            'profiles':              _decode_volumes(row.get('volumes') or ''),
            'remarks':               remarks,
            'flag_remarks':          flag_remarks,
            'clean_record_flag':     clean_flag,
            'billed_flag':           billed_flag,
            'tax_error':             tax_error,
            'tax_error_certificate': tax_error_cert,
            'tax_certificate':       tax_cert,
        })

    await db.commit()
    return results


# ── confirmations pending enrollments ─────────────────────────────────────────


async def get_pending_confirmations(
    search_text: Optional[str], db: AsyncSession
) -> List[dict]:
    """
    Replication of confirmations_report_enrollment.php lines 33-173.

    1. All confirmation_log (New/Addition, enroll_check<>1).
    2. SIDs that already have a matching enrollment via fuzzy JOIN.
    3. Return step-1 rows NOT in step-2.
    """
    start = '1341100800'

    if search_text:
        all_confs = (await db.execute(text("""
            SELECT * FROM confirmation_log
            WHERE date_modified >= :start
              AND (type_of_contract LIKE 'New' OR type_of_contract LIKE 'Addition')
              AND enroll_check <> 1
              AND customer_name LIKE :search
            ORDER BY SUBSTRING(start_date, 7, 4),
                     SUBSTRING(start_date, 1, 2),
                     SUBSTRING(start_date, 4, 2)
        """), {'start': start, 'search': f'%{search_text}%'})).fetchall()
    else:
        all_confs = (await db.execute(text("""
            SELECT * FROM confirmation_log
            WHERE date_modified >= :start
              AND (type_of_contract LIKE 'New' OR type_of_contract LIKE 'Addition')
              AND enroll_check <> 1
            ORDER BY SUBSTRING(start_date, 7, 4),
                     SUBSTRING(start_date, 1, 2),
                     SUBSTRING(start_date, 4, 2)
        """), {'start': start})).fetchall()

    matched = (await db.execute(text(f"""
        SELECT c.sid
        FROM enrollment e
        LEFT JOIN confirmation_log c ON {_CONF_FUZZY_ON}
        WHERE e.type LIKE 'enrollment'
          AND e.enroll_check <> 1
          AND c.enroll_check <> 1
          AND e.enrollment_cleared = 1
          AND e.status = 1
          AND (c.type_of_contract LIKE 'New' OR c.type_of_contract LIKE 'Addition')
    """))).fetchall()
    matched_sids = {r[0] for r in matched}

    results = []
    for r in all_confs:
        row = dict(r._mapping)
        if row['sid'] in matched_sids:
            continue
        row['profiles'] = _decode_volumes(row.get('volumes') or '')
        results.append(row)
    return results


# ── enrollments with no confirmations ─────────────────────────────────────────


async def get_no_confirmations(db: AsyncSession) -> List[dict]:
    """
    Replication of view_clear_enrollments.php lines 53-125 (?unchecked=1 mode).

    Three-step exclusion:
    1. Cleared enrollments with compare_check=0.
    2. Remove those matching confirmation_log via fuzzy JOIN.
    3. Remove those matching additional_esid_template.
    """
    unmatched_rows = (await db.execute(text("""
        SELECT * FROM enrollment
        WHERE enroll_check <> 1
          AND enrollment_cleared = 1
          AND compare_check = 0
          AND status = 1
    """))).fetchall()
    records_unmatch = {dict(r._mapping)['esid']: dict(r._mapping)
                       for r in unmatched_rows}

    conf_matched = (await db.execute(text(f"""
        SELECT e.esid
        FROM enrollment e
        LEFT JOIN confirmation_log c ON {_CONF_FUZZY_ON}
        WHERE e.type LIKE 'enrollment'
          AND e.enroll_check <> 1
          AND e.enrollment_cleared = 1
          AND e.status = 1
          AND (c.type_of_contract LIKE 'New' OR c.type_of_contract LIKE 'Addition')
        ORDER BY SUBSTRING(c.start_date, 7, 4),
                 SUBSTRING(c.start_date, 1, 2),
                 SUBSTRING(c.start_date, 4, 2),
                 c.sid, e.load_profile_isigma, e.billed_30 ASC
    """))).fetchall()
    conf_esids = {r[0] for r in conf_matched}

    records_1 = {esid: row for esid, row in records_unmatch.items()
                 if esid not in conf_esids}

    tmpl_matched = (await db.execute(text("""
        SELECT e.esid
        FROM enrollment e
        LEFT JOIN additional_esid_template c
            ON TRIM(c.broker_code) LIKE TRIM(e.broker_code)
           AND (c.customer_name LIKE e.company_name
                OR TRIM(c.customer_name) LIKE TRIM(e.company_name)
                OR c.customer_name LIKE CONCAT('%%', e.company_name, '%%'))
           AND TRIM(c.contract_rate) LIKE TRIM(e.contract_rate)
        WHERE TRIM(c.commission) LIKE TRIM(e.commission)
          AND e.type LIKE 'enrollment'
          AND e.enroll_check <> 1
          AND e.enrollment_cleared = 1
        ORDER BY c.customer_name,
                 SUBSTRING(e.contract_start_date, 7, 4),
                 SUBSTRING(e.contract_start_date, 1, 2),
                 SUBSTRING(e.contract_start_date, 4, 2)
    """))).fetchall()
    tmpl_esids = {r[0] for r in tmpl_matched}

    return [row for esid, row in records_1.items() if esid not in tmpl_esids]


# ── template enrollment comparison ────────────────────────────────────────────


async def get_template_comparison(db: AsyncSession) -> List[dict]:
    """
    Replication of template_enrollment_report.php lines 326-491.

    LEFT JOIN enrollment + additional_esid_template on broker+name+rate+commission.
    Computes remarks, flag_remarks, tax_error (0/1/2), clean_record_flag.
    Auto-updates meter_fee_check when enrollment.meter_fees == template.meter_fee.
    """
    raw_rows = (await db.execute(text("""
        SELECT e.esid, e.broker_code, e.commission AS enrol_comm,
               e.contract_rate, e.contract_term, e.zone, e.email_id,
               e.enrolled_status, e.billed_30, e.load_profile_isigma,
               e.meter_fees, e.meter_fee_check, e.tax_exempt_check,
               e.tax_exempt1, e.tax_exempt2, e.tax_exempt3, e.tax_exempt4,
               e.tax_exempt5, e.tax_exempt6, e.tax_exempt7, e.tax_exempt8,
               e.contract_start_date, e.comparison_comment,
               e.commission_check, e.contract_rate_check,
               c.sid, c.customer_name, c.commission AS template_comm,
               c.contract_rate AS contract_rate_template, c.tax_exempt,
               c.contract_end_date, c.meter_fee
        FROM enrollment e
        LEFT JOIN additional_esid_template c
            ON TRIM(c.broker_code) LIKE TRIM(e.broker_code)
           AND (c.customer_name LIKE e.company_name
                OR TRIM(c.customer_name) LIKE TRIM(e.company_name)
                OR c.customer_name LIKE CONCAT('%%', e.company_name, '%%'))
           AND TRIM(c.contract_rate) LIKE TRIM(e.contract_rate)
        WHERE TRIM(c.commission) LIKE TRIM(e.commission)
          AND e.type LIKE 'enrollment'
          AND e.enroll_check <> 1
          AND e.enrollment_cleared = 1
        ORDER BY c.customer_name,
                 SUBSTRING(e.contract_start_date, 7, 4),
                 SUBSTRING(e.contract_start_date, 1, 2),
                 SUBSTRING(e.contract_start_date, 4, 2)
    """))).fetchall()

    results = []
    for r in raw_rows:
        row          = dict(r._mapping)
        remarks      = ''
        flag_remarks = 0
        tax_error    = 0
        clean_flag   = 0
        billed_flag  = 0

        if row.get('contract_rate_template') != row.get('contract_rate'):
            if str(row.get('contract_rate_check', 0)) != '1':
                remarks = 'Contract rate'
                flag_remarks = 1

        if row.get('enrol_comm') != row.get('template_comm'):
            if str(row.get('commission_check', 0)) != '1':
                remarks = (remarks + ', Commission') if remarks else 'Commission'
                flag_remarks = 1

        if str(row.get('meter_fees') or '') == str(row.get('meter_fee') or ''):
            await db.execute(
                text("UPDATE enrollment SET meter_fee_check = 1 WHERE esid LIKE :esid"),
                {'esid': row['esid']}
            )
            row['meter_fee_check'] = 1
        elif str(row.get('meter_fee_check', 0)) == '0':
            flag_remarks = 1
            remarks = 'Meter fees doesnt match'

        if flag_remarks == 1:
            remarks += ' not matching'

        # tax error — template_enrollment_report.php lines 390-420
        tax_exempt = str(row.get('tax_exempt') or '')
        t = {i: str(row.get(f'tax_exempt{i}') or '') for i in (1, 2, 3, 4, 5, 8)}
        if tax_exempt == 'Residential':
            if not (t[1] == '100' and t[2] != '100' and t[3] == '100' and
                    t[4] == '100' and t[5] == '100' and t[8] == '100'):
                if str(row.get('tax_exempt_check', 0)) == '0':
                    tax_error = 1
        elif tax_exempt == 'Certificate':
            if all(t[i] == '100' for i in (1, 2, 3, 4, 5, 8)):
                tax_error = 2
            else:
                if str(row.get('tax_exempt_check', 0)) == '0':
                    tax_error = 1
        elif not tax_exempt:
            if any(t[i] == '100' for i in (1, 2, 3, 4, 5, 8)):
                if str(row.get('tax_exempt_check', 0)) == '0':
                    tax_error = 1

        if (str(row.get('billed_30', 0)) == '1' and
                str(row.get('meter_fee_check', 0)) == '1' and
                flag_remarks == 0):
            clean_flag = 1
        if str(row.get('billed_30', 0)) == '1':
            billed_flag = 1

        results.append({
            **row,
            'remarks':           remarks,
            'flag_remarks':      flag_remarks,
            'clean_record_flag': clean_flag,
            'billed_flag':       billed_flag,
            'tax_error':         tax_error,
        })

    await db.commit()
    return results


# ── completed enrollment check list ──────────────────────────────────────────


async def get_checked(db: AsyncSession) -> List[dict]:
    """Replication of enrollments_checked.php line 254."""
    rows = (await db.execute(text(
        "SELECT * FROM enrollment WHERE type LIKE 'enrollment' AND enroll_check = 1"
    ))).fetchall()
    return [dict(r._mapping) for r in rows]


# ── non-billed accounts > 35 days ─────────────────────────────────────────────


async def get_non_billed(db: AsyncSession) -> List[dict]:
    """
    Replication of non_billed.php lines 68-73.
    Completed, billed_30=0, date_added < 35 days ago.
    Adds 'days_diff' = floor((today - date_added) / 86400).
    """
    today_ts = int(time.time())
    cutoff   = today_ts - (35 * 86400)

    rows = (await db.execute(text("""
        SELECT * FROM enrollment
        WHERE type = 'enrollment'
          AND billed_30 = 0
          AND date_added < :cutoff
          AND enrollment_status LIKE '%Completed%'
    """), {'cutoff': cutoff})).fetchall()

    results = []
    for r in rows:
        row = dict(r._mapping)
        row['days_diff'] = (today_ts - int(row.get('date_added') or 0)) // 86400
        results.append(row)
    return results


# ── edit enrollment (edit_enrollment.php) ────────────────────────────────────


async def edit_enrollment_record(
    esid: str, payload, db: AsyncSession, username: str
) -> dict:
    """
    Replication of edit_enrollment.php lines 57-73.
    contract_rate entered as display value (÷100); stored ×100.
    Log entry only if company_name OR contract_end_date changed.
    """
    try:
        stored_rate = str(float(payload.contract_rate) * 100)
    except (ValueError, TypeError):
        stored_rate = payload.contract_rate

    await db.execute(text("""
        UPDATE enrollment SET
            company_name        = :company_name,
            broker_code         = :broker_code,
            contract_rate       = :contract_rate,
            commission          = :commission,
            contract_start_date = :contract_start_date,
            contract_end_date   = :contract_end_date,
            zone                = :zone,
            meter_fees          = :meter_fees
        WHERE esid LIKE :esid
    """), {
        'company_name':        payload.company_name,
        'broker_code':         payload.broker_code,
        'contract_rate':       stored_rate,
        'commission':          payload.commission,
        'contract_start_date': payload.contract_start_date,
        'contract_end_date':   payload.contract_end_date,
        'zone':                payload.zone,
        'meter_fees':          payload.meter_fees,
        'esid':                esid,
    })

    # log comment built from changes — lines 64-70
    comment = ''
    if payload.company_name_old != payload.company_name:
        comment += (
            f'<br />Changed company name from '
            f'{payload.company_name_old} to {payload.company_name}'
        )
    if payload.contract_end_date_old != payload.contract_end_date:
        comment += (
            f'<br />Changed contract term from '
            f'{payload.contract_end_date_old} to {payload.contract_end_date}'
        )

    if comment:
        await db.execute(text("""
            INSERT INTO enrollment_log (esid, user, date_modified, comments)
            VALUES (:esid, :user, :date_modified, :comments)
        """), {
            'esid':          esid,
            'user':          username,
            'date_modified': str(int(time.time())),
            'comments':      comment,
        })

    await db.commit()
    return {'message': 'Updated successfully'}


# ── status check (enrollment_status_check.php) ────────────────────────────────


async def status_check_enrollment(
    esid: str, payload, db: AsyncSession, username: str, role: int
) -> dict:
    """
    Replication of enrollment_status_check.php lines 299-378.

    Status construction (lines 308-319):
      'Scheduled' -> 'Scheduled-{txtdate}'
      'Completed'  -> 'Completed-{txtdate1}'  (txtdate1 required)
      Others       -> plain string

    Comment construction (lines 321-328):
      if comment_others=='on': comment_other = '<br />others-{textarea}'
      stored: enrollment_status_comment = comment + comment_other

    Log (lines 350-363):
      status changed  -> 'Changed enrollment status from {old} to {new}'
      comment changed -> 'Comment changed - {comment}'

    Admin delete (lines 365-372, role==1 only):
      status == 'Cancelled By Customer' -> DELETE enrollment + log

    NOTE: comment_active and archive columns absent from DB — those UPDATE
    statements from PHP lines 332-376 are excluded.
    """
    status = payload.radio1 or ''
    if status == 'Scheduled':
        status = f'Scheduled-{payload.txtdate or ""}'
    elif status == 'Completed':
        if not payload.txtdate1:
            raise ValueError('Date cannot be empty for completed enrollments')
        status = f'Completed-{payload.txtdate1}'

    if payload.comment_others == 'on':
        comment_other = f"<br />others-{payload.txtarea or ''}"
        comment_other = comment_other.replace("'", "\\'")
    else:
        comment_other = ''

    full_comment = (payload.comment or '') + comment_other

    await db.execute(text("""
        UPDATE enrollment
        SET enrollment_status          = :status,
            enrollment_status_comment  = :comment
        WHERE TRIM(esid) LIKE TRIM(:esid)
    """), {'status': status, 'comment': full_comment, 'esid': esid})

    # log — lines 350-363
    status_old  = payload.status_old or '-'
    log_comment = None
    if status_old != status:
        if full_comment:
            log_comment = (
                f'Changed enrollment status from {status_old} to {status}'
                f' with comment - {full_comment}'
            )
        else:
            log_comment = f'Changed enrollment status from {status_old} to {status}'
    else:
        if full_comment:
            log_comment = f'Comment changed - {full_comment}'

    if log_comment:
        await db.execute(text("""
            INSERT INTO enrollment_log (esid, user, date_modified, comments)
            VALUES (:esid, :user, :date_modified, :comments)
        """), {
            'esid':          esid,
            'user':          username,
            'date_modified': str(int(time.time())),
            'comments':      log_comment,
        })

    # admin-only delete — lines 365-372
    if status == 'Cancelled By Customer' and role == 1:
        await db.execute(
            text("DELETE FROM enrollment WHERE esid LIKE :esid"), {'esid': esid}
        )
        await db.execute(text("""
            INSERT INTO enrollment_log (esid, user, date_modified, comments)
            VALUES (:esid, :user, :date_modified, :comments)
        """), {
            'esid':          esid,
            'user':          username,
            'date_modified': str(int(time.time())),
            'comments':      'Deleted this esid from the enrollment',
        })

    await db.commit()
    return {'message': 'Updated successfully'}


# ── clear enrollment (clear_enrollment.php) ──────────────────────────────────


async def clear_enrollment_record(esid: str, db: AsyncSession) -> dict:
    """
    Replication of clear_enrollment.php lines 7-9.
    UPDATE enrollment SET enrollment_cleared = 1 WHERE esid = ?
    """
    existing = (await db.execute(
        text("SELECT esid FROM enrollment WHERE esid = :esid"), {'esid': esid}
    )).fetchone()
    if not existing:
        return {'result': 0}
    await db.execute(
        text("UPDATE enrollment SET enrollment_cleared = 1 WHERE esid = :esid"),
        {'esid': esid}
    )
    await db.commit()
    return {'result': 1}


# ── toggle additional ESI ID check (enrollment_additional_esid_check.php) ─────


async def toggle_additional_esid_check(
    esid: str, db: AsyncSession, username: str
) -> dict:
    """
    Replication of enrollment_additional_esid_check.php lines 8-19.
    Toggle additional_esid_check 0<->1; INSERT enrollment_log.
    """
    row = (await db.execute(
        text("SELECT additional_esid_check FROM enrollment WHERE esid = :esid"),
        {'esid': esid}
    )).fetchone()
    if not row:
        return {'result': 0}

    if int(row.additional_esid_check or 0) == 0:
        new_val = 1
        log_msg = 'Meter added to additional ESIID'
    else:
        new_val = 0
        log_msg = 'Meter removed from additional ESIID'

    await db.execute(
        text("UPDATE enrollment SET additional_esid_check = :val WHERE esid = :esid"),
        {'val': new_val, 'esid': esid}
    )
    await db.execute(text("""
        INSERT INTO enrollment_log (esid, user, date_modified, comments)
        VALUES (:esid, :user, :date_modified, :comments)
    """), {
        'esid':          esid,
        'user':          username,
        'date_modified': str(int(time.time())),
        'comments':      log_msg,
    })
    await db.commit()
    return {'result': 1}


# ── approve enrollment (clear_enrollment_report.php) ─────────────────────────


async def approve_enrollment(esid: str, payload, db: AsyncSession) -> dict:
    """
    Replication of clear_enrollment_report.php lines 8-38.
    type='confirmation': UPDATE enrollment enroll_check=1
                         UPDATE confirmation_log enroll_check=1
    other:               UPDATE enrollment enroll_check=1 only
    """
    exists_e = (await db.execute(
        text("SELECT esid FROM enrollment WHERE esid = :esid"), {'esid': esid}
    )).fetchone()
    if not exists_e:
        return {'result': 0}

    if payload.type == 'confirmation' and payload.sid:
        exists_c = (await db.execute(
            text("SELECT sid FROM confirmation_log WHERE sid = :sid"),
            {'sid': payload.sid}
        )).fetchone()
        if not exists_c:
            return {'result': 0}
        await db.execute(
            text("UPDATE enrollment SET enroll_check = 1 WHERE esid LIKE :esid"),
            {'esid': esid}
        )
        await db.execute(
            text("UPDATE confirmation_log SET enroll_check = 1 WHERE sid = :sid"),
            {'sid': payload.sid}
        )
    else:
        await db.execute(
            text("UPDATE enrollment SET enroll_check = 1 WHERE esid LIKE :esid"),
            {'esid': esid}
        )

    await db.commit()
    return {'result': 1}


# ── action enrollment (delete_enrollment.php) ─────────────────────────────────


async def action_enrollment(esid: str, payload, db: AsyncSession) -> dict:
    """
    Replication of delete_enrollment.php lines 8-16.
    type='delete': UPDATE SET enrollment_status='Delete'  (soft delete)
    type='update': UPDATE SET enroll_check=1              (admin clearance)
    """
    existing = (await db.execute(
        text("SELECT esid FROM enrollment WHERE esid LIKE :esid"), {'esid': esid}
    )).fetchone()
    if not existing:
        return {'result': 0}

    if payload.type == 'delete':
        await db.execute(
            text("UPDATE enrollment SET enrollment_status = 'Delete' WHERE esid LIKE :esid"),
            {'esid': esid}
        )
    else:
        await db.execute(
            text("UPDATE enrollment SET enroll_check = 1 WHERE esid LIKE :esid"),
            {'esid': esid}
        )

    await db.commit()
    return {'result': 1}


# ── dismiss confirmation (delete_confirmation_enrollment.php) ─────────────────


async def dismiss_confirmation(sid: int, db: AsyncSession) -> dict:
    """
    Replication of delete_confirmation_enrollment.php lines 6-11.
    Sets enroll_check=1 (not a DELETE — PHP naming is misleading).
    """
    existing = (await db.execute(
        text("SELECT sid FROM confirmation_log WHERE sid = :sid"), {'sid': sid}
    )).fetchone()
    if not existing:
        return {'result': 0}
    await db.execute(
        text("UPDATE confirmation_log SET enroll_check = 1 WHERE sid = :sid"),
        {'sid': sid}
    )
    await db.commit()
    return {'result': 1}
