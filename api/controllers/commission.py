"""
Commission module controller
Handles: upload, broker sync, view/edit/delete, payment upload,
         adjustments, calculate commission, review summary, user logging
"""

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from fastapi import HTTPException
import openpyxl
import xlrd
import io
import os
from datetime import datetime
from typing import Optional

TENANT_LOGO_URL = os.getenv("TENANT_LOGO_URL", "")
TENANT_DISPLAY_NAME = os.getenv("TENANT_DISPLAY_NAME", "Logo")


def _load_workbook_rows(file_bytes: bytes, filename: str = "") -> list:
    """
    Load Excel file rows as list of tuples.
    Handles both .xlsx (openpyxl) and .xls (xlrd) formats.
    Returns list of rows, each row is a tuple of cell values.
    """
    is_xls = filename.lower().endswith(".xls") and not filename.lower().endswith(
        ".xlsx"
    )

    if is_xls:
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        rows = []
        for i in range(ws.nrows):
            row = []
            for j in range(ws.ncols):
                cell = ws.cell(i, j)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    # Convert date serial to string
                    dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    row.append(dt.strftime("%b-%y"))
                else:
                    row.append(cell.value)
            rows.append(tuple(row))
        return rows
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        return list(ws.iter_rows(values_only=True))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _current_month_name() -> str:
    """Returns full month name e.g. 'April'"""
    return datetime.now().strftime("%B")


def _current_month_short() -> str:
    """Returns e.g. 'Apr-26'"""
    return datetime.now().strftime("%b-%y")


def _current_end_date_pattern() -> str:
    """Returns LIKE pattern e.g. '2026-04-%'"""
    return datetime.now().strftime("%Y-%m-%%")


async def _log_user_action(
    db: AsyncSession,
    uid: int,
    user_name: str,
    action: str,
    broker_name: str = "",
):
    await db.execute(
        text(
            "INSERT INTO user_log (uid, user_name, broker_name, action, date, flag) "
            "VALUES (:uid, :user_name, :broker_name, :action, :date, 'commission')"
        ),
        {
            "uid": uid,
            "user_name": user_name,
            "broker_name": broker_name,
            "action": action,
            "date": str(int(datetime.now().timestamp())),
        },
    )


async def _get_vendor_map(db: AsyncSession) -> dict:
    """
    Returns {broker_code: vendor} from broker_new
    e.g. {'V0364': 'V1', 'V0059': 'V2', ...}
    """
    result = await db.execute(
        text("SELECT DISTINCT broker_code, vendor FROM broker_new")
    )
    return {row.broker_code.strip(): row.vendor.strip() for row in result}


def _clean_cell(value) -> str:
    """Strip $, commas, single quotes from cell values"""
    if value is None:
        return ""
    s = str(value).strip()
    for ch in ("'", "$", ","):
        s = s.replace(ch, "")
    return s


# ---------------------------------------------------------------------------
# 1. UPLOAD COMMISSION FILE
# ---------------------------------------------------------------------------


async def upload_commission_file(
    file_bytes: bytes,
    filename: str,
    start_date: str,
    end_date: str,
    uid: int,
    user_name: str,
    db: AsyncSession,
) -> dict:
    """
    Parses Excel commission file and inserts into comm_bank.
    Pre-flight checks:
      1. Duplicate month check
      2. All vendor_ids must exist in broker_new
    """
    # -- Parse end_date to get month name (end_date = MM/DD/YYYY)
    try:
        end_dt = datetime.strptime(end_date, "%m/%d/%Y")
    except ValueError:
        raise HTTPException(
            status_code=400, detail="Invalid end_date format. Use MM/DD/YYYY"
        )

    month_name = end_dt.strftime("%B")
    year_str = end_dt.strftime("%Y")

    # -- Check 1: duplicate upload for this month
    dup_check = await db.execute(
        text(
            "SELECT COUNT(*) as cnt FROM comm_bank "
            "WHERE month = :month AND SUBSTR(end_date, 7, 4) = :year"
        ),
        {"month": month_name, "year": year_str},
    )
    dup_count = dup_check.scalar()
    if dup_count and dup_count > 1:
        raise HTTPException(
            status_code=409, detail=f"Data already uploaded for {month_name} {year_str}"
        )

    # -- Load vendor map from broker_new
    vendor_map = await _get_vendor_map(db)

    # -- Parse Excel (supports both .xls and .xlsx)
    rows = _load_workbook_rows(file_bytes, filename)

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="File has no data rows")

    # -- Check 2: pre-flight broker existence scan
    # -- Load all brokers including those with blank vendor/commission ID
    all_brokers_result = await db.execute(
        text("SELECT DISTINCT broker_code, vendor FROM broker_new")
    )
    all_brokers = {
        str(r.broker_code).strip().upper(): (r.vendor or "").strip()
        for r in all_brokers_result
    }

    # -- Pre-flight check
    not_in_db = []
    no_commission_id = []

    for row in rows[1:]:
        if not row or not row[0]:
            continue
        vendor_id = str(row[0]).strip().upper()
        if vendor_id == "V0562":
            continue
        if vendor_id not in all_brokers:
            not_in_db.append(vendor_id)
        elif all_brokers[vendor_id] == "":
            no_commission_id.append(vendor_id)

    print(f"DEBUG: Total rows: {len(rows)}")
    print(f"DEBUG: not_in_db: {not_in_db}")
    print(f"DEBUG: no_commission_id: {no_commission_id}")
    print(f"DEBUG: vendor_map sample: {list(vendor_map.items())[:3]}")

    if not_in_db:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "These broker codes are not in the Broker Database. Add them first.",
                "missing_brokers": list(set(not_in_db)),
                "no_commission_id": [],
            },
        )

    if no_commission_id:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "These brokers exist but have no Commission ID yet. Assign it in Broker Database first.",
                "missing_brokers": [],
                "no_commission_id": list(set(no_commission_id)),
            },
        )

    # -- Insert rows
    total_inserted = 0
    print(f"DEBUG: Starting insert, {len(rows)-1} rows to process")
    print(f"DEBUG: vendor_map has {len(vendor_map)} entries")
    print(f"DEBUG: Sample row[1]: {rows[1]}")
    col_names = [
        "vendor",
        "vendor_id",
        "vendor_name",
        "customer_id",
        "premise_id",
        "bill_no",
        "first_name",
        "last_name",
        "company_name",
        "cust_status",
        "last_recd_date",
        "bill_date",
        "service_start_date",
        "service_end_date",
        "plan_group",
        "plan_group_type",
        "plan_group_desc",
        "contract_id",
        "contract_no",
        "contract_status",
        "contract_start_date",
        "contract_end_date",
        "contract_category",
        "contract_type",
        "contract_rate",
        "inv_amount",
        "due_amount",
        "kwh_usage",
        "metered_usage",
        "usage_factor",
        "billed_prorated_usage",
        "actual_prorated_usage",
        "commission_rate",
        "commission_amount",
        "comm_amt_by_actual_usage",
        "comm_min_amt",
        "comm_max_amt",
        "agent_commission_terms",
        "pay_date",
        "pay_period_start",
        "pay_period_end",
        "comments",
    ]

    for row in rows[1:]:
        if not row or not row[0]:
            continue
        vendor_id = str(row[0]).strip().upper()
        if vendor_id == "V0562":
            continue

        vendor_label = vendor_map.get(vendor_id, "")

        # Get broker_status from commission_status
        broker_status_result = await db.execute(
            text(
                "SELECT commission_status FROM broker_new "
                "WHERE broker_code LIKE :bc LIMIT 1"
            ),
            {"bc": vendor_id},
        )
        comm_status_row = broker_status_result.fetchone()
        broker_status = (
            0 if (comm_status_row and str(comm_status_row[0]).lower() == "false") else 1
        )

        # Build values — Excel cols map 1:1 to comm_bank cols (after vendor prepend)
        cleaned = [_clean_cell(c) for c in row]

        # commission_amount is col index 33 (0-based) — keep as numeric
        try:
            comm_amount = float(cleaned[33]) if cleaned[33] else 0.0
        except (ValueError, IndexError):
            comm_amount = 0.0

        try:
            await db.execute(
                text(
                    """
                    INSERT INTO comm_bank (
                        vendor, vendor_id, vendor_name, customer_id, premise_id, 
                        bill_no, first_name, last_name, company_name, cust_status, 
                        last_recd_date, bill_date, service_start_date, service_end_date, 
                        plan_group, plan_group_type, plan_group_desc, contract_id, 
                        contract_no, contract_status, contract_start_date, contract_end_date, 
                        contract_category, contract_type, contract_rate, inv_amount, 
                        due_amount, kwh_usage, metered_usage, usage_factor, 
                        billed_prorated_usage, actual_prorated_usage, commission_rate, 
                        commission_amount, comm_amt_by_actual_usage, comm_min_amt, 
                        comm_max_amt, agent_commission_terms, pay_date, pay_period_start, 
                        pay_period_end, comments, month, start_date, end_date, broker_status
                    ) VALUES (
                        :vendor, :vendor_id, :vendor_name, :customer_id, :premise_id, 
                        :bill_no, :first_name, :last_name, :company_name, :cust_status, 
                        :last_recd_date, :bill_date, :service_start_date, :service_end_date, 
                        :plan_group, :plan_group_type, :plan_group_desc, :contract_id, 
                        :contract_no, :contract_status, :contract_start_date, :contract_end_date, 
                        :contract_category, :contract_type, :contract_rate, :inv_amount, 
                        :due_amount, :kwh_usage, :metered_usage, :usage_factor, 
                        :billed_prorated_usage, :actual_prorated_usage, :commission_rate, 
                        :commission_amount, :comm_amt_by_actual_usage, :comm_min_amt, 
                        :comm_max_amt, :agent_commission_terms, :pay_date, :pay_period_start, 
                        :pay_period_end, :comments, :month, :start_date, :end_date, :broker_status
                    )
                """
                ),
                {
                    "vendor": vendor_label,
                    "vendor_id": cleaned[0] if len(cleaned) > 0 else "",
                    "vendor_name": cleaned[1] if len(cleaned) > 1 else "",
                    "customer_id": cleaned[2] if len(cleaned) > 2 else "",
                    "premise_id": cleaned[3] if len(cleaned) > 3 else "",
                    "bill_no": cleaned[4] if len(cleaned) > 4 else "",
                    "first_name": cleaned[5] if len(cleaned) > 5 else "",
                    "last_name": cleaned[6] if len(cleaned) > 6 else "",
                    "company_name": cleaned[7] if len(cleaned) > 7 else "",
                    "cust_status": cleaned[8] if len(cleaned) > 8 else "",
                    "last_recd_date": cleaned[9] if len(cleaned) > 9 else "",
                    "bill_date": cleaned[10] if len(cleaned) > 10 else "",
                    "service_start_date": cleaned[11] if len(cleaned) > 11 else "",
                    "service_end_date": cleaned[12] if len(cleaned) > 12 else "",
                    "plan_group": cleaned[13] if len(cleaned) > 13 else "",
                    "plan_group_type": cleaned[14] if len(cleaned) > 14 else "",
                    "plan_group_desc": cleaned[15] if len(cleaned) > 15 else "",
                    "contract_id": cleaned[16] if len(cleaned) > 16 else "",
                    "contract_no": cleaned[17] if len(cleaned) > 17 else "",
                    "contract_status": cleaned[18] if len(cleaned) > 18 else "",
                    "contract_start_date": cleaned[19] if len(cleaned) > 19 else "",
                    "contract_end_date": cleaned[20] if len(cleaned) > 20 else "",
                    "contract_category": cleaned[21] if len(cleaned) > 21 else "",
                    "contract_type": cleaned[22] if len(cleaned) > 22 else "",
                    "contract_rate": cleaned[23] if len(cleaned) > 23 else "",
                    "inv_amount": cleaned[24] if len(cleaned) > 24 else "",
                    "due_amount": cleaned[25] if len(cleaned) > 25 else "",
                    "kwh_usage": cleaned[26] if len(cleaned) > 26 else "",
                    "metered_usage": cleaned[27] if len(cleaned) > 27 else "",
                    "usage_factor": cleaned[28] if len(cleaned) > 28 else "",
                    "billed_prorated_usage": cleaned[29] if len(cleaned) > 29 else "",
                    "actual_prorated_usage": cleaned[30] if len(cleaned) > 30 else "",
                    "commission_rate": cleaned[31] if len(cleaned) > 31 else "",
                    "commission_amount": comm_amount,
                    "comm_amt_by_actual_usage": (
                        cleaned[33] if len(cleaned) > 33 else ""
                    ),
                    "comm_min_amt": cleaned[34] if len(cleaned) > 34 else "",
                    "comm_max_amt": cleaned[35] if len(cleaned) > 35 else "",
                    "agent_commission_terms": cleaned[36] if len(cleaned) > 36 else "",
                    "pay_date": cleaned[37] if len(cleaned) > 37 else "",
                    "pay_period_start": cleaned[38] if len(cleaned) > 38 else "",
                    "pay_period_end": cleaned[39] if len(cleaned) > 39 else "",
                    "comments": cleaned[40] if len(cleaned) > 40 else "",
                    "month": month_name,
                    "start_date": datetime.strptime(start_date, "%m/%d/%Y").strftime(
                        "%Y-%m-%d"
                    ),
                    "end_date": datetime.strptime(end_date, "%m/%d/%Y").strftime(
                        "%Y-%m-%d"
                    ),
                    "broker_status": broker_status,
                },
            )
            total_inserted += 1
            if total_inserted == 1:
                print(f"DEBUG: First row inserted successfully")
        except Exception as e:
            print(f"DEBUG INSERT ERROR row {total_inserted}: {e}")
            raise
    await db.commit()
    print(f"DEBUG: Committed. Total inserted: {total_inserted}")
    await _log_user_action(
        db,
        uid,
        user_name,
        f"Uploaded commission file for {month_name} {year_str} — {total_inserted} rows",
    )
    await db.commit()

    return {
        "inserted": total_inserted,
        "month": month_name,
        "year": year_str,
        "month_param": f"{year_str}-{end_dt.strftime('%m')}",  # e.g. '2026-04'
    }


# ---------------------------------------------------------------------------
# 2. VIEW DATA  (with audit modes)
# ---------------------------------------------------------------------------


async def get_commission_data(
    db: AsyncSession,
    vendor: Optional[str] = None,
    vendors: Optional[list] = None,
    from_month: Optional[str] = None,
    to_month: Optional[str] = None,
    quick_period: Optional[int] = None,
    audit_mode: Optional[str] = None,
    mon_count: int = 6,
) -> list:
    if audit_mode == "double_payment":
        return await _audit_double_payment(db, vendor, mon_count)
    elif audit_mode == "variance_30":
        return await _audit_variance_30(db, vendor, mon_count)
    elif audit_mode == "inactive":
        return await _audit_inactive(db, vendor, from_month)

    conditions = ["broker_status = 1"]
    params: dict = {}

    # Multiple vendors
    if vendors:
        placeholders = ", ".join([f":v{i}" for i in range(len(vendors))])
        conditions.append(f"vendor IN ({placeholders})")
        for i, v in enumerate(vendors):
            params[f"v{i}"] = v
    elif vendor:
        conditions.append("vendor = :vendor")
        params["vendor"] = vendor

    # Quick period — last N months
    if quick_period:
        month_conditions = []
        now = datetime.now()
        for j in range(quick_period):
            month_num = (now.month - j - 1) % 12 + 1
            year = now.year - ((now.month - j - 1) // 12)
            month_conditions.append(f"end_date LIKE :mp{j}")
            params[f"mp{j}"] = f"{year}-{month_num:02d}-%"
        conditions.append(f"({' OR '.join(month_conditions)})")

    # Manual range
    elif from_month and to_month:
        conditions.append("end_date >= :from_m AND end_date <= :to_m")
        params["from_m"] = from_month
        params["to_m"] = to_month
    elif from_month:
        conditions.append("end_date LIKE :from_pat")
        params["from_pat"] = f"{from_month}-%"
    else:
        conditions.append("end_date LIKE :end_pat")
        params["end_pat"] = _current_end_date_pattern()

    where = " AND ".join(conditions)
    result = await db.execute(
        text(
            f"SELECT * FROM comm_bank WHERE {where} "
            "ORDER BY vendor, premise_id, service_start_date"
        ),
        params,
    )
    print(f"DEBUG data query - end_pat: {params.get('end_pat')}")
    print(f"DEBUG data query - conditions: {conditions}")
    return [dict(row) for row in result.mappings()]


async def _build_month_conditions(vendor: str, mon_count: int) -> str:
    """Build OR conditions for last N months for a vendor"""
    parts = []
    now = datetime.now()
    for j in range(mon_count):
        # subtract j months
        month_num = (now.month - j - 1) % 12 + 1
        year = now.year - ((now.month - j - 1) // 12)
        parts.append(
            f"(vendor LIKE '{vendor}' AND end_date LIKE '{month_num:02d}/%/{year}')"
        )
    return " OR ".join(parts)


async def _audit_double_payment(
    db: AsyncSession, vendor: Optional[str], mon_count: int
) -> list:
    """Find same premise_id + same service dates across different months (or same amount/month)"""
    if vendor:
        vendors_list = [vendor]
    else:
        result = await db.execute(
            text(
                "SELECT DISTINCT vendor FROM comm_bank WHERE broker_status = 1 ORDER BY vendor"
            )
        )
        vendors_list = [row[0] for row in result]

    flagged = []
    for v in vendors_list:
        month_cond = await _build_month_conditions(v, mon_count)
        result = await db.execute(
            text(
                f"SELECT * FROM comm_bank WHERE ({month_cond}) AND broker_status = 1 "
                "ORDER BY vendor_id, premise_id, service_start_date, service_end_date, end_date"
            )
        )
        rows = [dict(r) for r in result.mappings()]
        for i in range(1, len(rows)):
            curr = rows[i]
            prev = rows[i - 1]
            same_premise = curr["premise_id"] == prev["premise_id"]
            same_start = curr["service_start_date"] == prev["service_start_date"]
            same_end = curr["service_end_date"] == prev["service_end_date"]
            diff_month = curr["month"] != prev["month"]
            same_amount = curr["commission_amount"] == prev["commission_amount"]
            same_month = curr["month"] == prev["month"]

            if (
                same_premise
                and same_start
                and same_end
                and (diff_month or (same_amount and same_month))
            ):
                prev["double_payment"] = True
                curr["double_payment"] = True
                if prev not in flagged:
                    flagged.append(prev)
                flagged.append(curr)

    return flagged


async def _audit_variance_30(
    db: AsyncSession, vendor: Optional[str], mon_count: int
) -> list:
    """Find premise_ids where commission changed >= 30% month over month (and amount > $30)"""
    if vendor:
        vendors_list = [vendor]
    else:
        result = await db.execute(
            text(
                "SELECT DISTINCT vendor FROM comm_bank WHERE broker_status = 1 ORDER BY vendor"
            )
        )
        vendors_list = [row[0] for row in result]

    flagged = []
    for v in vendors_list:
        month_cond = await _build_month_conditions(v, mon_count)
        result = await db.execute(
            text(
                f"SELECT * FROM comm_bank WHERE ({month_cond}) AND broker_status = 1 "
                "ORDER BY vendor_id, premise_id, service_start_date, service_end_date, end_date"
            )
        )
        rows = [dict(r) for r in result.mappings()]
        for i in range(1, len(rows)):
            curr = rows[i]
            prev = rows[i - 1]
            if curr["premise_id"] != prev["premise_id"]:
                continue
            try:
                curr_amt = float(curr["commission_amount"] or 0)
                prev_amt = float(prev["commission_amount"] or 0)
            except (ValueError, TypeError):
                continue
            if prev_amt == 0:
                continue
            if curr_amt > 30 or prev_amt > 30:
                pct_change = ((curr_amt - prev_amt) / prev_amt) * 100
                if pct_change >= 30:
                    prev["variance_pct"] = round(pct_change, 2)
                    curr["variance_pct"] = round(pct_change, 2)
                    if prev not in flagged:
                        flagged.append(prev)
                    flagged.append(curr)

    return flagged


async def _audit_inactive(
    db: AsyncSession, vendor: Optional[str], month: Optional[str]
) -> list:
    end_pattern = _current_end_date_pattern()
    conditions = ["broker_status = 1", "(cust_status = 'I' OR cust_status = 'G')"]
    params: dict = {}

    if vendor:
        conditions.append("vendor = :vendor")
        params["vendor"] = vendor
    if month:
        conditions.append("end_date IN (:month)")
        params["month"] = month
    else:
        conditions.append("end_date LIKE :end_pat")
        params["end_pat"] = end_pattern

    where = " AND ".join(conditions)
    result = await db.execute(
        text(
            f"SELECT * FROM comm_bank WHERE {where} "
            "ORDER BY vendor_id, premise_id, service_start_date"
        ),
        params,
    )
    return [dict(row) for row in result.mappings()]


# ---------------------------------------------------------------------------
# 3. EDIT / DELETE comm_bank rows
# ---------------------------------------------------------------------------


async def update_commission_row(
    sid: int,
    data: dict,
    uid: int,
    user_name: str,
    db: AsyncSession,
) -> dict:
    allowed = {
        "commission_rate",
        "commission_amount",
        "comm_amt_by_actual_usage",
        "comm_min_amt",
        "comm_max_amt",
        "comments",
        "cust_status",
        "kwh_usage",
        "metered_usage",
        "broker_status",
    }
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        raise HTTPException(status_code=400, detail="No valid fields to update")

    set_clause = ", ".join([f"{k} = :{k}" for k in updates])
    updates["sid"] = sid

    await db.execute(
        text(f"UPDATE comm_bank SET {set_clause} WHERE sid = :sid"),
        updates,
    )
    await db.commit()
    await _log_user_action(
        db,
        uid,
        user_name,
        f"Edited comm_bank row sid={sid} fields={list(updates.keys())}",
    )
    await db.commit()
    return {"status": "updated"}


async def _recalculate_vendor_summary(
    vendor: str,
    db: AsyncSession,
    month: str = None,
) -> None:
    if month:
        try:
            dt_obj = datetime.strptime(month, "%Y-%m")
            curr_month_label = dt_obj.strftime("%b-%y")
            end_pattern = dt_obj.strftime("%Y-%m-%%")
        except ValueError:
            curr_month_label = _current_month_short()
            end_pattern = _current_end_date_pattern()
    else:
        curr_month_label = _current_month_short()
        end_pattern = _current_end_date_pattern()

    # Re-sum commission for this vendor this month
    result = await db.execute(
        text(
            "SELECT SUM(commission_amount) as total FROM comm_bank "
            "WHERE vendor = :vendor AND end_date LIKE :pat AND broker_status = 1"
        ),
        {"vendor": vendor, "pat": end_pattern},
    )
    row = result.fetchone()
    owed = round(float(row[0] or 0), 2)

    # Get current month row in summary_payments
    curr_row_result = await db.execute(
        text(
            "SELECT * FROM summary_payments "
            "WHERE vendor = :vendor AND month = :month "
            "ORDER BY sid DESC LIMIT 1"
        ),
        {"vendor": vendor, "month": curr_month_label},
    )
    curr_row = curr_row_result.fetchone()

    # Get previous balance
    prev_row_result = await db.execute(
        text(
            "SELECT balance FROM summary_payments "
            "WHERE vendor = :vendor AND month != :month "
            "ORDER BY sid DESC LIMIT 1"
        ),
        {"vendor": vendor, "month": curr_month_label},
    )
    prev_row = prev_row_result.fetchone()
    prev_balance = round(float(prev_row[0] or 0), 2) if prev_row else 0.0

    if curr_row:
        payment = round(float(curr_row[4] or 0), 2)
        new_balance = round(prev_balance + owed + payment, 2)
        await db.execute(
            text(
                "UPDATE summary_payments SET owed = :owed, balance = :balance "
                "WHERE sid = :sid"
            ),
            {"owed": owed, "balance": new_balance, "sid": curr_row[0]},
        )
    else:
        # No summary row yet — insert one
        new_balance = round(prev_balance + owed, 2)
        await db.execute(
            text(
                "INSERT INTO summary_payments "
                "(vendor, month, owed, payment, balance) "
                "VALUES (:vendor, :month, :owed, :payment, :balance)"
            ),
            {
                "vendor": vendor,
                "month": curr_month_label,
                "owed": owed,
                "payment": 0,
                "balance": new_balance,
            },
        )
    await db.commit()


async def delete_commission_row(
    sid: int,
    uid: int,
    user_name: str,
    db: AsyncSession,
) -> dict:
    # Get vendor before deleting
    # Get vendor AND end_date before deleting
    row_result = await db.execute(
        text("SELECT vendor, end_date FROM comm_bank WHERE sid = :sid"), {"sid": sid}
    )
    row = row_result.fetchone()
    vendor = row[0] if row else None
    end_date_val = str(row[1])[:7] if row and row[1] else None  # '2026-04'

    await db.execute(text("DELETE FROM comm_bank WHERE sid = :sid"), {"sid": sid})
    await db.commit()

    if vendor:
        await _recalculate_vendor_summary(vendor, db, end_date_val)

    await _log_user_action(db, uid, user_name, f"Deleted comm_bank row sid={sid}")
    await db.commit()
    return {"status": "deleted"}


async def delete_commission_month(
    month: str,
    uid: int,
    user_name: str,
    db: AsyncSession,
) -> dict:
    """Delete all comm_bank rows for a given month (e.g. 'March')"""
    result = await db.execute(
        text("DELETE FROM comm_bank WHERE month = :month"),
        {"month": month},
    )
    await db.commit()
    await _log_user_action(
        db, uid, user_name, f"Deleted all comm_bank data for month={month}"
    )
    await db.commit()
    return {"status": "deleted", "month": month}


# ---------------------------------------------------------------------------
# 4. UPLOAD PAYMENT SUMMARY
# ---------------------------------------------------------------------------


async def upload_payment_summary(
    file_bytes: bytes,
    filename: str,
    month: str = None,
    uid: int = 0,
    user_name: str = "",
    db: AsyncSession = None,
) -> dict:
    rows = _load_workbook_rows(file_bytes, filename)

    if not rows:
        raise HTTPException(status_code=400, detail="Empty file")

    # Use user-selected month
    if month:
        try:
            dt = datetime.strptime(month, "%Y-%m")
            curr_month_label = dt.strftime("%b-%y")
        except ValueError:
            curr_month_label = _current_month_short()
    else:
        curr_month_label = _current_month_short()

    # Verify commission data exists for this month
    comm_check = await db.execute(
        text("SELECT COUNT(*) FROM comm_bank WHERE end_date LIKE :pat"),
        {"pat": f"{month}-%"},
    )
    if comm_check.scalar() == 0:
        raise HTTPException(
            status_code=400,
            detail=f"No commission data found for {curr_month_label}. Upload commission file first.",
        )

    # Data rows start at row 4
    # Month validation disabled for historical uploads
    # current_month_short = datetime.now().strftime("%b").lower()
    # if header_month != current_month_short:
    # raise HTTPException(
    # status_code=400,
    # detail=f"File is for '{header_month}' but current month is '{current_month_short}'. "
    # "Use Adjustments for previous months.",
    # )
    total = 0
    for row in rows[3:]:
        # ... rest of existing code
        if not row or not row[1]:
            continue

        vendor_label = str(row[1]).strip()
        raw_payment = row[3] if len(row) > 3 else None
        payment_str = _clean_cell(raw_payment)
        try:
            payment = float(payment_str) if payment_str else 0.0
        except ValueError:
            payment = 0.0

        # Check if record exists for this vendor + this month
        existing = await db.execute(
            text(
                "SELECT * FROM summary_payments WHERE vendor LIKE :vendor "
                "AND month LIKE :month ORDER BY sid DESC LIMIT 1"
            ),
            {"vendor": vendor_label, "month": curr_month_label},
        )
        existing_row = existing.fetchone()

        if existing_row:
            # Update payment (stored as negative per PHP logic)
            await db.execute(
                text("UPDATE summary_payments SET payment = :payment WHERE sid = :sid"),
                {"payment": -payment, "sid": existing_row[0]},
            )
        else:
            # Get last balance to carry forward
            last_row = await db.execute(
                text(
                    "SELECT * FROM summary_payments WHERE vendor LIKE :vendor "
                    "ORDER BY sid DESC LIMIT 1"
                ),
                {"vendor": vendor_label},
            )
            last = last_row.fetchone()
            last_balance = float(last[6]) if last and last[6] else 0.0

            # Get vendor_id — try comm_bank first, then broker_new
            vid_result = await db.execute(
                text(
                    "SELECT vendor_id FROM comm_bank WHERE vendor LIKE :vendor LIMIT 1"
                ),
                {"vendor": vendor_label},
            )
            vid_row = vid_result.fetchone()
            vendor_id = vid_row[0] if vid_row else None

            if not vendor_id:
                vid_result2 = await db.execute(
                    text(
                        "SELECT broker_code FROM broker_new WHERE vendor LIKE :vendor LIMIT 1"
                    ),
                    {"vendor": vendor_label},
                )
                vid_row2 = vid_result2.fetchone()
                vendor_id = vid_row2[0] if vid_row2 else ""

            await db.execute(
                text(
                    "INSERT INTO summary_payments (vendor, vendor_id, month, payment, balance) "
                    "VALUES (:vendor, :vendor_id, :month, :payment, :balance)"
                ),
                {
                    "vendor": vendor_label,
                    "vendor_id": vendor_id,
                    "month": curr_month_label,
                    "payment": -payment,
                    "balance": last_balance,
                },
            )
        total += 1

    await db.commit()
    await _log_user_action(
        db,
        uid,
        user_name,
        f"Uploaded payment summary for {curr_month_label} — {total} vendors",
    )
    await db.commit()
    return {"inserted": total, "month": curr_month_label}


# ---------------------------------------------------------------------------
# 5. ADJUSTMENTS
# ---------------------------------------------------------------------------


async def get_adjustments(db: AsyncSession) -> list:
    result = await db.execute(text("SELECT * FROM adjustments ORDER BY sid DESC"))
    return [dict(row) for row in result.mappings()]


async def add_adjustment(
    data: dict,
    uid: int,
    user_name: str,
    db: AsyncSession,
) -> dict:
    await db.execute(
        text(
            "INSERT INTO adjustments (vendor, month, owed, comments) "
            "VALUES (:vendor, :month, :owed, :comments)"
        ),
        {
            "vendor": data["vendor"],
            "month": data["month"],
            "owed": data["owed"],
            "comments": data.get("comments", ""),
        },
    )
    await db.commit()
    await _log_user_action(
        db,
        uid,
        user_name,
        f"Added adjustment for vendor={data['vendor']} month={data['month']} owed={data['owed']}",
    )
    await db.commit()
    return {"status": "created"}


async def delete_adjustment(
    sid: int,
    uid: int,
    user_name: str,
    db: AsyncSession,
) -> dict:
    await db.execute(text("DELETE FROM adjustments WHERE sid = :sid"), {"sid": sid})
    await db.commit()
    await _log_user_action(db, uid, user_name, f"Deleted adjustment sid={sid}")
    await db.commit()
    return {"status": "deleted"}


# ---------------------------------------------------------------------------
# 6. CALCULATE COMMISSION
# ---------------------------------------------------------------------------


async def calculate_commission(
    uid: int,
    user_name: str,
    db: AsyncSession,
    month: str = None,  # format '2026-04', defaults to current month
) -> dict:

    # Convert month param to required formats
    if month:
        try:
            dt = datetime.strptime(month, "%Y-%m")
            curr_month_label = dt.strftime("%b-%y")
            end_pattern = dt.strftime("%Y-%m-%%")
        except ValueError:
            curr_month_label = _current_month_short()
            end_pattern = _current_end_date_pattern()
    else:
        # Use most recently uploaded month from comm_bank
        latest_result = await db.execute(
            text("SELECT end_date FROM comm_bank ORDER BY sid DESC LIMIT 1")
        )
        latest_row = latest_result.fetchone()
        if latest_row and latest_row[0]:
            latest_end = str(latest_row[0])[:7]  # '2025-12'
            dt = datetime.strptime(latest_end, "%Y-%m")
            curr_month_label = dt.strftime("%b-%y")  # 'Dec-25'
            end_pattern = dt.strftime("%Y-%m-%%")  # '2025-12-%'
        else:
            curr_month_label = _current_month_short()
            end_pattern = _current_end_date_pattern()
    # Check payment summary exists for this month
    print(f"DEBUG calc: curr_month_label={curr_month_label}, end_pattern={end_pattern}")

    # Sum commission per vendor from comm_bank for current month
    result = await db.execute(
        text(
            "SELECT vendor, SUM(commission_amount) as total_comm "
            "FROM comm_bank WHERE end_date LIKE :end_pat AND broker_status = 1 "
            "GROUP BY vendor"
        ),
        {"end_pat": end_pattern},
    )
    sum_commission: dict = {}
    for row in result:
        sum_commission[str(row[0]).upper()] = round(float(row[1] or 0), 2)
    # After sum_commission built:
    print(f"DEBUG sum_commission: {sum_commission}")
    # Add vendors in summary_payments but not in comm_bank (owed=0)
    sp_vendors_result = await db.execute(
        text("SELECT DISTINCT vendor FROM summary_payments WHERE month = :month"),
        {"month": curr_month_label},
    )
    for row in sp_vendors_result:
        v = str(row[0]).upper()
        if v not in sum_commission:
            sum_commission[v] = 0.0
    updated = 0
    for vendor_key, owed in sum_commission.items():
        # Get current month row (inserted by payment upload)
        curr_row_result = await db.execute(
            text(
                "SELECT * FROM summary_payments "
                "WHERE vendor = :vendor AND month = :month "
                "ORDER BY sid DESC LIMIT 1"
            ),
            {"vendor": vendor_key, "month": curr_month_label},
        )
        curr_row = curr_row_result.fetchone()
        # After sp_vendors added:
        print(f"DEBUG all vendors to process: {list(sum_commission.keys())}")

        # Get last closing balance (most recent row BEFORE current month)
        prev_row_result = await db.execute(
            text(
                "SELECT balance FROM summary_payments "
                "WHERE vendor = :vendor AND month != :month "
                "ORDER BY sid DESC LIMIT 1"
            ),
            {"vendor": vendor_key, "month": curr_month_label},
        )
        prev_row = prev_row_result.fetchone()
        prev_balance = round(float(prev_row[0] or 0), 2) if prev_row else 0.0

        if curr_row:
            payment = round(float(curr_row[4] or 0), 2)
            new_balance = round(prev_balance + owed + payment, 2)
            await db.execute(
                text(
                    "UPDATE summary_payments "
                    "SET owed = :owed, balance = :balance "
                    "WHERE sid = :sid"
                ),
                {"owed": owed or 0, "balance": new_balance or 0, "sid": curr_row[0]},
            )
        else:
            # No payment row — insert with payment=0
            new_balance = round(prev_balance + owed, 2)
            await db.execute(
                text(
                    "INSERT INTO summary_payments "
                    "(vendor, month, owed, payment, balance) "
                    "VALUES (:vendor, :month, :owed, :payment, :balance)"
                ),
                {
                    "vendor": vendor_key,
                    "month": curr_month_label,
                    "owed": owed,
                    "payment": 0,
                    "balance": new_balance,
                },
            )
        updated += 1

    await db.commit()
    await _log_user_action(
        db,
        uid,
        user_name,
        f"Calculated commission for {curr_month_label} — {updated} vendors updated",
    )
    await db.commit()
    return {
        "status": "calculated",
        "vendors_updated": updated,
        "month": curr_month_label,
    }


# ---------------------------------------------------------------------------
# 7. REVIEW SUMMARY
# ---------------------------------------------------------------------------


async def get_review_summary(
    db: AsyncSession,
    vendor: Optional[str] = None,
    month: Optional[str] = None,
    full_history: bool = False,
) -> list:
    """
    Returns summary_payments rows for a vendor.
    Default: last 12 months. full_history=True returns all.
    Calculates audit_status per row by verifying balance chain.
    """
    if not vendor:
        # Return distinct vendors with their latest balance for overview
        result = await db.execute(
            text(
                """
                SELECT sp.vendor, sp.vendor_id, sp.month, sp.payment,
                       sp.owed, sp.balance, sp.comments, sp.status,
                       bn.company_name, bn.commission_email
                FROM summary_payments sp
                LEFT JOIN broker_new bn ON bn.broker_code = sp.vendor_id
                WHERE sp.sid IN (
                    SELECT MAX(sid) FROM summary_payments GROUP BY vendor
                )
                ORDER BY sp.vendor
                """
            )
        )
        return [dict(r) for r in result.mappings()]

    # Single vendor — full history or last 12 months
    if full_history:
        result = await db.execute(
            text(
                "SELECT * FROM summary_payments "
                "WHERE vendor = :vendor "
                "ORDER BY STR_TO_DATE(CONCAT('01-', month), '%d-%b-%y') ASC, sid ASC"
            ),
            {"vendor": vendor},
        )
    else:
        result = await db.execute(
            text(
                "SELECT * FROM summary_payments "
                "WHERE vendor = :vendor "
                "ORDER BY STR_TO_DATE(CONCAT('01-', month), '%d-%b-%y') ASC, sid ASC LIMIT 12"
            ),
            {"vendor": vendor},
        )

    rows = [dict(r) for r in result.mappings()]

    # rows already in ASC order (oldest first) from query
    rows_asc = rows  # no reversal needed

    # Audit status — verify balance chain (runs oldest to newest correctly)
    for i, row in enumerate(rows_asc):
        if i == 0:
            row["audit_status"] = "ok"
            continue
        prev = rows_asc[i - 1]
        try:
            prev_bal = round(float(prev["balance"] or 0), 2)
            owed = round(float(row["owed"] or 0), 2)
            payment = round(float(row["payment"] or 0), 2)
            expected = round(prev_bal + owed + payment, 2)
            actual = round(float(row["balance"] or 0), 2)
            row["audit_status"] = "ok" if abs(expected - actual) < 0.02 else "error"
        except (ValueError, TypeError):
            row["audit_status"] = "ok"

    return rows_asc  # oldest first — Dec-25 row 1, Jan-26 row 2


async def get_summary_months(db: AsyncSession) -> list:
    """Returns distinct months from summary_payments, newest first"""
    result = await db.execute(
        text(
            "SELECT DISTINCT month FROM summary_payments " "ORDER BY sid DESC LIMIT 36"
        )
    )
    # Deduplicate while preserving order (sid DESC handles recency)
    seen = set()
    months = []
    for row in result:
        m = row[0]
        if m not in seen:
            seen.add(m)
            months.append(m)
    return months


# ---------------------------------------------------------------------------
# 8. USER LOG
# ---------------------------------------------------------------------------


async def get_user_log(
    db: AsyncSession,
    limit: int = 500,
) -> list:
    result = await db.execute(
        text(
            "SELECT * FROM user_log WHERE flag = 'commission' "
            "ORDER BY sid DESC LIMIT :limit"
        ),
        {"limit": limit},
    )
    return [dict(row) for row in result.mappings()]


# ---------------------------------------------------------------------------
# 9. VENDORS DROPDOWN (for filters)
# ---------------------------------------------------------------------------


async def get_vendor_dropdown(db: AsyncSession) -> list:
    result = await db.execute(
        text(
            "SELECT DISTINCT cb.vendor, bn.company_name "
            "FROM comm_bank cb "
            "LEFT JOIN broker_new bn ON bn.vendor = cb.vendor "
            "WHERE cb.broker_status = 1 "
            "AND cb.vendor IS NOT NULL AND cb.vendor != '' "
            "ORDER BY cb.vendor"
        )
    )
    return [{"vendor": row[0], "company_name": row[1] or row[0]} for row in result]


async def get_months_dropdown(db: AsyncSession) -> list:
    result = await db.execute(
        text(
            "SELECT DISTINCT month, end_date "
            "FROM comm_bank "
            "ORDER BY end_date DESC LIMIT 200"
        )
    )
    seen = set()
    months = []
    for row in result:
        if not row[1]:
            continue
        year_month = str(row[1])[:7]  # '2026-04' from '2026-04-10'
        label = f"{row[0]} {year_month[:4]}"  # 'April 2026'
        if label not in seen:
            seen.add(label)
            months.append({"label": label, "value": year_month})
    return months


async def get_commission_exceptions(
    db: AsyncSession,
    month_pattern: str,  # e.g. '2026-04-%'
) -> dict:
    """
    Runs all exception checks against comm_bank for the given month.
    Returns dict with exception type as key, list of rows as value.
    """
    # Base query — get all rows for this month
    result = await db.execute(
        text(
            "SELECT * FROM comm_bank WHERE end_date LIKE :pat AND broker_status = 1 "
            "ORDER BY vendor, premise_id, service_start_date"
        ),
        {"pat": month_pattern},
    )
    all_rows = [dict(r) for r in result.mappings()]

    exceptions = {
        "duplicate": [],
        "variance_30": [],
        "inactive": [],
        "zero_commission": [],
        "negative_commission": [],
        "expired_contract": [],
        "rate_anomaly": [],
        "missing_data": [],
    }

    # -- Get last 6 months of data for duplicate/variance checks
    from datetime import datetime, timedelta

    now = datetime.now()
    six_months_ago = (now.replace(day=1) - timedelta(days=180)).strftime("%Y-%m-%%")
    hist_result = await db.execute(
        text(
            "SELECT * FROM comm_bank WHERE end_date LIKE :pat AND broker_status = 1 "
            "ORDER BY vendor, premise_id, service_start_date, end_date"
        ),
        {"pat": six_months_ago},
    )
    hist_rows = [dict(r) for r in hist_result.mappings()]

    # 1. DUPLICATE — same premise_id + service_start + service_end across months
    seen = {}
    for row in hist_rows:
        key = (
            str(row["premise_id"]),
            str(row["service_start_date"]),
            str(row["service_end_date"]),
        )
        if key in seen:
            prev = seen[key]
            diff_month = prev["month"] != row["month"]
            same_amt = prev["commission_amount"] == row["commission_amount"]
            if diff_month or same_amt:
                prev["exception_type"] = "duplicate"
                row["exception_type"] = "duplicate"
                if prev not in exceptions["duplicate"]:
                    exceptions["duplicate"].append(prev)
                exceptions["duplicate"].append(row)
        seen[key] = row

    # 2. VARIANCE 30% — commission changed >= 30% for same premise month over month
    premise_by_month = {}
    for row in hist_rows:
        key = str(row["premise_id"])
        if key not in premise_by_month:
            premise_by_month[key] = []
        premise_by_month[key].append(row)

    for premise_id, rows in premise_by_month.items():
        for i in range(1, len(rows)):
            curr = rows[i]
            prev = rows[i - 1]
            try:
                curr_amt = float(curr["commission_amount"] or 0)
                prev_amt = float(prev["commission_amount"] or 0)
            except (ValueError, TypeError):
                continue
            if prev_amt == 0 or (curr_amt <= 30 and prev_amt <= 30):
                continue
            pct = ((curr_amt - prev_amt) / abs(prev_amt)) * 100
            if abs(pct) >= 30:
                curr["exception_type"] = "variance_30"
                curr["variance_pct"] = round(pct, 2)
                # Only flag current month rows
                if curr in all_rows:
                    exceptions["variance_30"].append(curr)

    # 3. INACTIVE — active billing but customer status is I or G
    for row in all_rows:
        if str(row.get("cust_status", "")).upper() in ("I", "G"):
            row["exception_type"] = "inactive"
            exceptions["inactive"].append(row)

    # 4. ZERO COMMISSION — active customer, commission = 0
    for row in all_rows:
        if (
            str(row.get("cust_status", "")).upper() == "A"
            and float(row.get("commission_amount") or 0) == 0
        ):
            row["exception_type"] = "zero_commission"
            exceptions["zero_commission"].append(row)

    # 5. NEGATIVE COMMISSION
    for row in all_rows:
        try:
            if float(row.get("commission_amount") or 0) < 0:
                row["exception_type"] = "negative_commission"
                exceptions["negative_commission"].append(row)
        except (ValueError, TypeError):
            pass

    # 6. EXPIRED CONTRACT — contract_end_date in the past but still billed
    today = datetime.now().date()
    for row in all_rows:
        ced = row.get("contract_end_date")
        if not ced:
            continue
        try:
            exp_date = datetime.strptime(str(ced), "%Y-%m-%d").date()
            if exp_date < today:
                row["exception_type"] = "expired_contract"
                exceptions["expired_contract"].append(row)
        except (ValueError, TypeError):
            pass

    # 7. RATE ANOMALY — commission_rate outside 0.001 to 0.05
    # 7. RATE ANOMALY
    # rate >= 0.01 = red (too high)
    # rate <= 0.001 = yellow (too low)
    # rate == 0 = yellow (missing)
    for row in all_rows:
        try:
            rate = float(row.get("commission_rate") or 0)
            if rate >= 0.01:
                row["exception_type"] = "rate_anomaly"
                row["anomaly_level"] = "red"
                row["anomaly_reason"] = f"Rate {rate} is too high (>=0.01)"
                exceptions["rate_anomaly"].append(row)
            elif rate == 0 or rate <= 0.001:
                row["exception_type"] = "rate_anomaly"
                row["anomaly_level"] = "yellow"
                row["anomaly_reason"] = f"Rate {rate} is too low or zero"
                exceptions["rate_anomaly"].append(row)
        except (ValueError, TypeError):
            pass

    # 8. MISSING DATA — blank premise_id, commission_rate, or kwh_usage
    for row in all_rows:
        missing = []
        if not row.get("premise_id"):
            missing.append("premise_id")
        if not row.get("commission_rate"):
            missing.append("commission_rate")
        if not row.get("kwh_usage"):
            missing.append("kwh_usage")
        if missing:
            row["exception_type"] = "missing_data"
            row["missing_fields"] = missing
            exceptions["missing_data"].append(row)

    # Build summary counts
    summary = {k: len(v) for k, v in exceptions.items()}
    total = sum(summary.values())

    return {
        "summary": summary,
        "total": total,
        "exceptions": exceptions,
        "month": month_pattern.replace("-%", ""),
    }


async def add_manual_payment(
    vendor: str,
    amount: float,
    comments: str,
    entry_type: str,  # 'payment' or 'adjustment'
    uid: int,
    user_name: str,
    db: AsyncSession,
) -> dict:
    curr_month = _current_month_short()

    # Get last balance
    prev_row = await db.execute(
        text(
            "SELECT balance FROM summary_payments "
            "WHERE vendor = :vendor "
            "ORDER BY sid DESC LIMIT 1"
        ),
        {"vendor": vendor},
    )
    prev = prev_row.fetchone()
    prev_balance = round(float(prev[0] or 0), 2) if prev else 0.0
    new_balance = round(prev_balance + amount, 2)

    # Get vendor_id
    vid_result = await db.execute(
        text("SELECT vendor_id FROM summary_payments WHERE vendor = :vendor LIMIT 1"),
        {"vendor": vendor},
    )
    vid_row = vid_result.fetchone()
    vendor_id = vid_row[0] if vid_row else ""

    await db.execute(
        text(
            "INSERT INTO summary_payments "
            "(vendor, vendor_id, month, payment, owed, balance, comments) "
            "VALUES (:vendor, :vendor_id, :month, :payment, :owed, :balance, :comments)"
        ),
        {
            "vendor": vendor,
            "vendor_id": vendor_id,
            "month": curr_month,
            "payment": amount,
            "owed": 0,
            "balance": new_balance,
            "comments": f"[{entry_type.upper()}] {comments}",
        },
    )
    await db.commit()
    await _log_user_action(
        db,
        uid,
        user_name,
        f"Added manual {entry_type} for {vendor} amount={amount} month={curr_month}",
    )
    await db.commit()
    return {"status": "created", "month": curr_month, "balance": new_balance}


async def generate_commission_excel(
    vendor: str,
    month_label: str,  # e.g. 'Apr-26'
    db: AsyncSession,
) -> tuple[bytes, str]:
    """
    Generates commission Excel file for a broker with 3 sheets:
    - Summary: last 12 months from summary_payments
    - Commission Details: current month comm_bank rows
    - Commission Analysis: premise × month grid last 12 months
    Returns (file_bytes, filename)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Get broker info
    broker_res = await db.execute(
        text("SELECT * FROM broker_new WHERE vendor = :vendor LIMIT 1"),
        {"vendor": vendor},
    )
    broker = broker_res.mappings().fetchone()
    company_name = dict(broker)["company_name"] if broker else vendor

    # Get last 12 months summary
    summary_res = await db.execute(
        text(
            "SELECT * FROM summary_payments WHERE vendor = :vendor "
            "ORDER BY sid DESC LIMIT 12"
        ),
        {"vendor": vendor},
    )
    summary_rows = [dict(r) for r in summary_res.mappings()]
    summary_rows = list(reversed(summary_rows))

    # Get current month commission details
    details_res = await db.execute(
        text(
            "SELECT * FROM comm_bank WHERE vendor = :vendor "
            "AND month = :month AND broker_status = 1 "
            "ORDER BY premise_id"
        ),
        {"vendor": vendor, "month": _month_short_to_full(month_label)},
    )
    details_rows = [dict(r) for r in details_res.mappings()]

    # Get last 12 months for analysis grid
    analysis_res = await db.execute(
        text(
            "SELECT premise_id, company_name, service_start_date, service_end_date, "
            "month, commission_amount, end_date FROM comm_bank "
            "WHERE vendor = :vendor AND broker_status = 1 "
            "ORDER BY premise_id, end_date DESC LIMIT 5000"
        ),
        {"vendor": vendor},
    )
    analysis_rows = [dict(r) for r in analysis_res.mappings()]

    wb = Workbook()

    # ── Styles ──────────────────────────────────────────
    header_font = Font(name="Arial", bold=True, size=11)
    header_fill = PatternFill("solid", start_color="17365D", end_color="17365D")
    header_font_white = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    red_font = Font(name="Arial", color="FF0000", bold=True)
    green_font = Font(name="Arial", color="008000", bold=True)
    normal_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = header_fill
            cell.font = header_font_white
            cell.alignment = center
            cell.border = border

    def style_data_row(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = normal_font
            cell.border = border

    # ── Sheet 1: Summary ────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"

    # Title
    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"{company_name} — Commission Summary"
    ws1["A1"].font = Font(name="Arial", bold=True, size=14, color="17365D")
    ws1["A1"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A2:F2")
    ws1["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    ws1["A2"].font = Font(name="Arial", size=10, color="666666")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Month", "Payment", "Owed", "Balance", "Comments"]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=4, column=col, value=h)
    style_header_row(ws1, 4, len(headers))

    # Data
    for i, row in enumerate(summary_rows, 5):
        ws1.cell(row=i, column=1, value=row.get("month", ""))
        payment = float(row.get("payment") or 0)
        owed = float(row.get("owed") or 0)
        balance = float(row.get("balance") or 0)

        pay_cell = ws1.cell(row=i, column=2, value=payment)
        pay_cell.number_format = "$#,##0.00"
        pay_cell.font = (
            Font(name="Arial", size=10, color="CC0000") if payment < 0 else normal_font
        )

        owed_cell = ws1.cell(row=i, column=3, value=owed)
        owed_cell.number_format = "$#,##0.00"
        owed_cell.font = (
            Font(name="Arial", size=10, color="008000") if owed > 0 else normal_font
        )

        bal_cell = ws1.cell(row=i, column=4, value=balance)
        bal_cell.number_format = "$#,##0.00"
        bal_cell.font = (
            Font(name="Arial", size=10, color="CC0000") if balance < 0 else normal_font
        )

        ws1.cell(row=i, column=5, value=row.get("comments", ""))
        style_data_row(ws1, i, len(headers))

    # Totals row
    total_row = len(summary_rows) + 5
    ws1.cell(row=total_row, column=1, value="TOTAL").font = Font(
        name="Arial", bold=True
    )
    ws1.cell(row=total_row, column=2, value=f"=SUM(B5:B{total_row-1})")
    ws1.cell(row=total_row, column=2).number_format = "$#,##0.00"
    ws1.cell(row=total_row, column=3, value=f"=SUM(C5:C{total_row-1})")
    ws1.cell(row=total_row, column=3).number_format = "$#,##0.00"
    style_header_row(ws1, total_row, len(headers))

    # Column widths
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 40

    # ── Sheet 2: Commission Details ──────────────────────
    ws2 = wb.create_sheet("Commission Details")

    detail_headers = [
        "Vendor ID",
        "Vendor Name",
        "Customer ID",
        "Premise ID",
        "Bill No",
        "First Name",
        "Last Name",
        "Company Name",
        "Cust Status",
        "Service Start Date",
        "Service End Date",
        "Plan Group",
        "Contract Rate",
        "KWH Usage",
        "Commission Rate",
        "Commission Amount",
        "Month",
    ]
    detail_keys = [
        "vendor_id",
        "vendor_name",
        "customer_id",
        "premise_id",
        "bill_no",
        "first_name",
        "last_name",
        "company_name",
        "cust_status",
        "service_start_date",
        "service_end_date",
        "plan_group",
        "contract_rate",
        "kwh_usage",
        "commission_rate",
        "commission_amount",
        "month",
    ]

    for col, h in enumerate(detail_headers, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header_row(ws2, 1, len(detail_headers))

    for i, row in enumerate(details_rows, 2):
        for col, key in enumerate(detail_keys, 1):
            val = row.get(key, "")
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = normal_font
            cell.border = border
            if key == "commission_amount":
                cell.number_format = "$#,##0.0000"
            if key == "commission_rate":
                cell.number_format = "0.00000"

    # Column widths
    col_widths = [10, 20, 14, 22, 14, 12, 12, 28, 8, 14, 14, 10, 10, 10, 12, 14, 10]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Commission Analysis ────────────────────
    ws3 = wb.create_sheet("Commission Analysis")

    # Build month list from last 12 months
    month_order = []
    seen_months = set()
    for row in analysis_rows:
        pid = row.get("premise_id", "")
        # Build month label from end_date e.g. '2026-04-10' → 'Apr-26'
        end_date_val = str(row.get("end_date") or "")
        if end_date_val and len(end_date_val) >= 7:
            try:
                dt = datetime.strptime(end_date_val[:7], "%Y-%m")
                m = dt.strftime("%b-%y")
            except ValueError:
                m = row.get("month", "")
        else:
            m = row.get("month", "")
        if m and m not in seen_months:
            seen_months.add(m)
            month_order.append(m)
    month_order = month_order[:12]

    # Headers
    ws3.cell(row=1, column=1, value="Premise ID")
    ws3.cell(row=1, column=2, value="Company Name")
    ws3.cell(row=1, column=3, value="Service Start")
    ws3.cell(row=1, column=4, value="Service End")
    for col, m in enumerate(month_order, 5):
        ws3.cell(row=1, column=col, value=m)
    style_header_row(ws3, 1, 4 + len(month_order))

    # Build premise → month → commission lookup
    premise_data: dict = {}
    for row in analysis_rows:
        pid = row.get("premise_id", "")
        m = row.get("month", "")
        if not pid:
            continue
        if pid not in premise_data:
            premise_data[pid] = {
                "company_name": row.get("company_name", ""),
                "service_start_date": row.get("service_start_date", ""),
                "service_end_date": row.get("service_end_date", ""),
                "months": {},
            }
        if m not in premise_data[pid]["months"]:
            premise_data[pid]["months"][m] = 0
        premise_data[pid]["months"][m] = float(row.get("commission_amount") or 0)

    for i, (pid, pdata) in enumerate(premise_data.items(), 2):
        ws3.cell(row=i, column=1, value=pid)
        ws3.cell(row=i, column=2, value=pdata["company_name"])
        ws3.cell(row=i, column=3, value=str(pdata["service_start_date"] or ""))
        ws3.cell(row=i, column=4, value=str(pdata["service_end_date"] or ""))
        for col, m in enumerate(month_order, 5):
            val = pdata["months"].get(m)
            cell = ws3.cell(row=i, column=col, value=val)
            if val:
                cell.number_format = "$#,##0.0000"
            cell.font = normal_font
            cell.border = border

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 28
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 14
    for col in range(5, 5 + len(month_order)):
        ws3.column_dimensions[get_column_letter(col)].width = 12

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Filename: CompanyName_Month_Year.xlsx
    safe_name = company_name.replace(" ", "_").replace("/", "-")
    month_parts = month_label.split("-")
    month_full = (
        datetime.strptime(month_parts[0], "%b").strftime("%B")
        if month_parts
        else month_label
    )
    year_full = (
        f"20{month_parts[1]}" if len(month_parts) > 1 else datetime.now().strftime("%Y")
    )
    filename = f"{safe_name}_{month_full}_{year_full}.xlsx"

    return buf.read(), filename


def _month_short_to_full(month_short: str) -> str:
    """Convert 'Apr-26' → 'April'"""
    try:
        return datetime.strptime(month_short.split("-")[0], "%b").strftime("%B")
    except (ValueError, IndexError):
        return month_short


async def send_commission_emails(
    vendor_ids: list,
    month_label: str,
    uid: int,
    user_name: str,
    db: AsyncSession,
) -> dict:
    """
    Generates commission Excel per broker and emails it to commission_email
    where commission_flag = 1
    """
    from controllers.email_pricing import send_email_async, build_email_html
    from utils.email_routing import get_tenant_email, get_tenant_display_name

    commission_contact = get_tenant_email("commission")  # fail fast before touching any brokers
    _tenant_name = get_tenant_display_name()

    results = {"sent": [], "failed": []}
    curr_month_full = _month_short_to_full(month_label)

    # Get brokers — either specific list or all with commission_flag=1
    if vendor_ids:
        placeholders = ",".join([f":v{i}" for i in range(len(vendor_ids))])
        params = {f"v{i}": v for i, v in enumerate(vendor_ids)}
        brokers_res = await db.execute(
            text(
                f"SELECT * FROM broker_new WHERE vendor IN ({placeholders}) "
                "AND commission_flag = 1"
            ),
            params,
        )
    else:
        brokers_res = await db.execute(
            text(
                "SELECT * FROM broker_new WHERE commission_flag = 1 "
                "AND regular_status != 'inactive' ORDER BY company_name"
            )
        )
    brokers = [dict(r) for r in brokers_res.mappings()]

    for broker in brokers:
        try:
            commission_email = broker.get("commission_email", "")
            if not commission_email:
                results["failed"].append(
                    f"{broker['company_name']} — no commission email"
                )
                continue

            vendor = broker.get("vendor", "")
            if not vendor:
                results["failed"].append(f"{broker['company_name']} — no vendor code")
                continue

            # Generate Excel file
            excel_bytes, filename = await generate_commission_excel(
                vendor, month_label, db
            )

            logo_html = f'<img src="{TENANT_LOGO_URL}" style="height:60px" alt="{TENANT_DISPLAY_NAME}"/>' if TENANT_LOGO_URL else ""
            html = f"""
<html><body style='font-family:Arial,sans-serif;color:#333'>
    <table width='100%' style='max-width:800px;margin:auto'>
        <tr>
            <td style='padding:20px 0'>
                {logo_html}
            </td>
            <td style='text-align:right;font-size:18px;font-weight:bold;color:#DC2626'>
                Commission Statement
            </td>
        </tr>
        <tr>
            <td colspan='2' style='padding:10px 0;border-bottom:2px solid #DC2626'>
                <strong>Broker:</strong> {broker['company_name']} 
                &nbsp;&nbsp;&nbsp; 
                <strong>Period:</strong> {curr_month_full}
            </td>
        </tr>
        <tr>
            <td colspan='2' style='padding:20px 0;color:#333;font-family:Arial;font-size:13px'>
                <p>Dear {broker['company_name']},</p>
                <p>Please find attached your commission statement for <strong>{curr_month_full}</strong>.</p>
                <p>The attached Excel file contains:</p>
                <ul style='color:#333;line-height:1.8'>
                    <li><strong>Summary</strong> — Last 12 months payment history and running balance</li>
                    <li><strong>Commission Details</strong> — Detailed breakdown of all commissions for {curr_month_full}</li>
                    <li><strong>Commission Analysis</strong> — Month-by-month commission grid per premise</li>
                </ul>
                <p style='color:#666;font-size:12px'>
                    If you have any questions regarding your commission statement,
                    please contact us at <a href='mailto:{commission_contact}'>{commission_contact}</a>
                </p>
                <p>Thank you for your business.</p>
                <p><strong>{_tenant_name} Team</strong></p>
            </td>
        </tr>
    </table>
</body></html>
"""
            subject = (
                f"Commission Statement — {broker['company_name']} — {curr_month_full}"
            )

            # Send email
            await send_email_async(
                commission_email, subject, html,
                purpose="commission",
                attachment=excel_bytes,
                attachment_name=filename,
            )

            # Log to broker_logs
            await db.execute(
                text(
                    "INSERT INTO broker_logs "
                    "(broker_code, company_name, email_type, sent_to, status) "
                    "VALUES (:broker_code, :company_name, :email_type, :sent_to, :status)"
                ),
                {
                    "broker_code": broker["broker_code"],
                    "company_name": broker["company_name"],
                    "email_type": "commission",
                    "sent_to": commission_email,
                    "status": "sent",
                },
            )

            results["sent"].append(broker["company_name"])

        except Exception as e:
            import traceback

            print(f"Commission email failed for {broker.get('company_name')}: {e}")
            print(traceback.format_exc())
            results["failed"].append(f"{broker['company_name']} — {str(e)}")

    await db.commit()
    await _log_user_action(
        db,
        uid,
        user_name,
        f"Sent commission emails for {month_label} — "
        f"{len(results['sent'])} sent, {len(results['failed'])} failed",
    )
    await db.commit()
    return results


async def download_commission_file(
    vendor: str,
    month_label: str,
    db: AsyncSession,
) -> tuple[bytes, str]:
    """Generate and return commission Excel for manual download"""
    return await generate_commission_excel(vendor, month_label, db)


async def generate_payment_sheet(
    month: str,  # '2026-04'
    db: AsyncSession,
) -> tuple[bytes, str]:
    """
    Generates payment sheet Excel for upload back to system.
    Format matches payment summary upload format:
    Row 1: blank, Commission Month, Apr-26
    Row 2: blank
    Row 3: headers
    Row 4+: data — S.No, Vendor Code, Vendor Name, Balance (positive)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    # Convert month to label
    dt = datetime.strptime(month, "%Y-%m")
    month_label = dt.strftime("%b-%y")  # 'Apr-26'

    # Get all vendors with their latest balance for this month
    result = await db.execute(
        text(
            """
            SELECT sp.vendor, sp.balance,
                   bn.company_name
            FROM summary_payments sp
            LEFT JOIN broker_new bn ON bn.vendor = sp.vendor
            WHERE sp.month = :month
            AND sp.sid IN (
                SELECT MAX(sid) FROM summary_payments 
                WHERE month = :month GROUP BY vendor
            )
            ORDER BY sp.vendor
            """
        ),
        {"month": month_label},
    )
    rows = [dict(r) for r in result.mappings()]

    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Sheet"

    # Row 1: Commission Month header
    ws["A1"] = ""
    ws["B1"] = "Commission Month"
    ws["C1"] = month_label
    ws["B1"].font = Font(name="Arial", bold=True)
    ws["C1"].font = Font(name="Arial", bold=True, color="FF0000")

    # Row 2: blank
    ws.append([])

    # Row 3: headers
    headers = ["S.No", "Vendor Code", "Vendor Name", "Amount (USD)"]
    ws.append(headers)
    for col in range(1, 5):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(name="Arial", bold=True)
        cell.fill = PatternFill("solid", start_color="17365D", end_color="17365D")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Row 4+: data
    for i, row in enumerate(rows, 1):
        balance = float(row.get("balance") or 0)
        # Payment sheet shows positive balance (what broker is owed)
        amount = abs(balance) if balance < 0 else balance
        ws.append(
            [
                i,
                row.get("vendor", ""),
                row.get("company_name", "") or row.get("vendor", ""),
                round(amount, 2),
            ]
        )

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Payment_Sheet_{month_label.replace('-', '_')}.xlsx"
    return buf.read(), filename
