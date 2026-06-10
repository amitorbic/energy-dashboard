import json
import os
import re
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional

import pandas as pd
import shutil
from fastapi import HTTPException, UploadFile
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
import tempfile
from models.billing import BillingUploadLog, BillingExceptionLog, BillingEmailRecipient
from schemas.billing import BillingCommentSave, RecipientCreate
import requests as req
from bs4 import BeautifulSoup
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO
from email import encoders
import os

SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", 465))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
SMTP_FROM = os.getenv("SMTP_FROM", "AmeriPower Billing <billing@ameripower.com>")

# ── helpers ───────────────────────────────────────────────────────────────────


def _row(r: Any) -> dict:
    """Convert a billing_extract_raw row to a slim exception dict."""
    return {
        "cust_id": str(r.cust_id or ""),
        "bill_no": str(r.bill_no or ""),
        "company_name": str(r.company_name or ""),
        "cust_name": f"{r.cust_first_name or ''} {r.cust_last_name or ''}".strip(),
    }


def _jdump(lst: List[dict]) -> Optional[str]:
    """Return JSON string or None if list is empty."""
    return json.dumps(lst) if lst else None


def _safe_rate(energy_charge, metered_usage) -> str:
    try:
        e = float(energy_charge or 0)
        m = float(metered_usage or 0)
        return str(round(e / m, 4)) if m != 0 else "0"
    except:
        return "0"


# ── parse & load raw ──────────────────────────────────────────────────────────


async def parse_and_load(file: UploadFile, db: AsyncSession, uploaded_by: str) -> int:
    """
    Read the uploaded XLS/XLSX billing extract.
    Truncate billing_extract_raw + billing_extract.
    Insert every row into billing_extract_raw.
    Filter LMP Day-Ahead rows → billing_extract with status codes.
    Return upload_id.
    """
    suffix = os.path.splitext(file.filename)[1].lower()

    tmp_path = os.path.join(tempfile.gettempdir(), file.filename)
    with open(tmp_path, "wb") as f:
        f.write(await file.read())

    if suffix == ".xls":
        df_raw = pd.read_excel(tmp_path, header=0, dtype=str, engine="xlrd")
    else:
        df_raw = pd.read_excel(tmp_path, header=0, dtype=str, engine="openpyxl")

    df_raw = df_raw.fillna("")

    # Truncate both staging tables
    await db.execute(text("TRUNCATE TABLE billing_extract_raw"))
    await db.execute(text("TRUNCATE TABLE billing_extract"))
    await db.execute(
        text("DELETE FROM billing_exception_log WHERE upload_date = :d"),
        {"d": date.today()},
    )

    # Column mapping: PHP used letter columns A=0, B=1 … up to GV
    # We trust the header names from the actual file; fall back to positional if needed.
    cols = list(df_raw.columns)

    def col(idx: int) -> str:
        return cols[idx] if idx < len(cols) else ""

    clean_money = lambda v: re.sub(r"[$,()\s]", "", str(v))

    rows_inserted = 0
    for _, r in df_raw.iterrows():

        def v(col_name):
            return str(r.get(col_name, "") or "").strip()

        def vm(col_name):
            return clean_money(v(col_name))

        await db.execute(
            text("""
            INSERT INTO billing_extract_raw (
                bill_to_id, cust_type, cust_id, company_name,
                cust_first_name, cust_last_name, bill_no, bill_date,
                due_date, plan_group, bal_fwd_amount, curr_amount,
                tax_amount, adj_amount, pay_amount, due_amount,
                energy_charge, deposit_charges, passthru_charge,
                other_charge, kh_qty, metered_usage,
                city_tax, county_tax, gros_tax, mtacda_tax,
                pugra_tax, spdt_tax, spdt2_tax, state_tax,
                last_paid_date, geocode, bill_handling_code,
                service_start, service_end, auto_pay_type,
                bill_mode, load_profile, contract_type, cust_email,
                premise_id, premise_address2, premise_address3,
                premise_county, premise_city, premise_state, premise_zip,
                agent_code, energy_rate, passthrough_rate, no_of_contracts_billed
            ) VALUES (
                :bill_to_id, :cust_type, :cust_id, :company_name,
                :cust_first_name, :cust_last_name, :bill_no, :bill_date,
                :due_date, :plan_group, :bal_fwd_amount, :curr_amount,
                :tax_amount, :adj_amount, :pay_amount, :due_amount,
                :energy_charge, :deposit_charges, :passthru_charge,
                :other_charge, :kh_qty, :metered_usage,
                :city_tax, :county_tax, :gros_tax, :mtacda_tax,
                :pugra_tax, :spdt_tax, :spdt2_tax, :state_tax,
                :last_paid_date, :geocode, :bill_handling_code,
                :service_start, :service_end, :auto_pay_type,
                :bill_mode, :load_profile, :contract_type, :cust_email,
                :premise_id, :premise_address2, :premise_address3,
                :premise_county, :premise_city, :premise_state, :premise_zip,
                :agent_code, :energy_rate, :passthrough_rate, :no_of_contracts_billed
            )
        """),
            {
                "bill_to_id": v("Bill to ID"),
                "cust_type": v("Cust Type"),
                "cust_id": v("Cust ID"),
                "company_name": v("Company Name"),
                "cust_first_name": v("Cust First Name"),
                "cust_last_name": v("Cust Last Name"),
                "bill_no": v("Bill No"),
                "bill_date": v("Bill Date"),
                "due_date": v("Due Date"),
                "plan_group": v("Plan Group"),
                "bal_fwd_amount": vm("Bal Fwd Amount"),
                "curr_amount": vm("Curr Amount"),
                "tax_amount": vm("Tax Amount"),
                "adj_amount": vm("Adj Amount"),
                "pay_amount": vm("Pay Amount"),
                "due_amount": vm("Due Amount"),
                "energy_charge": vm("Energy Charge"),
                "deposit_charges": vm("Deposit Charge"),
                "passthru_charge": vm("Passthru Charge"),
                "other_charge": vm("Other Charge"),
                "kh_qty": v("Usage Qty"),
                "metered_usage": v("Metered Usage"),
                "city_tax": vm("City Tax"),
                "county_tax": vm("County Tax"),
                "gros_tax": vm("Gros Tax"),
                "mtacda_tax": vm("Mtacda Tax"),
                "pugra_tax": vm("Pugra Tax"),
                "spdt_tax": vm("Spdt Tax"),
                "spdt2_tax": vm("Spdt2 Tax"),
                "state_tax": vm("State Tax"),
                "last_paid_date": v("Last Paid Date"),
                "geocode": v("Geocode"),
                "bill_handling_code": v("Bill Handling Code"),
                "service_start": v("Service Start"),
                "service_end": v("Service End"),
                "auto_pay_type": v("Auto Pay Type"),
                "bill_mode": v("Bill Mode"),
                "load_profile": v("Load Profile"),
                "contract_type": v("Contract Type"),
                "cust_email": v("Cust Email"),
                "premise_id": v("Premise Id"),
                "premise_address2": v("Premise Address2"),
                "premise_address3": v("Premise Address3"),
                "premise_county": v("Premise County"),
                "premise_city": v("Premise City"),
                "premise_state": v("Premise State"),
                "premise_zip": v("Premise Zip"),
                "agent_code": v("Agent Code"),
                "energy_rate": v("Energy Rate"),
                "passthrough_rate": v("Passthrough Rate"),
                "no_of_contracts_billed": v("No Of Contracts Billed"),
            },
        )
        rows_inserted += 1

    await db.commit()

    # Run all 36 checks and build exception + comment rows
    exceptions = await _run_all_checks(db)

    # Log the upload
    today = date.today()
    await db.execute(
        text("""
        INSERT INTO billing_upload_log (upload_date, filename, uploaded_by, total_rows, email_sent)
        VALUES (:upload_date, :filename, :uploaded_by, :total_rows, 0)
        ON DUPLICATE KEY UPDATE
        filename = VALUES(filename),
        uploaded_by = VALUES(uploaded_by),
        total_rows = VALUES(total_rows),
        email_sent = 0
       """),
        {
            "upload_date": today,
            "filename": file.filename,
            "uploaded_by": uploaded_by,
            "total_rows": rows_inserted,
        },
    )
    await db.flush()
    log_row = (
        await db.execute(
            text("SELECT id FROM billing_upload_log WHERE upload_date = :d"),
            {"d": today},
        )
    ).fetchone()
    upload_id = log_row.id

    # Insert exception row
    await db.execute(
        text("""
    INSERT INTO billing_exception_log 
    (upload_id, upload_date, row_type, check_tax_zero, check_kh_qty_energy_zero,
    check_kh_qty_metered_mismatch, check_residential_puc_grt_city, check_residential_tax_exempt,
    check_mcpe_bills, check_lmp_rate_range, check_sub_only_no_master, check_commercial_tdsp,
    check_residential_price_low, check_residential_price_high, check_commercial_price_high,
    check_commercial_price_low, check_negative_balance, check_zero_usage, check_partial_payment,
    check_zero_meter_fee, check_first_bill, check_final_bill, check_master_sub_final,
    check_state_tax_100, check_credit_card_final, check_autopay_balance, check_wrong_meter_fee,
    check_renewal_energy_high, check_paid_amount_variance, check_single_bill_under_100,
    check_multi_contract_invoice, check_old_autopay_balance, check_deposit_charges,
    check_first_bill_going_final, check_potential_final, check_difference_one_day,
    check_different_due_date, check_master_sub_autopay_type, check_master_sub_bill_mode)
    VALUES (:upload_id, :upload_date, :row_type, :check_tax_zero, :check_kh_qty_energy_zero,
    :check_kh_qty_metered_mismatch, :check_residential_puc_grt_city, :check_residential_tax_exempt,
    :check_mcpe_bills, :check_lmp_rate_range, :check_sub_only_no_master, :check_commercial_tdsp,
    :check_residential_price_low, :check_residential_price_high, :check_commercial_price_high,
    :check_commercial_price_low, :check_negative_balance, :check_zero_usage, :check_partial_payment,
    :check_zero_meter_fee, :check_first_bill, :check_final_bill, :check_master_sub_final,
    :check_state_tax_100, :check_credit_card_final, :check_autopay_balance, :check_wrong_meter_fee,
    :check_renewal_energy_high, :check_paid_amount_variance, :check_single_bill_under_100,
    :check_multi_contract_invoice, :check_old_autopay_balance, :check_deposit_charges,
    :check_first_bill_going_final, :check_potential_final, :check_difference_one_day,
    :check_different_due_date, :check_master_sub_autopay_type, :check_master_sub_bill_mode)
    """),
        {
            "upload_id": upload_id,
            "upload_date": today,
            "row_type": "exception",
            **exceptions,
        },
    )

    await db.execute(
        text("""
      INSERT INTO billing_exception_log 
       (upload_id, upload_date, row_type)
      VALUES (:upload_id, :upload_date, :row_type)
      """),
        {
            "upload_id": upload_id,
            "upload_date": today,
            "row_type": "comment",
        },
    )

    await db.commit()
    save_dir = os.path.join(tempfile.gettempdir(), "billing_extracts")
    os.makedirs(save_dir, exist_ok=True)
    shutil.copy(tmp_path, os.path.join(save_dir, file.filename))
    os.remove(tmp_path)
    print(f"Saved billing extract to: {os.path.join(save_dir, file.filename)}")
    await compare_with_php(upload_id, today, db)
    return upload_id


# ── 36 checks ─────────────────────────────────────────────────────────────────
async def send_billing_email(upload_id: int, db: AsyncSession):
    # get upload log
    log = (
        await db.execute(
            text("SELECT * FROM billing_upload_log WHERE id = :id"), {"id": upload_id}
        )
    ).fetchone()
    if not log:
        raise HTTPException(status_code=404, detail="Upload not found")

    # get exception + comment rows
    rows = (
        await db.execute(
            text(
                "SELECT * FROM billing_exception_log WHERE upload_id = :id ORDER BY row_type ASC"
            ),
            {"id": upload_id},
        )
    ).fetchall()

    exc = next((dict(r._mapping) for r in rows if r.row_type == "exception"), {})
    com = next((dict(r._mapping) for r in rows if r.row_type == "comment"), {})
    raw_rows = (await db.execute(text("SELECT * FROM billing_extract_raw"))).fetchall()
    attachments = _build_attachments(exc, log.upload_date, raw_rows)

    # get active recipients
    recipients = (
        await db.execute(
            text("SELECT email FROM billing_email_recipients WHERE active = 1")
        )
    ).fetchall()
    if not recipients:
        raise HTTPException(status_code=400, detail="No active email recipients")

    to_list = [r.email for r in recipients]

    # build email body — same format as the original system
    body = _build_email_body(exc, com, log.upload_date)
    save_dir = os.path.join(tempfile.gettempdir(), "billing_extracts")
    saved_path = os.path.join(save_dir, log.filename)
    print(f"Looking for file at: {saved_path}")
    print(f"File exists: {os.path.exists(saved_path)}")
    file_attachment = []
    if os.path.exists(saved_path):
        with open(saved_path, "rb") as f:
            file_attachment = [(log.filename, f.read())]
    attachments = file_attachment + _build_attachments(exc, log.upload_date)

    # send to all recipients
    # attachments = _build_attachments(exc, log.upload_date)
    print(f"Total attachments: {len(attachments)}")
    for name, data in attachments:
        print(f"  - {name}: {len(data)} bytes")
    _send_smtp(
        to=", ".join(to_list),
        subject=f"System Generated Billing Extract of {log.upload_date}",
        body=body,
        attachments=attachments,
    )

    # mark email sent
    await db.execute(
        text("UPDATE billing_upload_log SET email_sent = 1 WHERE id = :id"),
        {"id": upload_id},
    )
    await db.commit()

    return {"message": "Email sent successfully"}


def _build_email_body(exc: dict, com: dict, upload_date) -> str:
    CHECKS_META = [
        (
            "check_tax_zero",
            "1",
            "Tax amount is ZERO",
            ["Cust ID", "Bill No", "Company Name", "Customer Name"],
        ),
        (
            "check_kh_qty_energy_zero",
            "2",
            "KH Qty not zero and energy charge Zero",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Energy Charge",
                "KH Qty",
            ],
        ),
        (
            "check_kh_qty_metered_mismatch",
            "3",
            "KH Qty and metered usage does not match",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "KH Qty",
                "Metered Usage",
            ],
        ),
        (
            "check_residential_puc_grt_city",
            "4",
            "Residential: 100 under PUC, GRT, City tax",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Gross Tax",
                "Pugra Tax",
            ],
        ),
        (
            "check_residential_tax_exempt",
            "5",
            "Residential: exempt from all taxes",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Gross Tax",
                "Pugra Tax",
            ],
        ),
        (
            "check_mcpe_bills",
            "6",
            "Filter MCPE bills",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Service Start",
                "Service End",
            ],
        ),
        (
            "check_lmp_rate_range",
            "7",
            "LMP rate > 8c or < 4c/kWh",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Computed Rate"],
        ),
        (
            "check_sub_only_no_master",
            "8",
            "Sub Only accounts with no Master",
            ["Cust ID", "Bill No", "Company Name", "Customer Name"],
        ),
        (
            "check_commercial_tdsp",
            "9",
            "Commercial TDSP < 30% of Energy Charges",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Energy Charge",
                "Passthru Charge",
            ],
        ),
        (
            "check_residential_price_low",
            "10",
            "Residential price < 7.50c/kWh",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Computed Rate"],
        ),
        (
            "check_residential_price_high",
            "11",
            "Residential price > 15c/kWh",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Computed Rate"],
        ),
        (
            "check_commercial_price_high",
            "12",
            "Commercial > 13c/kWh",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Computed Rate"],
        ),
        (
            "check_commercial_price_low",
            "13",
            "Commercial < 3.6c/kWh",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Computed Rate"],
        ),
        (
            "check_negative_balance",
            "14",
            "Negative or low total balance",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Curr Amount",
                "Due Amount",
            ],
        ),
        (
            "check_zero_usage",
            "15",
            "Zero usage customers",
            ["Cust ID", "Bill No", "Company Name", "Customer Name"],
        ),
        (
            "check_partial_payment",
            "16",
            "Partial payment customers",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Pay Amount",
                "Curr Amount",
            ],
        ),
        (
            "check_zero_meter_fee",
            "17",
            "Zero meter fees",
            ["Cust ID", "Bill No", "Company Name", "Customer Name"],
        ),
        (
            "check_first_bill",
            "18",
            "First bill",
            ["Cust ID", "Bill No", "Company Name", "Customer Name"],
        ),
        (
            "check_final_bill",
            "19",
            "Final bill",
            ["Cust ID", "Bill No", "Company Name", "Customer Name"],
        ),
        (
            "check_master_sub_final",
            "20",
            "Master/Sub account final billed",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Cust Type"],
        ),
        (
            "check_state_tax_100",
            "21",
            "State tax = 100",
            ["Cust ID", "Bill No", "Company Name", "Customer Name"],
        ),
        (
            "check_credit_card_final",
            "22",
            "Credit card final bill service fee",
            ["Cust ID", "Bill No", "Company Name", "Customer Name"],
        ),
        (
            "check_autopay_balance",
            "23",
            "Auto pay customer with balance",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Due Amount",
                "Auto Pay Type",
            ],
        ),
        (
            "check_wrong_meter_fee",
            "24",
            "Wrong meter fee",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Other Charge"],
        ),
        (
            "check_renewal_energy_high",
            "25",
            "Renewal energy charges >= 0.13",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Computed Rate"],
        ),
        (
            "check_paid_amount_variance",
            "26",
            "Paid amount variance > 80%",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Curr Amount",
                "Pay Amount",
            ],
        ),
        (
            "check_single_bill_under_100",
            "27",
            "Single bills under $100",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Curr Amount"],
        ),
        (
            "check_multi_contract_invoice",
            "28",
            "Invoice billed with 2+ contracts",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "No Of Contracts"],
        ),
        (
            "check_old_autopay_balance",
            "29",
            "Old auto pay customer with balance",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Last Paid Date"],
        ),
        (
            "check_deposit_charges",
            "30",
            "Deposit charges",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Deposit Charges"],
        ),
        (
            "check_first_bill_going_final",
            "31",
            "First bill going final",
            ["Cust ID", "Bill No", "Company Name", "Customer Name"],
        ),
        (
            "check_potential_final",
            "32",
            "Potential final billing",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Service End"],
        ),
        (
            "check_difference_one_day",
            "33",
            "Difference 1 day",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Service Start",
                "Service End",
            ],
        ),
        (
            "check_different_due_date",
            "34",
            "Different due date",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Bill Date",
                "Due Date",
            ],
        ),
        (
            "check_master_sub_autopay_type",
            "35",
            "Master/Sub different Auto Pay Type",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Master Auto Pay",
                "Sub Auto Pay",
            ],
        ),
        (
            "check_master_sub_bill_mode",
            "36",
            "Master/Sub different Bill Mode",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Master Bill Mode",
                "Sub Bill Mode",
            ],
        ),
    ]

    KEY_MAP = {
        "Cust ID": "cust_id",
        "Bill No": "bill_no",
        "Company Name": "company_name",
        "Customer Name": "cust_name",
        "Energy Charge": "energy_charge",
        "KH Qty": "kh_qty",
        "Metered Usage": "metered_usage",
        "Gross Tax": "gros_tax",
        "Pugra Tax": "pugra_tax",
        "Service Start": "service_start",
        "Service End": "service_end",
        "Computed Rate": "computed_rate",
        "Passthru Charge": "passthru_charge",
        "Curr Amount": "curr_amount",
        "Due Amount": "due_amount",
        "Pay Amount": "pay_amount",
        "Other Charge": "other_charge",
        "Cust Type": "cust_type",
        "Auto Pay Type": "auto_pay_type",
        "No Of Contracts": "no_of_contracts_billed",
        "Last Paid Date": "last_paid_date",
        "Deposit Charges": "deposit_charges",
        "Bill Date": "bill_date",
        "Bill Handling Code": "bill_handling_code",
        "Master Auto Pay": "master_auto_pay",
        "Sub Auto Pay": "sub_auto_pay",
        "Master Bill Mode": "master_bill_mode",
        "Sub Bill Mode": "sub_bill_mode",
    }

    html = f"""
    <html><body style="font-family:Arial,sans-serif;font-size:13px;color:#222;max-width:900px;margin:0 auto;">
    <div style="background:#1F4E79;padding:16px 24px;border-radius:6px 6px 0 0;">
      <h2 style="color:#fff;margin:0;font-size:18px;">AmeriPower — Billing Exception Report</h2>
      <p style="color:#B8D4F0;margin:4px 0 0;font-size:13px;">Upload date: {upload_date}</p>
    </div>
    <div style="padding:16px 0;">
    """

    for key, num, title, cols in CHECKS_META:
        data = exc.get(key)
        rows_list = json.loads(data) if data else []
        comment = com.get(key, "")
        has_data = len(rows_list) > 0

        status_color = "#C00000" if has_data else "#375623"
        status_bg = "#FFE7E7" if has_data else "#E7F4E4"
        status_text = (
            f"{len(rows_list)} exception{'s' if len(rows_list) != 1 else ''}"
            if has_data
            else "No exceptions"
        )

        html += f"""
        <div style="margin-bottom:18px;border:1px solid #D0D7E3;border-radius:5px;overflow:hidden;">
          <div style="background:#2E6DA4;padding:8px 14px;display:flex;justify-content:space-between;align-items:center;">
            <span style="color:#fff;font-weight:bold;font-size:13px;">#{num} — {title}</span>
            <span style="background:{status_bg};color:{status_color};padding:2px 10px;border-radius:12px;font-size:12px;font-weight:bold;">{status_text}</span>
          </div>
        """

        if has_data:
            html += """<div style="overflow-x:auto;"><table style="width:100%;border-collapse:collapse;font-size:12px;">"""
            html += "<tr style='background:#EBF3FB;'>"
            html += "<th style='border:1px solid #C5D8EE;padding:5px 8px;text-align:left;color:#1F4E79;'>#</th>"
            for col in cols:
                html += f"<th style='border:1px solid #C5D8EE;padding:5px 8px;text-align:left;color:#1F4E79;'>{col}</th>"
            html += "</tr>"

            for i, row in enumerate(rows_list):
                bg = "#FFFFFF" if i % 2 == 0 else "#F5F9FE"
                html += f"<tr style='background:{bg};'>"
                html += f"<td style='border:1px solid #D9E5F3;padding:4px 8px;color:#555;'>{i+1}</td>"
                for col in cols:
                    val = row.get(KEY_MAP.get(col, col.lower().replace(" ", "_")), "")
                    html += f"<td style='border:1px solid #D9E5F3;padding:4px 8px;'>{val}</td>"
                html += "</tr>"
            html += "</table></div>"

        if comment:
            html += f"<div style='padding:8px 14px;background:#FFFDE7;border-top:1px solid #D0D7E3;font-size:12px;color:#555;'><b>Comment:</b> {comment}</div>"

        html += "</div>"

    html += """
    </div>
    <div style="background:#F0F4FA;padding:12px 24px;border-top:2px solid #2E6DA4;font-size:12px;color:#888;border-radius:0 0 6px 6px;">
      AmeriPower Automated Billing System — This is a system generated email.
    </div>
    </body></html>
    """
    return html


def _send_smtp(to: str, subject: str, body: str, attachments: list = []):
    msg = MIMEMultipart("mixed")
    msg["From"] = SMTP_FROM
    msg["To"] = to
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "html"))

    for filename, data in attachments:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(data)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={filename}")
        msg.attach(part)

    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_FROM, to.split(", "), msg.as_string())
    print("SMTP sendmail completed successfully")


def _make_excel(title: str, headers: list, rows: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 16)

    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _run_all_checks(db: AsyncSession) -> Dict[str, Optional[str]]:
    result = {}

    # ── Check 1: Tax amount is ZERO ──────────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name "
                "FROM billing_extract_raw WHERE tax_amount = 0 OR tax_amount IS NULL"
            )
        )
    ).fetchall()
    result["check_tax_zero"] = _jdump([_row(r) for r in rows])

    # ── Check 2: KH Qty not zero AND energy charge zero ──────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "kh_qty, metered_usage "
                "FROM billing_extract_raw "
                "WHERE kh_qty != '' AND metered_usage != '' "
                "AND kh_qty != '0' AND metered_usage != '0' "
                "AND ROUND(CAST(kh_qty AS DECIMAL(15,2)), 2) <> ROUND(CAST(metered_usage AS DECIMAL(15,2)), 2)"
            )
        )
    ).fetchall()
    result["check_kh_qty_energy_zero"] = _jdump(
        [
            {**_row(r), "energy_charge": str(r.energy_charge), "kh_qty": str(r.kh_qty)}
            for r in rows
        ]
    )

    # ── Check 3: KH Qty does not match metered usage ──────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "kh_qty, metered_usage FROM billing_extract_raw "
                "WHERE kh_qty != '' AND metered_usage != '' "
                "AND kh_qty != '0' AND metered_usage != '0' "
                "AND CAST(kh_qty AS DECIMAL(15,2)) <> CAST(metered_usage AS DECIMAL(15,2))"
            )
        )
    ).fetchall()
    result["check_kh_qty_metered_mismatch"] = _jdump(
        [
            {**_row(r), "kh_qty": str(r.kh_qty), "metered_usage": str(r.metered_usage)}
            for r in rows
        ]
    )

    # ── Check 4: Residential — 100 under PUC, GRT, City tax ──────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "gros_tax, pugra_tax, city_tax, plan_group "
                "FROM billing_extract_raw "
                "WHERE (gros_tax = '0' OR gros_tax = '' OR gros_tax IS NULL) "
                "OR (pugra_tax = '0' OR pugra_tax = '' OR pugra_tax IS NULL)"
            )
        )
    ).fetchall()
    result["check_residential_puc_grt_city"] = _jdump(
        [
            {
                **_row(r),
                "gros_tax": str(r.gros_tax),
                "pugra_tax": str(r.pugra_tax),
                "city_tax": str(r.city_tax),
                "plan_group": str(r.plan_group),
            }
            for r in rows
        ]
    )

    # ── Check 5: Residential — exempt from all taxes ──────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "gros_tax_exempt, pugra_tax_exempt FROM billing_extract_raw "
                "WHERE pugra_tax_exempt = '100'"
            )
        )
    ).fetchall()
    result["check_residential_tax_exempt"] = _jdump(
        [
            {
                **_row(r),
                "gros_tax": str(r.gros_tax_exempt),
                "pugra_tax": str(r.pugra_tax_exempt),
            }
            for r in rows
        ]
    )

    # ── Check 6: MCPE bills — billing period start date ───────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "contract_type FROM billing_extract_raw "
                "WHERE contract_type = 'LMP Day-Ahead'"
            )
        )
    ).fetchall()
    result["check_mcpe_bills"] = _jdump(
        [
            {
                **_row(r),
                "service_start": str(r.service_start),
                "service_end": str(r.service_end),
            }
            for r in rows
        ]
    )

    # ── Check 7: LMP rate > 8c or < 4c/kWh ───────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "energy_charge, metered_usage FROM billing_extract_raw "
                "WHERE contract_type = 'LMP Day-Ahead' "
                "AND metered_usage != '' AND metered_usage != '0' "
                "AND (CAST(energy_charge AS DECIMAL(15,4)) / CAST(metered_usage AS DECIMAL(15,4)) < 0.04 "
                "OR CAST(energy_charge AS DECIMAL(15,4)) / CAST(metered_usage AS DECIMAL(15,4)) > 0.08)"
            )
        )
    ).fetchall()
    result["check_lmp_rate_range"] = _jdump(
        [
            {
                **_row(r),
                "computed_rate": _safe_rate(r.energy_charge, r.metered_usage),
            }
            for r in rows
        ]
    )

    # ── Check 8: Sub Only accounts with no Master ─────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name "
                "FROM billing_extract_raw WHERE bill_mode = 'SubOnly'"
            )
        )
    ).fetchall()
    result["check_sub_only_no_master"] = _jdump([_row(r) for r in rows])

    # ── Check 9: Commercial TDSP < 30% of energy charge ──────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "energy_charge, metered_usage FROM billing_extract_raw "
                "WHERE contract_type = 'LMP Day-Ahead' "
                "AND metered_usage != '' AND metered_usage != '0' "
                "AND (CAST(energy_charge AS DECIMAL(15,4)) / CAST(metered_usage AS DECIMAL(15,4)) < 0.04 "
                "OR CAST(energy_charge AS DECIMAL(15,4)) / CAST(metered_usage AS DECIMAL(15,4)) > 0.08)"
            )
        )
    ).fetchall()
    result["check_commercial_tdsp"] = _jdump(
        [
            {
                **_row(r),
                "energy_charge": str(r.energy_charge),
                "passthru_charge": str(r.passthru_charge),
            }
            for r in rows
        ]
    )

    # ── Check 10: Residential price < 7.50c/kWh ──────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "energy_charge, passthru_charge, metered_usage FROM billing_extract_raw "
                "WHERE plan_group = 'R1' "
                "AND metered_usage != '' AND metered_usage != '0' "
                "AND (CAST(energy_charge AS DECIMAL(15,4)) + CAST(passthru_charge AS DECIMAL(15,4))) "
                "/ CAST(metered_usage AS DECIMAL(15,4)) < 0.075"
            )
        )
    ).fetchall()
    result["check_residential_price_low"] = _jdump(
        [
            {
                **_row(r),
                "computed_rate": _safe_rate(
                    float(r.energy_charge or 0) + float(r.passthru_charge or 0),
                    r.metered_usage,
                ),
            }
            for r in rows
        ]
    )

    # ── Check 11: Residential price > 15c/kWh ────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "energy_charge, passthru_charge, metered_usage FROM billing_extract_raw "
                "WHERE plan_group = 'R1' "
                "AND metered_usage != '' AND metered_usage != '0' "
                "AND (CAST(energy_charge AS DECIMAL(15,4)) + CAST(passthru_charge AS DECIMAL(15,4))) "
                "/ CAST(metered_usage AS DECIMAL(15,4)) > 0.15"
            )
        )
    ).fetchall()
    result["check_residential_price_high"] = _jdump(
        [
            {
                **_row(r),
                "computed_rate": _safe_rate(
                    float(r.energy_charge or 0) + float(r.passthru_charge or 0),
                    r.metered_usage,
                ),
            }
            for r in rows
        ]
    )

    # ── Check 12: Commercial > 13c/kWh ───────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "energy_charge, kh_qty FROM billing_extract_raw "
                "WHERE plan_group IN ('C1', 'C3') "
                "AND kh_qty != '' AND kh_qty != '0' "
                "AND CAST(energy_charge AS DECIMAL(15,4)) / CAST(kh_qty AS DECIMAL(15,4)) >= 0.13"
            )
        )
    ).fetchall()
    result["check_commercial_price_high"] = _jdump(
        [
            {**_row(r), "computed_rate": _safe_rate(r.energy_charge, r.kh_qty)}
            for r in rows
        ]
    )

    # ── Check 13: Commercial < 3.6c/kWh ──────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "energy_charge, kh_qty FROM billing_extract_raw "
                "WHERE plan_group IN ('C1', 'C3') "
                "AND kh_qty != '' AND kh_qty != '0' "
                "AND CAST(energy_charge AS DECIMAL(15,4)) / CAST(kh_qty AS DECIMAL(15,4)) < 0.036"
            )
        )
    ).fetchall()
    result["check_commercial_price_low"] = _jdump(
        [
            {**_row(r), "computed_rate": _safe_rate(r.energy_charge, r.kh_qty)}
            for r in rows
        ]
    )

    # ── Check 14: Negative balance or balance < current charge ───────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "energy_charge, passthru_charge, curr_amount, tax_amount, due_amount, bill_handling_code "
                "FROM billing_extract_raw "
                "WHERE curr_amount != '' AND tax_amount != '' AND due_amount != '' "
                "AND ROUND(CAST(curr_amount AS DECIMAL(15,4)) + CAST(tax_amount AS DECIMAL(15,4)), 2) "
                "!= CAST(due_amount AS DECIMAL(15,4)) "
                "AND CAST(curr_amount AS DECIMAL(15,4)) + CAST(tax_amount AS DECIMAL(15,4)) "
                "> CAST(due_amount AS DECIMAL(15,4))"
            )
        )
    ).fetchall()
    result["check_negative_balance"] = _jdump(
        [
            {
                **_row(r),
                "curr_amount": str(r.curr_amount),
                "tax_amount": str(r.tax_amount),
                "due_amount": str(r.due_amount),
                "computed": str(
                    round(float(r.curr_amount or 0) + float(r.tax_amount or 0), 2)
                ),
            }
            for r in rows
        ]
    )

    # ── Check 15: Zero usage ──────────────────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "metered_usage, kh_qty "
                "FROM billing_extract_raw WHERE metered_usage = 0 OR metered_usage IS NULL"
            )
        )
    ).fetchall()
    result["check_zero_usage"] = _jdump([_row(r) for r in rows])

    # ── Check 16: Partial payment ─────────────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "bal_fwd_amount, pay_amount, due_amount "
                "FROM billing_extract_raw "
                "WHERE bal_fwd_amount != '' AND bal_fwd_amount != '0' "
                "AND pay_amount != '' AND pay_amount != '0' "
                "AND ROUND(CAST(pay_amount AS DECIMAL(15,4)) / CAST(bal_fwd_amount AS DECIMAL(15,4)), 2) > 0 "
                "AND ROUND(CAST(pay_amount AS DECIMAL(15,4)) / CAST(bal_fwd_amount AS DECIMAL(15,4)), 2) <= 0.75"
            )
        )
    ).fetchall()
    result["check_partial_payment"] = _jdump(
        [
            {
                **_row(r),
                "bal_fwd_amount": str(r.bal_fwd_amount),
                "pay_amount": str(r.pay_amount),
                "ratio": str(
                    round(float(r.pay_amount or 0) / float(r.bal_fwd_amount or 1), 2)
                ),
            }
            for r in rows
        ]
    )

    # ── Check 17: Zero meter fees ─────────────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name "
                "FROM billing_extract_raw "
                "WHERE other_charge = '0' OR other_charge = '' OR other_charge IS NULL"
            )
        )
    ).fetchall()
    result["check_zero_meter_fee"] = _jdump([_row(r) for r in rows])

    # ── Check 18: First bill ──────────────────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "bill_handling_code "
                "FROM billing_extract_raw WHERE bill_handling_code = '1'"
            )
        )
    ).fetchall()
    result["check_first_bill"] = _jdump([_row(r) for r in rows])

    # ── Check 19: Final bill ──────────────────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "bill_handling_code "
                "FROM billing_extract_raw WHERE bill_handling_code = '9999'"
            )
        )
    ).fetchall()
    result["check_final_bill"] = _jdump([_row(r) for r in rows])

    # ── Check 20: Master/Sub final billed ────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, cust_type "
                "FROM billing_extract_raw "
                "WHERE cust_type IN ('Master','Sub') AND bill_handling_code = '9999'"
            )
        )
    ).fetchall()
    result["check_master_sub_final"] = _jdump(
        [{**_row(r), "cust_type": str(r.cust_type)} for r in rows]
    )

    # ── Check 21: State tax = 100 ─────────────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, state_tax, load_profile "
                "FROM billing_extract_raw WHERE state_tax = '0' OR state_tax IS NULL OR state_tax = ''"
            )
        )
    ).fetchall()
    result["check_state_tax_100"] = _jdump(
        [
            {
                **_row(r),
                "state_tax": str(r.state_tax),
                "load_profile": str(r.load_profile),
            }
            for r in rows
        ]
    )

    # ── Check 22: Credit card final bill service fee ──────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "auto_pay_type, bill_handling_code "
                "FROM billing_extract_raw "
                "WHERE auto_pay_type = 'CreditCard' AND bill_handling_code = '9999'"
            )
        )
    ).fetchall()
    result["check_credit_card_final"] = _jdump([_row(r) for r in rows])

    # ── Check 23: Auto pay customer with balance ──────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "auto_pay_type, curr_amount, tax_amount, due_amount "
                "FROM billing_extract_raw "
                "WHERE auto_pay_type IS NOT NULL AND auto_pay_type != '' "
                "AND curr_amount != '' AND tax_amount != '' AND due_amount != '' "
                "AND ROUND(CAST(curr_amount AS DECIMAL(15,4)) + CAST(tax_amount AS DECIMAL(15,4)), 2) "
                "!= ROUND(CAST(due_amount AS DECIMAL(15,4)), 2)"
            )
        )
    ).fetchall()
    result["check_autopay_balance"] = _jdump(
        [
            {
                **_row(r),
                "auto_pay_type": str(r.auto_pay_type),
                "curr_amount": str(r.curr_amount),
                "tax_amount": str(r.tax_amount),
                "due_amount": str(r.due_amount),
                "computed": str(
                    round(float(r.curr_amount or 0) + float(r.tax_amount or 0), 2)
                ),
            }
            for r in rows
        ]
    )

    # ── Check 24: Wrong meter fee ─────────────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, other_charge "
                "FROM billing_extract_raw WHERE other_charge < 0"
            )
        )
    ).fetchall()
    result["check_wrong_meter_fee"] = _jdump(
        [{**_row(r), "other_charge": str(r.other_charge)} for r in rows]
    )

    # ── Check 25: Renewal customer energy charge >= 0.13 ─────────────────────
    rows = (await db.execute(text("""
    SELECT r.cust_id, r.bill_no, r.company_name, r.cust_first_name, r.cust_last_name,
           r.energy_rate, r.premise_id, c.contract_rate
    FROM billing_extract_raw r
    JOIN confirmation_log c ON r.premise_id = c.esiid
    WHERE c.type_of_contract IN ('Renewal', 'Renewal+Addition')
    AND r.energy_rate != '' AND r.energy_rate != '0'
    AND r.energy_rate != CAST(c.contract_rate / 100 AS DECIMAL(15,4))
    """))).fetchall()
    result["check_renewal_energy_high"] = _jdump(
        [
            {
                **_row(r),
                "energy_rate": str(r.energy_rate),
                "contract_rate": str(r.contract_rate),
                "premise_id": str(r.premise_id),
            }
            for r in rows
        ]
    )

    # ── Check 26: Paid amount variance > 80% ─────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "pay_amount, due_amount, curr_amount, tax_amount "
                "FROM billing_extract_raw "
                "WHERE pay_amount != '' AND CAST(pay_amount AS DECIMAL(15,2)) >= 500 "
                "AND curr_amount != '' AND tax_amount != '' "
                "AND CAST(pay_amount AS DECIMAL(15,2)) >= "
                "ROUND((CAST(curr_amount AS DECIMAL(15,2)) + CAST(tax_amount AS DECIMAL(15,2))) * 0.8, 2)"
            )
        )
    ).fetchall()
    result["check_paid_amount_variance"] = _jdump(
        [
            {
                **_row(r),
                "pay_amount": str(r.pay_amount),
                "due_amount": str(r.due_amount),
                "curr_amount": str(r.curr_amount),
                "tax_amount": str(r.tax_amount),
                "percent": str(
                    round(
                        (float(r.curr_amount or 0) + float(r.tax_amount or 0)) * 0.8, 2
                    )
                ),
            }
            for r in rows
        ]
    )

    # ── Check 27: Single bills under $100 ────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "bill_mode, cust_email, auto_pay_type, curr_amount "
                "FROM billing_extract_raw "
                "WHERE (cust_type = '' OR cust_type = 'StandAlone') "
                "AND bill_mode != 'Email' "
                "AND curr_amount != '' "
                "AND CAST(curr_amount AS DECIMAL(15,2)) <= 100"
            )
        )
    ).fetchall()
    result["check_single_bill_under_100"] = _jdump(
        [
            {
                **_row(r),
                "bill_mode": str(r.bill_mode),
                "cust_email": str(r.cust_email),
                "auto_pay_type": str(r.auto_pay_type),
                "curr_amount": str(r.curr_amount),
            }
            for r in rows
        ]
    )

    # ── Check 28: Invoice billed with 2+ contracts ────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "no_of_contracts_billed "
                "FROM billing_extract_raw WHERE CAST(no_of_contracts_billed AS UNSIGNED) >= 2"
            )
        )
    ).fetchall()
    result["check_multi_contract_invoice"] = _jdump(
        [
            {**_row(r), "no_of_contracts_billed": str(r.no_of_contracts_billed)}
            for r in rows
        ]
    )

    # ── Check 29: Old auto pay customer with balance ──────────────────────────
    three_months_ago = (date.today() - timedelta(days=90)).strftime("%Y-%m-%d")
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_to_id, bill_no, cust_type, company_name, "
                "cust_first_name, cust_last_name, bal_fwd_amount, pay_amount, auto_pay_type "
                "FROM billing_extract_raw "
                "WHERE (auto_pay_type = '6' OR auto_pay_type = 'C') "
                "AND bal_fwd_amount != pay_amount"
            )
        )
    ).fetchall()
    result["check_old_autopay_balance"] = _jdump(
        [
            {
                **_row(r),
                "auto_pay_type": str(r.auto_pay_type),
                "bal_fwd_amount": str(r.bal_fwd_amount),
                "pay_amount": str(r.pay_amount),
                "cust_type": str(r.cust_type),
            }
            for r in rows
        ]
    )

    # ── Check 30: Deposit charges ─────────────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_to_id, bill_no, cust_type, company_name, "
                "cust_first_name, cust_last_name, deposit_charges, bill_handling_code "
                "FROM billing_extract_raw "
                "WHERE deposit_charges != '' AND deposit_charges != '0'"
            )
        )
    ).fetchall()
    result["check_deposit_charges"] = _jdump(
        [
            {
                **_row(r),
                "deposit_charges": str(r.deposit_charges),
                "bill_handling_code": str(r.bill_handling_code),
                "cust_type": str(r.cust_type),
            }
            for r in rows
        ]
    )

    # ── Check 31: First bill going final ─────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "bill_handling_code, service_start "
                "FROM billing_extract_raw "
                "WHERE bill_handling_code = '1' AND service_start >= :cutoff"
            ),
            {"cutoff": (date.today() - timedelta(days=35)).strftime("%Y-%m-%d")},
        )
    ).fetchall()
    result["check_first_bill_going_final"] = _jdump([_row(r) for r in rows])

    # ── Check 32: Potential final billing ────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_to_id, bill_no, premise_id, company_name, "
                "cust_first_name, cust_last_name, service_start, service_end, bill_handling_code "
                "FROM billing_extract_raw "
                "WHERE bill_handling_code = '0' "
                "AND service_start != '' AND service_end != '' "
                "AND DATEDIFF(service_end, service_start) < 27"
            )
        )
    ).fetchall()
    result["check_potential_final"] = _jdump(
        [
            {
                **_row(r),
                "service_start": str(r.service_start),
                "service_end": str(r.service_end),
                "days": (
                    str(
                        (
                            (
                                __import__("datetime").datetime.strptime(
                                    str(r.service_end)[:10], "%Y-%m-%d"
                                )
                                - __import__("datetime").datetime.strptime(
                                    str(r.service_start)[:10], "%Y-%m-%d"
                                )
                            ).days
                        )
                    )
                    if r.service_start and r.service_end
                    else "0"
                ),
                "bill_handling_code": str(r.bill_handling_code),
            }
            for r in rows
        ]
    )

    # ── Check 33: Difference 1 day ────────────────────────────────────────────
    rows = (
        await db.execute(
            text(
                "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                "service_start, service_end, "
                "DATEDIFF(service_end, service_start) AS days_diff "
                "FROM billing_extract_raw "
                "WHERE DATEDIFF(service_end, service_start) = 1"
            )
        )
    ).fetchall()
    result["check_difference_one_day"] = _jdump(
        [
            {
                **_row(r),
                "service_start": str(r.service_start),
                "service_end": str(r.service_end),
            }
            for r in rows
        ]
    )

    # ── Check 34: Different due date ─────────────────────────────────────────
    masters = (
        await db.execute(
            text(
                "SELECT cust_id, due_date FROM billing_extract_raw WHERE cust_type = 'Master'"
            )
        )
    ).fetchall()

    flagged_34 = []
    for master in masters:
        subs = (
            await db.execute(
                text(
                    "SELECT cust_id, bill_to_id, bill_no, premise_id, company_name, "
                    "cust_first_name, cust_last_name, due_date "
                    "FROM billing_extract_raw "
                    "WHERE bill_to_id = :master_id AND due_date != :due_date"
                ),
                {"master_id": master.cust_id, "due_date": master.due_date},
            )
        ).fetchall()
        for s in subs:
            flagged_34.append(
                {
                    **_row(s),
                    "bill_to_id": str(s.bill_to_id),
                    "premise_id": str(s.premise_id),
                    "due_date": str(s.due_date),
                    "master_due_date": str(master.due_date),
                }
            )
    result["check_different_due_date"] = _jdump(flagged_34)

    # ── Check 35: Master/Sub different Auto Pay Type ──────────────────────────
    masters = (
        await db.execute(
            text(
                "SELECT cust_id, auto_pay_type FROM billing_extract_raw WHERE cust_type = 'Master'"
            )
        )
    ).fetchall()
    flagged_35 = []
    for master in masters:
        subs = (
            await db.execute(
                text(
                    "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                    "auto_pay_type, bill_to_id "
                    "FROM billing_extract_raw "
                    "WHERE bill_to_id = :master_id AND auto_pay_type <> :atype"
                ),
                {"master_id": master.cust_id, "atype": master.auto_pay_type},
            )
        ).fetchall()
        for s in subs:
            flagged_35.append(
                {
                    **_row(s),
                    "master_auto_pay": str(master.auto_pay_type),
                    "sub_auto_pay": str(s.auto_pay_type),
                }
            )
    result["check_master_sub_autopay_type"] = _jdump(flagged_35)

    # ── Check 36: Master/Sub different Bill Mode ──────────────────────────────
    masters = (
        await db.execute(
            text(
                "SELECT cust_id, bill_mode FROM billing_extract_raw WHERE cust_type = 'Master'"
            )
        )
    ).fetchall()
    flagged_36 = []
    for master in masters:
        subs = (
            await db.execute(
                text(
                    "SELECT cust_id, bill_no, company_name, cust_first_name, cust_last_name, "
                    "bill_mode, bill_to_id "
                    "FROM billing_extract_raw "
                    "WHERE bill_to_id = :master_id AND bill_mode <> :bmode"
                ),
                {"master_id": master.cust_id, "bmode": master.bill_mode},
            )
        ).fetchall()
        for s in subs:
            flagged_36.append(
                {
                    **_row(s),
                    "master_bill_mode": str(master.bill_mode),
                    "sub_bill_mode": str(s.bill_mode),
                }
            )
    result["check_master_sub_bill_mode"] = _jdump(flagged_36)

    return result


# ── save comments ─────────────────────────────────────────────────────────────


async def save_comments(
    payload: BillingCommentSave, db: AsyncSession, commented_by: str
):
    row = (
        await db.execute(
            text(
                "SELECT id FROM billing_exception_log "
                "WHERE upload_id = :uid AND row_type = 'comment'"
            ),
            {"uid": payload.upload_id},
        )
    ).fetchone()

    if not row:
        raise HTTPException(
            status_code=404, detail="Comment row not found for this upload"
        )

    updates = payload.dict(exclude={"upload_id"}, exclude_none=True)
    if not updates:
        return {"message": "Nothing to update"}

    set_clause = ", ".join([f"{k} = :{k}" for k in updates])
    updates["row_id"] = row.id
    await db.execute(
        text(f"UPDATE billing_exception_log SET {set_clause} WHERE id = :row_id"),
        updates,
    )
    await db.commit()
    return {"message": "Comments saved"}


# ── get exceptions for a date ─────────────────────────────────────────────────


def _build_attachments(exc: dict, upload_date, raw_rows: list = []) -> list:
    attachments = []
    date_str = str(upload_date).replace("-", "")

    def parse(key):
        data = exc.get(key)
        return json.loads(data) if data else []

    def rows(data, fields):
        return [[r.get(f, "") for f in fields] for r in data]

    # Full billing extract from DB

    if raw_rows:
        headers = list(raw_rows[0]._mapping.keys())
        data = _make_excel(
            "Billing Extract",
            headers,
            [[str(r._mapping[h] or "") for h in headers] for r in raw_rows],
        )
        attachments.append((f"Billing_Extract_{date_str}.xlsx", data))

    base_fields = ["cust_id", "bill_no", "company_name", "cust_name"]

    # 1. First Bills and Final Bills
    first = parse("check_first_bill")
    final = parse("check_final_bill")
    combined = [{"type": "First Bill", **r} for r in first] + [
        {"type": "Final Bill", **r} for r in final
    ]
    if combined:
        data = _make_excel(
            "First & Final Bills",
            ["Type", "Cust ID", "Bill No", "Company Name", "Customer Name"],
            [
                [
                    r.get("type"),
                    r.get("cust_id"),
                    r.get("bill_no"),
                    r.get("company_name"),
                    r.get("cust_name"),
                ]
                for r in combined
            ],
        )
        attachments.append((f"First_Bills_and_Final_Bills_{date_str}.xlsx", data))

    # 2. Previous dues
    autopay = parse("check_autopay_balance")
    neg_bal = parse("check_negative_balance")
    dues = autopay + neg_bal
    if dues:
        data = _make_excel(
            "Previous Dues",
            [
                "Cust ID",
                "Bill No",
                "Company Name",
                "Customer Name",
                "Due Amount",
                "Auto Pay Type",
            ],
            [
                [
                    r.get("cust_id"),
                    r.get("bill_no"),
                    r.get("company_name"),
                    r.get("cust_name"),
                    r.get("due_amount", ""),
                    r.get("auto_pay_type", ""),
                ]
                for r in dues
            ],
        )
        attachments.append(
            (f"List_of_customer_with_previous_dues_{date_str}.xlsx", data)
        )

    # 3. State Tax
    state_tax = parse("check_state_tax_100")
    data = _make_excel("State Tax", base_fields, rows(state_tax, base_fields))
    attachments.append((f"STATE_TAX_{date_str}.xlsx", data))

    # 4. Deposit charges
    deposit = parse("check_deposit_charges")

    data = _make_excel(
        "Deposit Charges",
        ["Cust ID", "Bill No", "Company Name", "Customer Name", "Deposit Charges"],
        [
            [
                r.get("cust_id"),
                r.get("bill_no"),
                r.get("company_name"),
                r.get("cust_name"),
                r.get("deposit_charges", ""),
            ]
            for r in deposit
        ],
    )
    attachments.append((f"Movin_charge_{date_str}.xlsx", data))

    # 5. Tax exemption
    exempt1 = parse("check_residential_tax_exempt")
    exempt2 = parse("check_residential_puc_grt_city")
    exempt = exempt1 + exempt2

    data = _make_excel(
        "Tax Exemption",
        [
            "Cust ID",
            "Bill No",
            "Company Name",
            "Customer Name",
            "Gross Tax",
            "Pugra Tax",
        ],
        [
            [
                r.get("cust_id"),
                r.get("bill_no"),
                r.get("company_name"),
                r.get("cust_name"),
                r.get("gros_tax", ""),
                r.get("pugra_tax", ""),
            ]
            for r in exempt
        ],
    )
    attachments.append((f"Tax_exemption_{date_str}.xlsx", data))

    # 6. Renewal accounts
    renewal = parse("check_renewal_energy_high")
    if renewal:
        data = _make_excel(
            "Renewal Accounts",
            ["Cust ID", "Bill No", "Company Name", "Customer Name", "Computed Rate"],
            [
                [
                    r.get("cust_id"),
                    r.get("bill_no"),
                    r.get("company_name"),
                    r.get("cust_name"),
                    r.get("computed_rate", ""),
                ]
                for r in renewal
            ],
        )
        attachments.append((f"Renewal_Account_{date_str}.xlsx", data))

    return attachments


async def get_exceptions_by_date(upload_date: str, db: AsyncSession):
    rows = (
        await db.execute(
            text(
                "SELECT * FROM billing_exception_log "
                "WHERE upload_date = :d ORDER BY row_type ASC"
            ),
            {"d": upload_date},
        )
    ).fetchall()

    if not rows:
        raise HTTPException(status_code=404, detail="No data found for this date")
    return rows


async def get_last_exceptions(db: AsyncSession):
    latest = (
        await db.execute(
            text(
                "SELECT upload_date FROM billing_upload_log ORDER BY upload_date DESC LIMIT 1"
            )
        )
    ).fetchone()
    if not latest:
        raise HTTPException(status_code=404, detail="No uploads found")
    return await get_exceptions_by_date(str(latest.upload_date), db)


async def get_upload_history(db: AsyncSession):
    rows = (
        await db.execute(
            text("SELECT * FROM billing_upload_log ORDER BY upload_date DESC LIMIT 60")
        )
    ).fetchall()
    return rows


# ── email recipients ──────────────────────────────────────────────────────────


async def get_recipients(db: AsyncSession):
    return (
        await db.execute(
            text("SELECT * FROM billing_email_recipients ORDER BY name ASC")
        )
    ).fetchall()


async def add_recipient(payload: RecipientCreate, db: AsyncSession):
    await db.execute(
        text(
            "INSERT INTO billing_email_recipients (name, email) VALUES (:name, :email)"
        ),
        {"name": payload.name, "email": payload.email},
    )
    await db.commit()
    return {"message": "Recipient added"}


async def toggle_recipient(recipient_id: int, db: AsyncSession):
    await db.execute(
        text("UPDATE billing_email_recipients SET active = 1 - active WHERE id = :id"),
        {"id": recipient_id},
    )
    await db.commit()
    return {"message": "Recipient updated"}


async def delete_recipient(recipient_id: int, db: AsyncSession):
    await db.execute(
        text("DELETE FROM billing_email_recipients WHERE id = :id"),
        {"id": recipient_id},
    )
    await db.commit()
    return {"message": "Recipient deleted"}


async def rerun_checks(db: AsyncSession):
    # get latest upload
    log_row = (
        await db.execute(
            text(
                "SELECT id, upload_date FROM billing_upload_log ORDER BY upload_date DESC LIMIT 1"
            )
        )
    ).fetchone()
    if not log_row:
        raise HTTPException(status_code=404, detail="No uploads found")

    # run all 36 checks
    exceptions = await _run_all_checks(db)

    # update exception row
    set_clause = ", ".join([f"{k} = :{k}" for k in exceptions])
    exceptions["upload_id"] = log_row.id
    await db.execute(
        text(
            f"UPDATE billing_exception_log SET {set_clause} "
            f"WHERE upload_id = :upload_id AND row_type = 'exception'"
        ),
        exceptions,
    )
    await db.commit()
    return {"message": "Checks re-run successfully", "upload_id": log_row.id}


async def compare_with_php(upload_id: int, upload_date, db: AsyncSession):
    try:
        print(
            f"compare_with_php called: upload_id={upload_id}, upload_date={upload_date}, req={req}"
        )
        print(f"Step 1: starting")
        session = req.Session()
        print(f"Step 2: session created")
        # Login
        login = session.post(
            "https://portal.enertsol.com/login.php",
            data={
                "login": os.getenv("PHP_PORTAL_USER", ""),
                "pass": os.getenv("PHP_PORTAL_PASS", ""),
                "submit": "Submit",
            },
            timeout=30,
        )
        print(f"Step 3: logged in {login.status_code}")

        # Fetch results page
        response = session.get(
            "https://portal.enertsol.com/billing_extract_result.php", timeout=60
        )
        print(f"Step 4: fetched page {response.status_code}")

        soup = BeautifulSoup(response.text, "html.parser")

        # PHP check number → our column name
        PHP_MAP = {
            "1": "check_tax_zero",
            "2": "check_kh_qty_energy_zero",
            "3": "check_kh_qty_metered_mismatch",
            "4": "check_residential_puc_grt_city",
            "5": "check_residential_tax_exempt",
            "6": "check_mcpe_bills",
            "7": "check_lmp_rate_range",
            "8": "check_sub_only_no_master",
            "9": "check_commercial_tdsp",
            "10": "check_residential_price_low",
            "11": "check_residential_price_high",
            "12": "check_commercial_price_high",
            "13": "check_commercial_price_low",
            "14": "check_negative_balance",
            "15": "check_zero_usage",
            "16": "check_partial_payment",
            "17": "check_zero_meter_fee",
            "18": "check_first_bill",
            "19": "check_final_bill",
            "20": "check_master_sub_final",
            "21": "check_state_tax_100",
            "22": "check_credit_card_final",
            "23": "check_autopay_balance",
            "24": "check_wrong_meter_fee",
            "25": "check_renewal_energy_high",
            "26": "check_paid_amount_variance",
            "27": "check_single_bill_under_100",
            "28": "check_multi_contract_invoice",
            "29": "check_old_autopay_balance",
            "30": "check_deposit_charges",
            "31": "check_first_bill_going_final",
            "32": "check_potential_final",
            "33": "check_difference_one_day",
            "34": "check_different_due_date",
            "35": "check_master_sub_autopay_type",
            "36": "check_master_sub_bill_mode",
        }

        # Parse counts
        php_counts = {}

        print(f"Step 5: parsed HTML")
        for tag in soup.find_all("td"):
            td_text = tag.get_text(strip=True)
            match = re.match(r"^Check\s*#\s*(\d+)\s*$", td_text)
            if match:
                print(f"Step 6: found check {match.group(1)}")
                check_num = match.group(1)
                col_name = PHP_MAP.get(check_num)
                if not col_name:
                    continue
                parent_table = tag.find_parent("table")
                if parent_table:
                    rows = parent_table.find_all("tr")
                    max_sr = 0
                    for r in rows:
                        if r.get("bgcolor"):
                            continue
                        if r.find("textarea") or r.find("input"):
                            continue
                        first_td = r.find("td")
                        if first_td and first_td.get_text(strip=True).isdigit():
                            sr_val = int(first_td.get_text(strip=True))
                            if sr_val > max_sr:
                                max_sr = sr_val
                    php_counts[col_name] = max_sr

        if not php_counts:
            print("PHP comparison: no counts parsed")
            return None

        # Save to billing_php_comparison
        set_clause = ", ".join([f"{k} = :{k}" for k in php_counts])
        php_counts["upload_id"] = upload_id
        php_counts["upload_date"] = upload_date
        print(f"Step 7: php_counts = {php_counts}")
        print(f"Step 8: about to insert")

        # separate data columns from keys
        data_cols = {
            k: v for k, v in php_counts.items() if k not in ["upload_id", "upload_date"]
        }
        col_names = ", ".join(data_cols.keys())
        col_params = ", ".join([f":{k}" for k in data_cols.keys()])
        update_clause = ", ".join([f"{k} = :{k}" for k in data_cols.keys()])

        params = {**data_cols, "upload_id": upload_id, "upload_date": upload_date}

        await db.execute(
            text(f"""
            INSERT INTO billing_php_comparison (upload_id, upload_date, {col_names})
            VALUES (:upload_id, :upload_date, {col_params})
          ON DUPLICATE KEY UPDATE {update_clause}
       """),
            params,
        )
        await db.commit()

        print(f"PHP comparison saved for upload_id {upload_id}")
        return php_counts

    except Exception as e:
        print(f"PHP comparison failed: {e}")
        return None
