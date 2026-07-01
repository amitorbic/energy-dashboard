from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from utils.database import get_db
from middleware.auth import require_auth
from pydantic import BaseModel
from typing import Optional
from decimal import Decimal, InvalidOperation
from datetime import datetime
import io
import openpyxl

router = APIRouter(prefix="/enrollment-engine", tags=["enrollment-engine"])

# ── Plan code lookup ──────────────────────────────────────────────────────────

_PR_LMP0 = {
    "0.00": "PR1503060001",  "2.99": "PR1503090001",
    "3.99": "PR1504230001",  "4.95": "PR1506170002",
    "4.99": "PR1503090002",  "5.00": "PR1503060003",
    "5.95": "PR1506170001",  "5.99": "PR1503050002",
    "6.95": "PR1509030001",  "7.95": "PR1503120001",
    "7.99": "PR1503060002",  "9.99": "PR1509030001",
    "10.00": "PR1503050003",
}
_PR_LMP1 = {"10.00": "PR1503050004"}
_PR_LMP2 = {"10.00": "PR1503050005"}


def _clean_fee(meter_fees) -> str:
    if not meter_fees:
        return "0.00"
    try:
        val = Decimal(str(meter_fees).replace("$", "").strip())
        return f"{val:.2f}"
    except (InvalidOperation, ValueError):
        return "0.00"


def _fee_to_pr(meter_fees, lmp) -> str:
    fee = _clean_fee(meter_fees)
    lmp_int = int(lmp or 0)
    if lmp_int == 1:
        return _PR_LMP1.get(fee, "")
    if lmp_int == 2:
        return _PR_LMP2.get(fee, "")
    return _PR_LMP0.get(fee, "")


def _parse_enrollment_date(d) -> Optional[str]:
    """Convert MM/DD/YYYY or YYYY-MM-DD to YYYY-MM-DD for MySQL DATE."""
    if not d:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(d).strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


# ── MasterRoll 2016 column layout (128 cols, A–DX) ───────────────────────────

_HEADERS = [
    "Batch_No", "Batch_File_Name", "Source", "source_line", "ENROLL_SERIAL_NO",
    "Sic_Code", "Cust_Class", "Referred_Code", "Referred_Cust_ID", "Credit_Rate2",
    "Credit_Score", "Credit_Score_Source", "Credit_Score_Date", "Cust_Type",
    "Master_Id", "Master_Ref", "Agent_Code", "Premise_ID", "Plan_Group",
    "Request_Date", "Enrol_Type", "Offcycle_Switch_Date", "Company_Name",
    "Cust_Firstname", "Cust_Lastname", "Cust_Mi", "SSN", "Phone1",
    "Cm_Address1", "Cm_Address2", "Cm_Address3", "Cm_City", "Cm_State", "Cm_Zip",
    "Phone2", "Email_Address", "Contact1", "Contact2", "Account_Rep",
    "Life_Support", "waiver_notice", "cust_status", "plan_id1", "plan_id2",
    "plan_id3", "cust_ref_id", "billto_cust_id", "lock_box", "pay_term",
    "cust_bill_type", "cust_bill_mode",
    "tax_exempt1", "tax_exempt2", "tax_exempt3", "tax_exempt4",
    "tax_exempt5", "tax_exempt6", "tax_exempt7", "tax_exempt8",
    "credit_id", "edi_bill_presenter", "deposit_plan", "deposit_amount",
    "deposit_pay_type", "deposit_aba_nbr", "deposit_account_no", "deposit_cc_no",
    "deposit_expiry_YYYY", "deposit_expiry_MM", "deposit_security_code",
    "deposit_pay_amount", "Deposit_charge", "deposit_acct_type", "deposit_card_type",
    "plan_id_t1", "plan_id_t1_rate", "plan_id_t1_rateper",
    "plan_id_t2", "plan_id_t2_rate", "plan_id_t2_rateper",
    "plan_id_t3", "plan_id_t3_rate", "plan_id_t3_rateper",
    "use_data_from_file",
    "pm_address1", "pm_address2", "pm_city", "pm_state", "pm_zip",
    "pm_county", "pm_country", "pm_duns", "pm_meter", "pm_multiplier",
    "priority_code", "multi_plan", "multi_plan_rate", "spouse_email", "cust_fax1",
    "use_cust_id", "promo_code", "cust_coments",
    "contract_ind", "contract_no", "master_contract_no",
    "contract_date", "contract_start_date", "contract_end_date", "contract_term",
    "contract_type", "calc_method", "sys_charge_code_st", "rate_type",
    "contract_rate", "adder1_rate", "adder2_rate", "agent_duns",
    "flat_chg_amt", "trueup_term", "agent_commission_rate", "contract_comments",
    "mcp_multiplier", "fixed_rate_factor", "default_contract_id",
    "enroll_product", "flow_status", "current_rate", "current_rate_json",
]

_COL = {
    "batch_no":        1,
    "batch_file_name": 2,
    "source":          3,
    "serial_no":       5,
    "agent_code":      17,
    "premise_id":      18,
    "plan_group":      19,
    "request_date":    20,
    "enrol_type":      21,
    "offcycle_switch": 22,
    "company_name":    23,
    "email_address":   36,
    "life_support":    40,
    "waiver_notice":   41,
    "cust_status":     42,
    "plan_id1":        43,
    "cust_bill_mode":  51,
    "contract_ind":   103,
    "contract_no":    104,
    "contract_date":  106,
    "contract_start": 107,
    "contract_term":  109,
    "contract_type":  110,
    "agent_comm":     120,
    "enroll_product": 125,
    "flow_status":    126,
    "current_rate":   127,
}


def _build_row(rec: dict, batch_no: int, serial: int) -> list:
    row = [""] * 128

    try:
        rate = round(float(str(rec.get("contract_rate") or "0")) / 100, 6)
    except (ValueError, TypeError):
        rate = ""

    plan_group = rec.get("plan_group") or "C1"

    row[_COL["batch_no"] - 1]        = f"B{batch_no}"
    row[_COL["batch_file_name"] - 1] = "X2"
    row[_COL["source"] - 1]          = "BATCH"
    row[_COL["serial_no"] - 1]       = serial
    row[_COL["agent_code"] - 1]      = rec.get("broker_code") or ""
    row[_COL["premise_id"] - 1]      = rec.get("esiid") or ""
    row[_COL["plan_group"] - 1]      = plan_group
    row[_COL["request_date"] - 1]    = rec.get("start_date") or ""
    row[_COL["enrol_type"] - 1]      = "S"
    row[_COL["offcycle_switch"] - 1] = rec.get("start_date") or ""
    row[_COL["company_name"] - 1]    = rec.get("customer_name") or ""
    row[_COL["email_address"] - 1]   = rec.get("customer_email") or ""
    row[_COL["life_support"] - 1]    = "N"
    row[_COL["waiver_notice"] - 1]   = "N"
    row[_COL["cust_status"] - 1]     = "P"
    row[_COL["plan_id1"] - 1]        = "PNCPOSTPAY"
    row[_COL["cust_bill_mode"] - 1]  = "Email" if rec.get("customer_email") else ""
    row[_COL["contract_ind"] - 1]    = "Y"
    row[_COL["contract_no"] - 1]     = rec.get("contract_no") or ""
    row[_COL["contract_date"] - 1]   = rec.get("start_date") or ""
    row[_COL["contract_start"] - 1]  = rec.get("start_date") or ""
    row[_COL["contract_term"] - 1]   = rec.get("term") or ""
    row[_COL["contract_type"] - 1]   = "FIXED"
    row[_COL["agent_comm"] - 1]      = rec.get("commission") or ""
    row[_COL["enroll_product"] - 1]  = _fee_to_pr(rec.get("meter_fees"), rec.get("lmp"))
    row[_COL["flow_status"] - 1]     = "-10"
    row[_COL["current_rate"] - 1]    = rate

    return row


def _load_plan_group_map_query() -> str:
    return "SELECT premise_id, plan_group FROM contract_renewal WHERE premise_id IS NOT NULL"


def _expand_esiids(rows: list, plan_group_map: dict) -> list:
    """Expand comma-separated esiid strings into one record per ESIID."""
    expanded: list = []
    for row in rows:
        esiid_raw = (row.get("esiid") or "").strip()
        if "," in esiid_raw:
            parts = [e.strip() for e in esiid_raw.split(",") if e.strip()]
            for esid in parts:
                nr = dict(row)
                nr["esiid"] = esid
                nr["plan_group"] = plan_group_map.get(esid)
                expanded.append(nr)
        else:
            row["plan_group"] = plan_group_map.get(esiid_raw) if esiid_raw else None
            expanded.append(row)
    return expanded


# ── Request models ────────────────────────────────────────────────────────────

class GenerateRequest(BaseModel):
    record_sids: list[int]
    date_from: Optional[str] = None
    date_to: Optional[str] = None


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/pending")
async def get_pending(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    broker_code: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(require_auth),
):
    filters = ["cl.enroll_check = 0"]
    params: dict = {}

    if date_from:
        filters.append("cl.start_date >= :date_from")
        params["date_from"] = date_from
    if date_to:
        filters.append("cl.start_date <= :date_to")
        params["date_to"] = date_to
    if broker_code:
        filters.append("cl.broker_code = :broker_code")
        params["broker_code"] = broker_code

    where = " AND ".join(filters)

    result = await db.execute(
        text(f"""
            SELECT
                cl.sid, cl.esiid, cl.customer_name, cl.broker_code, cl.broker_name,
                cl.start_date, cl.term, cl.contract_rate, cl.meter_fees, cl.lmp,
                cl.tax_exempt, cl.customer_email, cl.contract_no, cl.date_modified,
                cl.commission, cl.type_of_contract
            FROM confirmation_log cl
            WHERE {where}
            ORDER BY cl.date_modified DESC
        """),
        params,
    )
    rows = [dict(r) for r in result.mappings().all()]

    pg_result = await db.execute(text(_load_plan_group_map_query()))
    plan_group_map = {r[0]: r[1] for r in pg_result.fetchall()}

    expanded = _expand_esiids(rows, plan_group_map)

    plan_result = await db.execute(
        text("SELECT id, base_fee, plan_id, plan_name, paired_with FROM plan_codes WHERE active = 1 ORDER BY id")
    )
    plan_list = [dict(r) for r in plan_result.mappings().all()]

    fee_to_plan: dict = {}
    for p in plan_list:
        key = f"{float(p['base_fee']):.2f}"
        if key not in fee_to_plan:
            fee_to_plan[key] = p

    plan_by_id = {p["plan_id"]: p for p in plan_list}

    for row in expanded:
        fee = _clean_fee(row.get("meter_fees"))
        plan = fee_to_plan.get(fee)
        row["suggested_plan"] = plan["plan_id"] if plan else None
        row["suggested_plan_name"] = plan["plan_name"] if plan else None
        if plan and plan.get("paired_with"):
            paired = plan_by_id.get(plan["paired_with"])
            row["paired_plan"] = paired["plan_id"] if paired else None
            row["paired_plan_name"] = paired["plan_name"] if paired else None
        else:
            row["paired_plan"] = None
            row["paired_plan_name"] = None

    return {"records": expanded, "total": len(expanded)}


@router.get("/plan-codes")
async def get_plan_codes(
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(require_auth),
):
    result = await db.execute(
        text("SELECT * FROM plan_codes WHERE active = 1 ORDER BY customer_type, base_fee, plan_id")
    )
    rows = [dict(r) for r in result.mappings().all()]

    grouped: dict = {}
    for row in rows:
        ctype = row.get("customer_type") or "other"
        grouped.setdefault(ctype, []).append(row)
    return grouped


@router.post("/generate-masterroll")
async def generate_masterroll(
    body: GenerateRequest,
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(require_auth),
):
    if not body.record_sids:
        raise HTTPException(status_code=400, detail="No record sids provided")

    sid_list = ",".join(str(s) for s in body.record_sids)

    result = await db.execute(
        text(f"""
            SELECT
                cl.sid, cl.esiid, cl.customer_name, cl.broker_code, cl.broker_name,
                cl.start_date, cl.term, cl.contract_rate, cl.meter_fees, cl.lmp,
                cl.customer_email, cl.contract_no, cl.commission,
                cl.billing_address, cl.billing_city, cl.billing_state, cl.billing_zip,
                cl.plan_group, cl.plan_id, cl.cust_first_name, cl.cust_last_name
            FROM confirmation_log cl
            WHERE cl.sid IN ({sid_list})
            ORDER BY cl.customer_name
        """)
    )
    records = [dict(r) for r in result.mappings().all()]
    if not records:
        raise HTTPException(status_code=404, detail="No matching records found")

    pg_result = await db.execute(text(_load_plan_group_map_query()))
    plan_group_map = {r[0]: r[1] for r in pg_result.fetchall()}

    expanded = _expand_esiids(records, plan_group_map)

    batch_r = await db.execute(text("SELECT COALESCE(MAX(id), 0) + 1 FROM enrollment_batches"))
    batch_no = batch_r.scalar()

    # Build XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MasterRoll"
    ws.cell(row=1, column=1, value="enrolment_queue")
    for col_idx, header in enumerate(_HEADERS, start=1):
        ws.cell(row=2, column=col_idx, value=header)

    for serial, rec in enumerate(expanded, start=1):
        row_data = _build_row(rec, batch_no, serial)
        for col_idx, value in enumerate(row_data, start=1):
            if value != "":
                ws.cell(row=serial + 2, column=col_idx, value=value)

    # Generate cust_ids and insert into contract_renewal
    date_prefix = datetime.now().strftime("%y%m%d")
    seq_r = await db.execute(
        text("SELECT COUNT(*) FROM contract_renewal WHERE cust_id LIKE :prefix"),
        {"prefix": f"{date_prefix}%"},
    )
    base_seq = seq_r.scalar() or 0

    skipped = []
    inserted_count = 0
    for rec in expanded:
        esi_id = rec.get("esiid") or ""
        if esi_id:
            dup_r = await db.execute(
                text(
                    "SELECT cust_id FROM contract_renewal"
                    " WHERE premise_id = :esi"
                    " AND status IN ('active', 'pending', 'going_final') LIMIT 1"
                ),
                {"esi": esi_id},
            )
            if dup_r.fetchone():
                skipped.append({
                    "esi_id": esi_id,
                    "customer_name": rec.get("customer_name", ""),
                    "reason": "Active contract exists",
                })
                continue

        cust_id = f"{date_prefix}{base_seq + inserted_count + 1:04d}"
        inserted_count += 1
        raw_rate = rec.get("contract_rate")
        contract_rate = None
        if raw_rate:
            try:
                contract_rate = round(float(str(raw_rate)) / 100, 6)
            except (ValueError, TypeError):
                contract_rate = None
        await db.execute(
            text("""
                INSERT INTO contract_renewal (
                    premise_id, cust_id, status, batch_no, account_type,
                    contract_start_date, company_name,
                    cust_first_name, cust_last_name, cust_email,
                    billing_address, billing_city, billing_state, billing_zip,
                    broker_code, broker_name, plan_group, plan_id, other_charge,
                    contract_rate
                ) VALUES (
                    :premise_id, :cust_id, 'pending', :batch_no, 'standalone',
                    :contract_start_date, :company_name,
                    :cust_first_name, :cust_last_name, :cust_email,
                    :billing_address, :billing_city, :billing_state, :billing_zip,
                    :broker_code, :broker_name, :plan_group, :plan_id, :other_charge,
                    :contract_rate
                )
            """),
            {
                "premise_id":          esi_id,
                "cust_id":             cust_id,
                "batch_no":            str(batch_no),
                "contract_start_date": _parse_enrollment_date(rec.get("start_date")),
                "company_name":        rec.get("customer_name") or "",
                "cust_first_name":     rec.get("cust_first_name") or None,
                "cust_last_name":      rec.get("cust_last_name") or None,
                "cust_email":          rec.get("customer_email") or None,
                "billing_address":     rec.get("billing_address") or None,
                "billing_city":        rec.get("billing_city") or None,
                "billing_state":       rec.get("billing_state") or None,
                "billing_zip":         rec.get("billing_zip") or None,
                "broker_code":         rec.get("broker_code") or None,
                "broker_name":         rec.get("broker_name") or None,
                "plan_group":          rec.get("plan_group") or "C1",
                "plan_id":             rec.get("plan_id") or None,
                "other_charge":        str(float(_clean_fee(rec.get("meter_fees")))),
                "contract_rate":       contract_rate,
            },
        )

    # Mark original records enrolled and log batch
    await db.execute(
        text(f"UPDATE confirmation_log SET enroll_check = 1 WHERE sid IN ({sid_list})")
    )
    await db.execute(
        text("""
            INSERT INTO enrollment_batches (batch_no, generated_by, record_count, date_from, date_to)
            VALUES (:batch_no, :generated_by, :record_count, :date_from, :date_to)
        """),
        {
            "batch_no":      str(batch_no),
            "generated_by":  payload.get("username") or payload.get("email") or "unknown",
            "record_count":  inserted_count,
            "date_from":     body.date_from or None,
            "date_to":       body.date_to or None,
        },
    )
    await db.commit()

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    import json as _json
    filename = f"MasterRoll {datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Enrollment-Skipped": _json.dumps(skipped),
        },
    )


# ── Step 3: Batch management ──────────────────────────────────────────────────

@router.get("/batches")
async def list_batches(
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(require_auth),
):
    result = await db.execute(
        text("""
            SELECT id, batch_no, generated_by, generated_at,
                   record_count, date_from, date_to, status, submitted_at
            FROM enrollment_batches
            ORDER BY generated_at DESC
        """)
    )
    rows = [dict(r) for r in result.mappings().all()]

    # Coerce non-serialisable types
    for row in rows:
        if row.get("generated_at") and not isinstance(row["generated_at"], str):
            row["generated_at"] = row["generated_at"].isoformat()
        if row.get("submitted_at") and not isinstance(row["submitted_at"], str):
            row["submitted_at"] = row["submitted_at"].isoformat()

    return {"batches": rows}


@router.post("/batches/{batch_no}/submit")
async def submit_batch(
    batch_no: str,
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(require_auth),
):
    result = await db.execute(
        text("SELECT id FROM enrollment_batches WHERE batch_no = :batch_no"),
        {"batch_no": batch_no},
    )
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Batch not found")

    await db.execute(
        text("""
            UPDATE enrollment_batches
            SET status = 'submitted', submitted_at = NOW()
            WHERE batch_no = :batch_no
        """),
        {"batch_no": batch_no},
    )
    await db.commit()
    return {"ok": True, "batch_no": batch_no, "status": "submitted"}


# ── Step 4: Per-batch customer view ──────────────────────────────────────────

@router.get("/batches/{batch_no}")
async def get_batch_customers(
    batch_no: str,
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(require_auth),
):
    batch_r = await db.execute(
        text("""
            SELECT id, batch_no, generated_by, generated_at,
                   record_count, date_from, date_to, status, submitted_at
            FROM enrollment_batches WHERE batch_no = :batch_no
        """),
        {"batch_no": batch_no},
    )
    batch_row = batch_r.mappings().fetchone()
    if not batch_row:
        raise HTTPException(status_code=404, detail="Batch not found")

    batch = dict(batch_row)
    if batch.get("generated_at") and not isinstance(batch["generated_at"], str):
        batch["generated_at"] = batch["generated_at"].isoformat()
    if batch.get("submitted_at") and not isinstance(batch["submitted_at"], str):
        batch["submitted_at"] = batch["submitted_at"].isoformat()

    cust_r = await db.execute(
        text("""
            SELECT cust_id AS customer_id, premise_id AS esi_id, company_name, status,
                   broker_code AS broker_id, broker_name, plan_group,
                   other_charge AS meter_fee, contract_start_date AS enrollment_date
            FROM contract_renewal
            WHERE batch_no = :batch_no
            ORDER BY cust_id
        """),
        {"batch_no": batch_no},
    )
    customers = [dict(r) for r in cust_r.mappings().all()]
    for c in customers:
        if c.get("enrollment_date") and not isinstance(c["enrollment_date"], str):
            c["enrollment_date"] = str(c["enrollment_date"])

    return {"batch": batch, "customers": customers}


# ── Step 5: Activate / cancel customer ───────────────────────────────────────

@router.post("/activate/{customer_id}")
async def activate_customer(
    customer_id: str,
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(require_auth),
):
    cust_r = await db.execute(
        text("SELECT cust_id, premise_id FROM contract_renewal WHERE cust_id = :cid"),
        {"cid": customer_id},
    )
    row = cust_r.mappings().fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Customer not found")

    esi_id = row["premise_id"]

    # Pull rate from confirmation_log (¢/kWh → $/kWh) and refresh contract_rate
    rate_r = await db.execute(
        text("""
            SELECT contract_rate FROM confirmation_log
            WHERE esiid = :esi
               OR esiid LIKE CONCAT(:esi, ',%')
               OR esiid LIKE CONCAT('%, ', :esi)
               OR esiid LIKE CONCAT('%, ', :esi, ',%')
            ORDER BY date_modified DESC
            LIMIT 1
        """),
        {"esi": esi_id},
    )
    rate_row = rate_r.fetchone()
    contract_rate = None
    if rate_row and rate_row[0]:
        try:
            contract_rate = round(float(str(rate_row[0])) / 100, 6)
        except (ValueError, TypeError):
            contract_rate = None

    await db.execute(
        text("""
            UPDATE contract_renewal
            SET status = 'active',
                contract_rate = COALESCE(:contract_rate, contract_rate)
            WHERE cust_id = :cid
        """),
        {"cid": customer_id, "contract_rate": contract_rate},
    )
    await db.commit()
    return {"ok": True, "customer_id": customer_id, "esi_id": esi_id}


@router.post("/cancel/{customer_id}")
async def cancel_customer(
    customer_id: str,
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(require_auth),
):
    result = await db.execute(
        text("SELECT cust_id FROM contract_renewal WHERE cust_id = :cid"),
        {"cid": customer_id},
    )
    if not result.fetchone():
        raise HTTPException(status_code=404, detail="Customer not found")

    await db.execute(
        text("UPDATE contract_renewal SET status = 'cancelled' WHERE cust_id = :cid"),
        {"cid": customer_id},
    )
    await db.commit()
    return {"ok": True, "customer_id": customer_id, "status": "cancelled"}


# ── Create Internal Batch (no Excel — Renewals / Assignments / B&E) ───────────

class InternalBatchRequest(BaseModel):
    record_sids: list[int]


@router.post("/create-internal-batch")
async def create_internal_batch(
    body: InternalBatchRequest,
    db: AsyncSession = Depends(get_db),
    payload: dict = Depends(require_auth),
):
    if not body.record_sids:
        raise HTTPException(status_code=400, detail="No record sids provided")

    sid_list = ",".join(str(s) for s in body.record_sids)

    result = await db.execute(
        text(f"""
            SELECT
                cl.sid, cl.esiid, cl.customer_name, cl.broker_code, cl.broker_name,
                cl.start_date, cl.term, cl.contract_rate, cl.meter_fees,
                cl.customer_email, cl.commission,
                cl.billing_address, cl.billing_city, cl.billing_state, cl.billing_zip,
                cl.plan_group, cl.plan_id, cl.cust_first_name, cl.cust_last_name
            FROM confirmation_log cl
            WHERE cl.sid IN ({sid_list})
            ORDER BY cl.customer_name
        """)
    )
    records = [dict(r) for r in result.mappings().all()]
    if not records:
        raise HTTPException(status_code=404, detail="No matching records found")

    pg_result = await db.execute(text(_load_plan_group_map_query()))
    plan_group_map = {r[0]: r[1] for r in pg_result.fetchall()}

    expanded = _expand_esiids(records, plan_group_map)

    batch_r = await db.execute(text("SELECT COALESCE(MAX(id), 0) + 1 FROM enrollment_batches"))
    batch_no = batch_r.scalar()

    date_prefix = datetime.now().strftime("%y%m%d")
    seq_r = await db.execute(
        text("SELECT COUNT(*) FROM contract_renewal WHERE cust_id LIKE :prefix"),
        {"prefix": f"{date_prefix}%"},
    )
    base_seq = seq_r.scalar() or 0

    skipped = []
    inserted_count = 0
    for rec in expanded:
        esi_id = rec.get("esiid") or ""
        if esi_id:
            dup_r = await db.execute(
                text(
                    "SELECT cust_id FROM contract_renewal"
                    " WHERE premise_id = :esi"
                    " AND status IN ('active', 'pending', 'going_final') LIMIT 1"
                ),
                {"esi": esi_id},
            )
            if dup_r.fetchone():
                skipped.append({
                    "esi_id": esi_id,
                    "customer_name": rec.get("customer_name", ""),
                    "reason": "Active contract exists",
                })
                continue

        cust_id = f"{date_prefix}{base_seq + inserted_count + 1:04d}"
        inserted_count += 1
        raw_rate = rec.get("contract_rate")
        contract_rate = None
        if raw_rate:
            try:
                contract_rate = round(float(str(raw_rate)) / 100, 6)
            except (ValueError, TypeError):
                contract_rate = None
        await db.execute(
            text("""
                INSERT INTO contract_renewal (
                    premise_id, cust_id, status, batch_no, account_type,
                    contract_start_date, company_name,
                    cust_first_name, cust_last_name, cust_email,
                    billing_address, billing_city, billing_state, billing_zip,
                    broker_code, broker_name, plan_group, plan_id, other_charge,
                    contract_rate
                ) VALUES (
                    :premise_id, :cust_id, 'pending', :batch_no, 'standalone',
                    :contract_start_date, :company_name,
                    :cust_first_name, :cust_last_name, :cust_email,
                    :billing_address, :billing_city, :billing_state, :billing_zip,
                    :broker_code, :broker_name, :plan_group, :plan_id, :other_charge,
                    :contract_rate
                )
            """),
            {
                "premise_id":          esi_id,
                "cust_id":             cust_id,
                "batch_no":            str(batch_no),
                "contract_start_date": _parse_enrollment_date(rec.get("start_date")),
                "company_name":        rec.get("customer_name") or "",
                "cust_first_name":     rec.get("cust_first_name") or None,
                "cust_last_name":      rec.get("cust_last_name") or None,
                "cust_email":          rec.get("customer_email") or None,
                "billing_address":     rec.get("billing_address") or None,
                "billing_city":        rec.get("billing_city") or None,
                "billing_state":       rec.get("billing_state") or None,
                "billing_zip":         rec.get("billing_zip") or None,
                "broker_code":         rec.get("broker_code") or None,
                "broker_name":         rec.get("broker_name") or None,
                "plan_group":          rec.get("plan_group") or "C1",
                "plan_id":             rec.get("plan_id") or None,
                "other_charge":        str(float(_clean_fee(rec.get("meter_fees")))),
                "contract_rate":       contract_rate,
            },
        )

    await db.execute(
        text(f"UPDATE confirmation_log SET enroll_check = 1 WHERE sid IN ({sid_list})")
    )
    await db.execute(
        text("""
            INSERT INTO enrollment_batches (batch_no, generated_by, record_count, status)
            VALUES (:batch_no, :generated_by, :record_count, 'internal')
        """),
        {
            "batch_no":      str(batch_no),
            "generated_by":  payload.get("username") or payload.get("email") or "unknown",
            "record_count":  inserted_count,
        },
    )
    await db.commit()

    return {"batch_no": batch_no, "inserted": inserted_count, "skipped": skipped}
